#!/usr/bin/env python3
"""
Accufund Formatter Script
Formats Accufund Excel reports using pandas and openpyxl
"""

import argparse
import random
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def compute_fiscal_year(date_obj):
    """Compute fiscal year based on date (Oct-Sept logic)"""
    if pd.isna(date_obj):
        return ""
    
    try:
        if isinstance(date_obj, str):
            date_obj = pd.to_datetime(date_obj)
        
        year = date_obj.year
        month = date_obj.month
        
        if month > 9:
            return f"{year}-{year + 1}"
        else:
            return f"{year - 1}-{year}"
    except:
        return ""


def categorize_type(gl_code):
    """Map GL Code first digit to Type category"""
    if pd.isna(gl_code) or str(gl_code).strip() == "":
        return ""
    
    first_char = str(gl_code)[0]
    
    mapping = {
        "1": "Asset",
        "2": "Liabilities",
        "3": "Equities",
        "4": "Revenue",
        "5": "Expenditure",
        "6": "Expenditure",
        "7": "Expenditure",
        "8": "Expenditure",
        "9": "Expenditure"
    }
    
    return mapping.get(first_char, "")


def auto_fit_columns(ws, max_rows=1000):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for i, cell in enumerate(column):
            if i >= max_rows:
                break
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 10)


def apply_header_style(ws, header_row=2):
    """Apply styling to header row with random light color"""
    random.seed()
    light_colors = ["CCFFCC", "FFCCCC", "CCCCFF", "FFFFCC", "FFCCFF", "CCFFFF"]
    fill_color = random.choice(light_colors)
    
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align


def apply_borders(ws, data_min_row=2, data_max_row=None, data_min_col=1, data_max_col=None):
    """Apply borders to data region"""
    if data_max_row is None:
        data_max_row = ws.max_row
    if data_max_col is None:
        data_max_col = ws.max_column
    
    medium_side = Side(style='medium')
    thin_side = Side(style='thin')
    
    for row in range(data_min_row, data_max_row + 1):
        for col in range(data_min_col, data_max_col + 1):
            cell = ws.cell(row=row, column=col)
            
            top = medium_side if row == data_min_row else thin_side
            bottom = medium_side if row == data_max_row else thin_side
            left = medium_side if col == data_min_col else thin_side
            right = medium_side if col == data_max_col else thin_side
            
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def set_number_formats(ws, cols, data_min_row=3):
    """Apply accounting number format to specified columns"""
    accounting_format = '#,##0.00_);(#,##0.00);"-"'
    
    for row in range(data_min_row, ws.max_row + 1):
        for col_idx in cols:
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = accounting_format


def process_excel_file(file_path):
    """Main processing function"""
    print(f"Processing file: {file_path}")
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        df = pd.read_excel(file_path, sheet_name=0, header=None)
        
        initial_rows = len(df)
        print(f"  Loaded {initial_rows} rows from file")
    except Exception as e:
        print(f"Error loading file: {e}")
        raise
    
    if len(df) == 0:
        print("Error: Empty worksheet")
        return
    
    print("  Step 1: Processing headers...")
    # Skip row 0 (title row)
    df = df.iloc[1:].reset_index(drop=True)
    
    # Check if row 0 (originally row 1) is all NaN - if so, skip it too
    if df.iloc[0].isna().all():
        print("  Skipping blank row 1...")
        df = df.iloc[1:].reset_index(drop=True)
    
    # Now set the column headers from the first row
    df.columns = df.iloc[0].values
    df = df.iloc[1:].reset_index(drop=True)
    
    print("  Step 2: Filtering Total rows...")
    account_col = df.columns[0]
    df = df[~df.iloc[:, 0].astype(str).str.contains("Total", case=False, na=False)]
    rows_removed = initial_rows - len(df) - 2
    print(f"  Removed {rows_removed} rows containing 'Total', {len(df)} rows remaining")
    
    print("  Step 3: Splitting account codes...")
    account_parts = df.iloc[:, 0].astype(str).str.split('-', expand=True)
    
    if account_parts.shape[1] >= 4:
        df.insert(1, 'Fund', account_parts[0])
        df.insert(2, 'Dept', account_parts[1])
        df.insert(3, 'MEPA', account_parts[2])
        df.insert(4, 'GL Code', account_parts[3])
    else:
        df.insert(1, 'Fund', '')
        df.insert(2, 'Dept', '')
        df.insert(3, 'MEPA', '')
        df.insert(4, 'GL Code', '')
        
        print(f"  Warning: Account codes don't all have 4 parts, processing {len(df)} rows...")
        accounts_str = df.iloc[:, 0].astype(str)
        for idx in range(len(df)):
            parts = accounts_str.iloc[idx].split('-')
            if len(parts) >= 4:
                df.at[idx, 'Fund'] = parts[0]
                df.at[idx, 'Dept'] = parts[1]
                df.at[idx, 'MEPA'] = parts[2]
                df.at[idx, 'GL Code'] = parts[3]
    
    print("  Step 4: Processing columns...")
    if 'Demo Desc' in df.columns:
        df = df.drop(columns=['Demo Desc'])
    
    date_col = None
    for col in df.columns:
        if str(col).strip() == 'Date':
            date_col = col
            break
    
    print("  Step 5: Categorizing types...")
    df.insert(5, 'Type', df['GL Code'].apply(categorize_type))
    
    print("  Step 6: Computing fiscal years...")
    if date_col:
        df['_temp_date'] = pd.to_datetime(df[date_col], errors='coerce')
        df.insert(6, 'F.Y.', df['_temp_date'].apply(compute_fiscal_year))
        df = df.drop(columns=['_temp_date'])
    else:
        df.insert(6, 'F.Y.', '')
    
    print("  Step 7: Calculating balances...")
    debit_col_idx = None
    credit_col_idx = None
    
    for i, col in enumerate(df.columns):
        if str(col).strip().lower() in ['debit', 'dr', 'debits']:
            debit_col_idx = i
        elif str(col).strip().lower() in ['credit', 'cr', 'credits']:
            credit_col_idx = i
    
    balance_col_idx = None
    
    if debit_col_idx is not None and credit_col_idx is not None:
        debit_col = df.columns[debit_col_idx]
        credit_col = df.columns[credit_col_idx]
        
        df[debit_col] = pd.to_numeric(df[debit_col], errors='coerce')
        df[debit_col] = df[debit_col].fillna(0)
        df[credit_col] = pd.to_numeric(df[credit_col], errors='coerce')
        df[credit_col] = df[credit_col].fillna(0)
        
        if 'Balance' in df.columns:
            print("  Balance column already exists, recalculating all values...")
            balance_col_idx = df.columns.get_loc('Balance')
            df['Balance'] = df[debit_col] - df[credit_col]
        else:
            balance_col_idx = credit_col_idx + 1
            df.insert(balance_col_idx, 'Balance', df[debit_col] - df[credit_col])
    
    print("  Step 8: Creating formatted worksheet...")
    if '#Formatted' in wb.sheetnames:
        del wb['#Formatted']
    
    formatted_ws = wb.create_sheet('#Formatted')
    
    print(f"  Step 9: Writing {len(df)} rows to Excel (this may take a while for large files)...")
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            formatted_ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx % 1000 == 0:
            print(f"    Written {r_idx - 1} rows...")
    
    for c_idx, col_name in enumerate(df.columns, start=1):
        formatted_ws.cell(row=1, column=c_idx, value=col_name)
    
    print("  Step 10: Auto-fitting columns...")
    auto_fit_columns(formatted_ws)
    
    print("  Step 11: Applying header styles...")
    apply_header_style(formatted_ws, header_row=1)
    
    if len(df) <= 5000:
        print("  Step 12: Applying fonts and alignment...")
        default_font = Font(name='Calibri', size=11)
        left_align = Alignment(horizontal='left', vertical='center')
        
        for row in formatted_ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = default_font
                if cell.column < formatted_ws.max_column - 2:
                    cell.alignment = left_align
        
        print("  Step 13: Applying borders...")
        apply_borders(formatted_ws, data_min_row=1, data_max_row=formatted_ws.max_row, 
                      data_min_col=1, data_max_col=formatted_ws.max_column)
    else:
        print(f"  Step 12: Skipping cell-by-cell styling for large file ({len(df)} rows)")
        print("  Note: Header styling and number formats are still applied")
    
    if debit_col_idx is not None and credit_col_idx is not None and balance_col_idx is not None:
        debit_excel_col = debit_col_idx + 1
        credit_excel_col = credit_col_idx + 1
        balance_excel_col = balance_col_idx + 1
        
        set_number_formats(formatted_ws, [debit_excel_col, credit_excel_col, balance_excel_col], data_min_row=2)
    
    print("  Step 13: Applying date format to Date column...")
    date_col_idx = None
    for i, col in enumerate(df.columns):
        if str(col).strip() == 'Date':
            date_col_idx = i
            break
    
    if date_col_idx is not None:
        date_excel_col = date_col_idx + 1
        date_format = 'mm/dd/yyyy'
        for row in range(2, formatted_ws.max_row + 1):
            cell = formatted_ws.cell(row=row, column=date_excel_col)
            cell.number_format = date_format
    
    print("  Step 14: Adding AutoFilter...")
    formatted_ws.auto_filter.ref = f"A1:{get_column_letter(formatted_ws.max_column)}{formatted_ws.max_row}"
    
    print("  Step 15: Saving file...")
    wb.save(file_path)
    
    print(f"\n✓ Processing complete!")
    print(f"  - Rows processed: {len(df) - 1}")
    print(f"  - Rows removed (containing 'Total'): {rows_removed}")
    print(f"  - Columns added: Fund, Dept, MEPA, GL Code, Type, F.Y., Balance")
    print(f"  - Output: Sheet '#Formatted' in {file_path}")


def find_sample_file():
    """Try to find a sample Excel file in common locations"""
    possible_files = [
        'sample_input.xlsx',
        'uploads/sample_input.xlsx',
        Path('uploads') / 'sample_input.xlsx'
    ]
    
    for file_path in possible_files:
        path = Path(file_path)
        if path.exists():
            return path
    
    return None


def main():
    parser = argparse.ArgumentParser(
        description='Accufund Formatter - CLI Tool',
        epilog='Example: python main.py --file sample_input.xlsx'
    )
    parser.add_argument(
        '--file', 
        type=str, 
        default=None,
        help='Path to Excel file to process'
    )
    
    args = parser.parse_args()
    
    # Determine file path
    if args.file:
        file_path = Path(args.file)
    else:
        # Try to find a sample file
        file_path = find_sample_file()
        if file_path:
            print(f"No file specified, using found sample: {file_path}")
        else:
            file_path = Path('input.xlsx')
    
    # Check if file exists
    if not file_path.exists():
        print(f"\n❌ Error: File '{file_path}' not found")
        print("\n📝 Usage Options:")
        print("  1. Run the web application:")
        print("     python app.py")
        print("\n  2. Use the CLI tool with a specific file:")
        print("     python main.py --file path/to/your/file.xlsx")
        
        # Check if sample file exists and suggest it
        sample = find_sample_file()
        if sample:
            print(f"\n  3. Process the sample file:")
            print(f"     python main.py --file {sample}")
        
        # List available Excel files in current and uploads directory
        excel_files = list(Path('.').glob('*.xlsx')) + list(Path('uploads').glob('*.xlsx')) if Path('uploads').exists() else []
        if excel_files:
            print(f"\n📂 Available Excel files found:")
            for f in excel_files[:5]:  # Show first 5
                print(f"     - {f}")
        
        sys.exit(1)
    
    try:
        process_excel_file(file_path)
    except Exception as e:
        print(f"\n❌ Error processing file: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()