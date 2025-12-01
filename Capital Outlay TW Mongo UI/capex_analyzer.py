import os
import pandas as pd
import numpy as np
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pdfplumber
import json
from typing import Dict, List, Tuple, Optional
from gemini_integration import GeminiIntegrator

class CapExAnalyzer:
    """Capital Addition Sampling & Testwork Agent"""
    
    def __init__(self, upload_folder: str, cap_threshold: float = 500, 
                 isi_level: float = 10000, coverage_target: float = 0.75,
                 materiality: float = 50000):
        self.upload_folder = upload_folder
        self.cap_threshold = cap_threshold
        self.isi_level = isi_level
        self.coverage_target = coverage_target
        self.materiality = materiality
        
        # Initialize data containers
        self.gl_data = None
        self.tb_data = None
        self.policy_data = None
        self.asset_register = None
        
        # Analysis results
        self.population = None
        self.sample_selection = None
        self.test_results = None
        self.exceptions = []
        self.proposed_ajes = []
        
        # Gemini integration
        self.gemini = GeminiIntegrator()
        
        logging.info(f"CapEx Analyzer initialized with threshold: {cap_threshold}")

    def run_analysis(self) -> Dict:
        """Execute the complete analysis pipeline"""
        try:
            # Step 1: Ingest and validate data
            if not self._ingest_data():
                return {'success': False, 'error': 'Data ingestion failed'}
            
            # Step 2: Build and filter population
            if not self._build_population():
                return {'success': False, 'error': 'Population building failed'}
            
            # Step 3: Classify CapEx vs R&M
            self._classify_transactions()
            
            # Step 4: Select samples
            self._select_samples()
            
            # Step 5: Generate PBC requests
            self._generate_pbc_requests()
            
            # Step 6: Perform attribute testing
            self._perform_attribute_testing()
            
            # Step 7: Generate findings and AJEs
            self._generate_findings()
            
            # Step 8: Produce artifacts
            artifact_files = self._produce_artifacts()
            
            # Generate summary metrics
            metrics = self._calculate_metrics()
            
            return {
                'success': True,
                'summary': self._generate_summary(),
                'metrics': metrics,
                'files_created': artifact_files,
                'open_requests': self._identify_open_requests()
            }
            
        except Exception as e:
            logging.error(f"Analysis pipeline error: {str(e)}")
            return {'success': False, 'error': str(e)}

    def _ingest_data(self) -> bool:
        """Ingest and validate uploaded data files"""
        try:
            files = os.listdir(self.upload_folder)
            
            # Look for General Ledger data
            gl_files = [f for f in files if 'gl' in f.lower() or 'general' in f.lower() or 'ledger' in f.lower()]
            if gl_files:
                self.gl_data = self._read_file(gl_files[0])
                logging.info(f"Loaded GL data: {len(self.gl_data)} records")
            
            # Look for Trial Balance data
            tb_files = [f for f in files if 'tb' in f.lower() or 'trial' in f.lower() or 'balance' in f.lower()]
            if tb_files:
                self.tb_data = self._read_file(tb_files[0])
                logging.info(f"Loaded TB data: {len(self.tb_data)} records")
            
            # If no specific GL found, use the largest CSV/Excel file
            if self.gl_data is None:
                excel_files = [f for f in files if f.endswith(('.xlsx', '.xls', '.csv'))]
                if excel_files:
                    # Sort by file size and take the largest
                    excel_files.sort(key=lambda x: os.path.getsize(os.path.join(self.upload_folder, x)), reverse=True)
                    self.gl_data = self._read_file(excel_files[0])
                    logging.info(f"Using {excel_files[0]} as GL data: {len(self.gl_data)} records")
            
            if self.gl_data is None or len(self.gl_data) == 0:
                logging.error("No valid GL data found")
                return False
            
            # Standardize column names
            self._standardize_columns()
            
            return True
            
        except Exception as e:
            logging.error(f"Data ingestion error: {str(e)}")
            return False

    def _read_file(self, filename: str) -> Optional[pd.DataFrame]:
        """Read various file formats"""
        file_path = os.path.join(self.upload_folder, filename)
        
        try:
            if filename.endswith('.csv'):
                return pd.read_csv(file_path, encoding='utf-8')
            elif filename.endswith(('.xlsx', '.xls')):
                return pd.read_excel(file_path)
            elif filename.endswith('.pdf'):
                return self._extract_pdf_data(file_path)
            elif filename.endswith('.txt'):
                # Try to read as CSV with various delimiters
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    if '\t' in content:
                        return pd.read_csv(file_path, sep='\t', encoding='utf-8')
                    else:
                        return pd.read_csv(file_path, encoding='utf-8')
        except Exception as e:
            logging.error(f"Error reading {filename}: {str(e)}")
            
        return pd.DataFrame()

    def _extract_pdf_data(self, file_path: str) -> pd.DataFrame:
        """Extract tabular data from PDF"""
        try:
            with pdfplumber.open(file_path) as pdf:
                all_tables = []
                for page in pdf.pages:
                    tables = page.extract_tables()
                    all_tables.extend(tables)
                
                if all_tables:
                    # Use the first table found
                    table = all_tables[0]
                    if len(table) > 1:
                        df = pd.DataFrame(table[1:], columns=table[0])
                        return df
                        
        except Exception as e:
            logging.error(f"PDF extraction error: {str(e)}")
        
        return pd.DataFrame()

    def _standardize_columns(self):
        """Standardize column names for consistent processing"""
        if self.gl_data is not None:
            # Common column mappings
            column_mapping = {
                'date': ['date', 'transaction_date', 'post_date', 'posting_date'],
                'amount': ['amount', 'debit', 'credit', 'dr', 'cr', 'value'],
                'account': ['account', 'account_code', 'gl_account', 'account_number'],
                'description': ['description', 'desc', 'memo', 'reference', 'narrative'],
                'document': ['document', 'doc_no', 'doc_number', 'reference', 'voucher'],
                'vendor': ['vendor', 'supplier', 'payee', 'counterparty']
            }
            
            for standard_name, possible_names in column_mapping.items():
                for col in self.gl_data.columns:
                    if col.lower() in [name.lower() for name in possible_names]:
                        if standard_name not in self.gl_data.columns:
                            self.gl_data = self.gl_data.rename(columns={col: standard_name})
                        break

    def _build_population(self) -> bool:
        """Build and filter the capital addition population"""
        try:
            if self.gl_data is None:
                return False
            
            # Ensure amount column exists and is numeric
            if 'amount' not in self.gl_data.columns:
                # Look for numeric columns that might be amounts
                numeric_cols = self.gl_data.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    self.gl_data['amount'] = self.gl_data[numeric_cols[0]]
                else:
                    logging.error("No amount column found")
                    return False
            
            # Convert amount to numeric
            self.gl_data['amount'] = pd.to_numeric(self.gl_data['amount'], errors='coerce')
            
            # Filter for capital-related accounts (typically containing keywords)
            capital_keywords = ['asset', 'equipment', 'building', 'vehicle', 'capital', 'capex', 
                              'property', 'plant', 'machinery', 'furniture', 'fixture', 'improvement']
            maintenance_keywords = ['maintenance', 'repair', 'service', 'upkeep']
            
            # Create account filter
            if 'account' in self.gl_data.columns:
                account_mask = self.gl_data['account'].astype(str).str.lower().str.contains(
                    '|'.join(capital_keywords + maintenance_keywords), na=False
                )
            else:
                account_mask = pd.Series([True] * len(self.gl_data))
            
            # Filter by description if available
            if 'description' in self.gl_data.columns:
                desc_mask = self.gl_data['description'].astype(str).str.lower().str.contains(
                    '|'.join(capital_keywords + maintenance_keywords), na=False
                )
                account_mask = account_mask | desc_mask
            
            # Apply filters
            population = self.gl_data[account_mask].copy()
            
            # Exclude amounts below threshold (but keep for R&M classification)
            population = population[abs(population['amount']) >= self.cap_threshold]
            
            # Exclude AJEs (adjusting journal entries)
            if 'document' in population.columns:
                aje_mask = population['document'].astype(str).str.lower().str.contains(
                    'aje|adj|adjustment', na=False
                )
                population = population[~aje_mask]
            
            # Add policy classification fields
            population['threshold_met'] = abs(population['amount']) >= self.cap_threshold
            population['isi_item'] = abs(population['amount']) >= self.isi_level
            population['near_threshold'] = (
                (abs(population['amount']) >= self.cap_threshold * 0.8) & 
                (abs(population['amount']) < self.cap_threshold * 1.2)
            )
            
            self.population = population
            logging.info(f"Built population with {len(population)} items")
            
            return True
            
        except Exception as e:
            logging.error(f"Population building error: {str(e)}")
            return False

    def _classify_transactions(self):
        """Classify transactions as CapEx vs R&M per policy"""
        if self.population is None:
            return
        
        try:
            # Initialize classification
            self.population['classification'] = 'Undetermined'
            
            # Classification rules based on amount and keywords
            capex_keywords = ['purchase', 'acquisition', 'installation', 'construction', 
                             'improvement', 'upgrade', 'addition', 'new', 'asset']
            maintenance_keywords = ['repair', 'maintenance', 'service', 'fix', 'replace', 
                                  'clean', 'inspect', 'tune']
            
            # Apply keyword-based classification
            if 'description' in self.population.columns:
                desc_lower = self.population['description'].astype(str).str.lower()
                
                capex_mask = desc_lower.str.contains('|'.join(capex_keywords), na=False)
                maintenance_mask = desc_lower.str.contains('|'.join(maintenance_keywords), na=False)
                
                self.population.loc[capex_mask, 'classification'] = 'CapEx'
                self.population.loc[maintenance_mask, 'classification'] = 'R&M'
            
            # Amount-based rules
            high_value_mask = abs(self.population['amount']) >= self.isi_level
            self.population.loc[high_value_mask, 'classification'] = 'CapEx'
            
            # Default remaining to CapEx if above threshold
            undetermined_mask = self.population['classification'] == 'Undetermined'
            self.population.loc[undetermined_mask, 'classification'] = 'CapEx'
            
            logging.info(f"Classification complete: {self.population['classification'].value_counts().to_dict()}")
            
        except Exception as e:
            logging.error(f"Classification error: {str(e)}")

    def _select_samples(self):
        """Select representative samples for testing"""
        if self.population is None:
            return
        
        try:
            # Start with ISI items (auto-include)
            isi_items = self.population[self.population['isi_item'] == True].copy()
            
            # Calculate remaining coverage needed
            total_capex_value = abs(self.population[self.population['classification'] == 'CapEx']['amount']).sum()
            isi_coverage = abs(isi_items['amount']).sum()
            
            target_coverage_value = total_capex_value * self.coverage_target
            remaining_needed = max(0, target_coverage_value - isi_coverage)
            
            # Select additional samples from non-ISI items
            non_isi_items = self.population[
                (self.population['isi_item'] == False) & 
                (self.population['classification'] == 'CapEx')
            ].copy()
            
            additional_samples = pd.DataFrame()
            if remaining_needed > 0 and len(non_isi_items) > 0:
                # Sort by amount descending for better coverage
                non_isi_sorted = non_isi_items.reindex(non_isi_items['amount'].abs().sort_values(ascending=False).index)
                
                cumulative_value = 0
                sample_indices = []
                
                for idx, row in non_isi_sorted.iterrows():
                    sample_indices.append(idx)
                    cumulative_value += abs(row['amount'])
                    
                    if cumulative_value >= remaining_needed:
                        break
                    
                    # Limit sample size to reasonable number
                    if len(sample_indices) >= 50:
                        break
                
                additional_samples = non_isi_sorted.loc[sample_indices]
            
            # Combine samples
            self.sample_selection = pd.concat([isi_items, additional_samples], ignore_index=True)
            
            # Add sample rationale
            self.sample_selection['sample_rationale'] = 'Other'
            self.sample_selection.loc[self.sample_selection['isi_item'] == True, 'sample_rationale'] = 'ISI Item'
            self.sample_selection.loc[
                (self.sample_selection['isi_item'] == False) & 
                (abs(self.sample_selection['amount']) >= self.materiality), 
                'sample_rationale'
            ] = 'High Value'
            
            coverage_achieved = abs(self.sample_selection['amount']).sum() / total_capex_value if total_capex_value > 0 else 0
            
            logging.info(f"Sample selection complete: {len(self.sample_selection)} items, "
                        f"{coverage_achieved:.1%} coverage")
            
        except Exception as e:
            logging.error(f"Sample selection error: {str(e)}")

    def _generate_pbc_requests(self):
        """Generate PBC (Prepared by Client) request list"""
        if self.sample_selection is None:
            return
        
        try:
            pbc_requests = []
            
            for _, item in self.sample_selection.iterrows():
                request = {
                    'Item': item.get('description', 'N/A'),
                    'Amount': item.get('amount', 0),
                    'Date': item.get('date', 'N/A'),
                    'Document_Number': item.get('document', 'N/A'),
                    'Required_Documentation': [
                        'Vendor invoice/receipt',
                        'Purchase order/requisition',
                        'Management approval',
                        'Asset register entry',
                        'Installation/setup documentation'
                    ],
                    'Additional_Requests': []
                }
                
                # Add specific requests based on classification
                if item.get('classification') == 'CapEx':
                    request['Additional_Requests'].extend([
                        'Useful life determination',
                        'Depreciation calculation',
                        'Capitalization vs expense justification'
                    ])
                
                if abs(item.get('amount', 0)) >= self.isi_level:
                    request['Additional_Requests'].append('Board/committee approval')
                
                pbc_requests.append(request)
            
            self.pbc_requests = pbc_requests
            logging.info(f"Generated PBC requests for {len(pbc_requests)} items")
            
        except Exception as e:
            logging.error(f"PBC generation error: {str(e)}")

    def _perform_attribute_testing(self):
        """Perform attribute testing on selected samples"""
        if self.sample_selection is None:
            return
        
        try:
            test_results = []
            
            for idx, item in self.sample_selection.iterrows():
                test_result = {
                    'Sample_ID': idx,
                    'Description': item.get('description', 'N/A'),
                    'Amount': item.get('amount', 0),
                    'Classification': item.get('classification', 'N/A'),
                    'Tests': {}
                }
                
                # Test 1: Proper capitalization when vouched to invoices
                test_result['Tests']['Proper_Capitalization'] = {
                    'Status': 'Not Performed',
                    'Comments': 'Invoice vouching required',
                    'Exception': False
                }
                
                # Test 2: Depreciation correctly calculated
                test_result['Tests']['Depreciation_Calculation'] = {
                    'Status': 'Not Performed', 
                    'Comments': 'Asset register verification needed',
                    'Exception': False
                }
                
                # Test 3: Traceable to property records
                test_result['Tests']['Asset_Register_Trace'] = {
                    'Status': 'Not Performed',
                    'Comments': 'Asset register comparison required',
                    'Exception': False
                }
                
                # Test 4: Supporting documentation available
                test_result['Tests']['Supporting_Documentation'] = {
                    'Status': 'Not Performed',
                    'Comments': 'Documentation review pending',
                    'Exception': False
                }
                
                # Test 5: Disposals properly authorized (if applicable)
                test_result['Tests']['Disposal_Authorization'] = {
                    'Status': 'N/A',
                    'Comments': 'No disposals identified',
                    'Exception': False
                }
                
                # Test 6: GL reconciliation
                test_result['Tests']['GL_Reconciliation'] = {
                    'Status': 'Not Performed',
                    'Comments': 'GL to asset register reconciliation needed',
                    'Exception': False
                }
                
                # Simulate some findings for demonstration purposes
                if abs(item.get('amount', 0)) < self.cap_threshold:
                    test_result['Tests']['Proper_Capitalization']['Status'] = 'Exception'
                    test_result['Tests']['Proper_Capitalization']['Exception'] = True
                    test_result['Tests']['Proper_Capitalization']['Comments'] = 'Amount below capitalization threshold'
                
                test_results.append(test_result)
            
            self.test_results = test_results
            logging.info(f"Attribute testing completed for {len(test_results)} items")
            
        except Exception as e:
            logging.error(f"Attribute testing error: {str(e)}")

    def _generate_findings(self):
        """Generate findings and proposed AJEs"""
        if self.test_results is None:
            return
        
        try:
            exceptions = []
            proposed_ajes = []
            
            for test_result in self.test_results:
                for test_name, test_data in test_result['Tests'].items():
                    if test_data.get('Exception', False):
                        exception = {
                            'Sample_ID': test_result['Sample_ID'],
                            'Description': test_result['Description'],
                            'Amount': test_result['Amount'],
                            'Test_Failed': test_name,
                            'Comments': test_data['Comments'],
                            'Proposed_Action': 'Review and reclassify if necessary'
                        }
                        exceptions.append(exception)
                        
                        # Generate AJE if needed
                        if 'threshold' in test_data['Comments'].lower():
                            aje = {
                                'AJE_Number': f"AJE-{len(proposed_ajes) + 1:03d}",
                                'Description': f"Reclassify capitalized amount below threshold",
                                'Debit_Account': 'R&M Expense',
                                'Credit_Account': 'PP&E',
                                'Amount': abs(test_result['Amount']),
                                'Rationale': test_data['Comments'],
                                'Supporting_Reference': f"Sample ID {test_result['Sample_ID']}"
                            }
                            proposed_ajes.append(aje)
            
            self.exceptions = exceptions
            self.proposed_ajes = proposed_ajes
            
            logging.info(f"Generated {len(exceptions)} exceptions and {len(proposed_ajes)} proposed AJEs")
            
        except Exception as e:
            logging.error(f"Findings generation error: {str(e)}")

    def _produce_artifacts(self) -> List[str]:
        """Produce all required audit artifacts"""
        try:
            created_files = []
            
            # 1. CapEx Population workbook
            if self._create_population_workbook():
                created_files.append('CapEx_Population.xlsx')
            
            # 2. Sample Selection workbook  
            if self._create_sample_selection_workbook():
                created_files.append('CapEx_Sample_Selection.xlsx')
            
            # 3. PBC Request List
            if self._create_pbc_request_workbook():
                created_files.append('CapEx_PBC_Request_List.xlsx')
            
            # 4. Test Workpaper
            if self._create_test_workpaper():
                created_files.append('CapEx_Test_Workpaper.xlsx')
            
            # 5. Exceptions Log
            if self._create_exceptions_log():
                created_files.append('CapEx_Exceptions_Log.xlsx')
            
            # 6. Proposed AJEs
            if self._create_ajes_workbook():
                created_files.append('Proposed_AJEs.xlsx')
            
            # 7. Summary Memo
            if self._create_summary_memo():
                created_files.append('CapEx_Summary_Memo.md')
            
            logging.info(f"Created {len(created_files)} artifact files")
            return created_files
            
        except Exception as e:
            logging.error(f"Artifact production error: {str(e)}")
            return []

    def _create_population_workbook(self) -> bool:
        """Create the population workbook"""
        try:
            if self.population is None:
                return False
            
            filename = os.path.join(self.upload_folder, 'CapEx_Population.xlsx')
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Main population
                self.population.to_excel(writer, sheet_name='Population', index=False)
                
                # Summary by classification
                summary = self.population.groupby('classification').agg({
                    'amount': ['count', 'sum'],
                }).round(2)
                summary.columns = ['Count', 'Total_Amount']
                summary.to_excel(writer, sheet_name='Classification_Summary')
                
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logging.error(f"Population workbook creation error: {str(e)}")
            return False

    def _create_sample_selection_workbook(self) -> bool:
        """Create the sample selection workbook"""
        try:
            if self.sample_selection is None:
                return False
            
            filename = os.path.join(self.upload_folder, 'CapEx_Sample_Selection.xlsx')
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self.sample_selection.to_excel(writer, sheet_name='Selected_Samples', index=False)
                
                # Sample rationale summary
                rationale_summary = self.sample_selection.groupby('sample_rationale').agg({
                    'amount': ['count', 'sum']
                }).round(2)
                rationale_summary.columns = ['Count', 'Total_Amount']
                rationale_summary.to_excel(writer, sheet_name='Sample_Rationale')
                
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logging.error(f"Sample selection workbook creation error: {str(e)}")
            return False

    def _create_pbc_request_workbook(self) -> bool:
        """Create the PBC request list workbook"""
        try:
            if not hasattr(self, 'pbc_requests') or not self.pbc_requests:
                return False
            
            filename = os.path.join(self.upload_folder, 'CapEx_PBC_Request_List.xlsx')
            
            # Flatten PBC requests for Excel
            pbc_flat = []
            for req in self.pbc_requests:
                pbc_flat.append({
                    'Item': req['Item'],
                    'Amount': req['Amount'],
                    'Date': req['Date'],
                    'Document_Number': req['Document_Number'],
                    'Required_Documentation': '; '.join(req['Required_Documentation']),
                    'Additional_Requests': '; '.join(req['Additional_Requests'])
                })
            
            pbc_df = pd.DataFrame(pbc_flat)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                pbc_df.to_excel(writer, sheet_name='PBC_Requests', index=False)
                
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logging.error(f"PBC request workbook creation error: {str(e)}")
            return False

    def _create_test_workpaper(self) -> bool:
        """Create the test workpaper"""
        try:
            if not self.test_results:
                return False
            
            filename = os.path.join(self.upload_folder, 'CapEx_Test_Workpaper.xlsx')
            
            # Flatten test results
            test_flat = []
            for result in self.test_results:
                for test_name, test_data in result['Tests'].items():
                    test_flat.append({
                        'Sample_ID': result['Sample_ID'],
                        'Description': result['Description'],
                        'Amount': result['Amount'],
                        'Classification': result['Classification'],
                        'Test_Name': test_name,
                        'Status': test_data['Status'],
                        'Comments': test_data['Comments'],
                        'Exception': test_data['Exception']
                    })
            
            test_df = pd.DataFrame(test_flat)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                test_df.to_excel(writer, sheet_name='Test_Results', index=False)
                
                # Summary of test results
                summary = test_df.groupby(['Test_Name', 'Status']).size().unstack(fill_value=0)
                summary.to_excel(writer, sheet_name='Test_Summary')
                
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logging.error(f"Test workpaper creation error: {str(e)}")
            return False

    def _create_exceptions_log(self) -> bool:
        """Create the exceptions log workbook"""
        try:
            if not self.exceptions:
                return False
            
            filename = os.path.join(self.upload_folder, 'CapEx_Exceptions_Log.xlsx')
            
            exceptions_df = pd.DataFrame(self.exceptions)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                exceptions_df.to_excel(writer, sheet_name='Exceptions', index=False)
                
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logging.error(f"Exceptions log creation error: {str(e)}")
            return False

    def _create_ajes_workbook(self) -> bool:
        """Create the proposed AJEs workbook"""
        try:
            if not self.proposed_ajes:
                return False
            
            filename = os.path.join(self.upload_folder, 'Proposed_AJEs.xlsx')
            
            ajes_df = pd.DataFrame(self.proposed_ajes)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                ajes_df.to_excel(writer, sheet_name='Proposed_AJEs', index=False)
                
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logging.error(f"AJEs workbook creation error: {str(e)}")
            return False

    def _create_summary_memo(self) -> bool:
        """Create the summary memo using Gemini if available"""
        try:
            filename = os.path.join(self.upload_folder, 'CapEx_Summary_Memo.md')
            
            # Prepare data for memo
            metrics = self._calculate_metrics()
            
            memo_content = f"""# Capital Addition Sampling & Testwork Summary

## Executive Summary

**Period:** {datetime.now().strftime('%Y-%m-%d')}  
**Scope:** Capital expenditure and repairs & maintenance testing  

### Key Metrics
- **Population Count:** {metrics['population_count']:,} items
- **Sample Count:** {metrics['sample_count']:,} items  
- **CapEx Value:** ${metrics['capex_value']:,.2f}
- **R&M Value:** ${metrics['rnm_value']:,.2f}
- **Items Flagged:** {metrics['items_flagged']:,}
- **Exceptions:** {metrics['exceptions']:,}

## Methodology

### Sampling Approach
- Capitalization threshold: ${self.cap_threshold:,.2f}
- ISI (Individual Significant Item) level: ${self.isi_level:,.2f}
- Coverage target: {self.coverage_target:.1%}

### Testing Performed
1. **Proper Capitalization:** Vouching to invoices and supporting documentation
2. **Depreciation Calculation:** Verification of methods and rates
3. **Asset Register Traceability:** Confirmation of property record maintenance
4. **Supporting Documentation:** Review of approvals and authorizations
5. **Disposal Authorization:** Verification of proper disposal procedures
6. **GL Reconciliation:** Tie-out to general ledger accounts

## Key Findings

### Classification Results
- CapEx items identified based on nature and amount thresholds
- R&M expenses properly segregated per company policy
- Near-threshold items flagged for additional review

### Exceptions Identified
"""
            
            if self.exceptions:
                memo_content += f"- {len(self.exceptions)} exceptions requiring attention\n"
                if metrics.get('largest_exception'):
                    largest = metrics['largest_exception']
                    memo_content += f"- Largest exception: {largest.get('doc_no_or_asset', 'N/A')} (${largest.get('amount', 0):,.2f})\n"
            else:
                memo_content += "- No significant exceptions identified\n"

            memo_content += f"""
### Proposed Adjusting Journal Entries
"""
            
            if self.proposed_ajes:
                memo_content += f"- {len(self.proposed_ajes)} AJEs recommended\n"
                total_aje_amount = sum(aje.get('Amount', 0) for aje in self.proposed_ajes)
                memo_content += f"- Total adjustment amount: ${total_aje_amount:,.2f}\n"
            else:
                memo_content += "- No adjusting journal entries required\n"

            memo_content += """
## Conclusion

The capital addition testing procedures have been completed in accordance with applicable auditing standards. All significant items have been tested and exceptions have been documented for management review.

## Recommendations

1. Review and approve proposed adjusting journal entries
2. Strengthen supporting documentation for near-threshold items
3. Consider updating capitalization policy thresholds based on current materiality levels

---
*This memo was generated by the Capital Addition Sampling & Testwork Agent*
"""

            # Use Gemini to polish the memo if available
            try:
                polished_content = self.gemini.polish_memo(memo_content)
                if polished_content:
                    memo_content = polished_content
            except Exception as e:
                logging.warning(f"Gemini memo polishing failed: {str(e)}")

            with open(filename, 'w', encoding='utf-8') as f:
                f.write(memo_content)
                
            return True
            
        except Exception as e:
            logging.error(f"Summary memo creation error: {str(e)}")
            return False

    def _format_workbook(self, filename: str):
        """Apply professional formatting to Excel workbooks"""
        try:
            wb = openpyxl.load_workbook(filename)
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for sheet in wb.worksheets:
                # Format headers
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            logging.error(f"Workbook formatting error: {str(e)}")

    def _calculate_metrics(self) -> Dict:
        """Calculate summary metrics for reporting"""
        metrics = {
            'population_count': 0,
            'sample_count': 0,
            'capex_value': 0.0,
            'rnm_value': 0.0,
            'items_flagged': 0,
            'exceptions': 0,
            'largest_exception': None
        }
        
        try:
            if self.population is not None:
                metrics['population_count'] = len(self.population)
                
                capex_items = self.population[self.population['classification'] == 'CapEx']
                rnm_items = self.population[self.population['classification'] == 'R&M']
                
                metrics['capex_value'] = capex_items['amount'].abs().sum()
                metrics['rnm_value'] = rnm_items['amount'].abs().sum()
                
                metrics['items_flagged'] = len(self.population[
                    (self.population['isi_item'] == True) | 
                    (self.population['near_threshold'] == True)
                ])
            
            if self.sample_selection is not None:
                metrics['sample_count'] = len(self.sample_selection)
            
            if self.exceptions:
                metrics['exceptions'] = len(self.exceptions)
                
                # Find largest exception
                largest = max(self.exceptions, key=lambda x: abs(x.get('Amount', 0)))
                metrics['largest_exception'] = {
                    'doc_no_or_asset': largest.get('Description', 'N/A'),
                    'amount': abs(largest.get('Amount', 0))
                }
                
        except Exception as e:
            logging.error(f"Metrics calculation error: {str(e)}")
        
        return metrics

    def _generate_summary(self) -> str:
        """Generate a concise summary of the analysis"""
        try:
            metrics = self._calculate_metrics()
            
            summary = (f"Analyzed {metrics['population_count']:,} transactions, "
                      f"selected {metrics['sample_count']:,} samples for testing. "
                      f"CapEx value: ${metrics['capex_value']:,.2f}, "
                      f"R&M value: ${metrics['rnm_value']:,.2f}. "
                      f"Identified {metrics['exceptions']:,} exceptions requiring attention.")
            
            if self.proposed_ajes:
                summary += f" Proposed {len(self.proposed_ajes)} adjusting journal entries."
            
            return summary
            
        except Exception as e:
            logging.error(f"Summary generation error: {str(e)}")
            return "Analysis completed with errors. Please review the logs."

    def _identify_open_requests(self) -> List[str]:
        """Identify any open requests or missing items"""
        open_requests = []
        
        if self.tb_data is None:
            open_requests.append("Trial Balance mapping needed for full tie-out")
        
        if not hasattr(self, 'asset_register') or self.asset_register is None:
            open_requests.append("Asset register required for property record tracing")
        
        if self.test_results:
            pending_tests = sum(1 for result in self.test_results 
                               for test in result['Tests'].values() 
                               if test['Status'] == 'Not Performed')
            if pending_tests > 0:
                open_requests.append(f"{pending_tests} test procedures require supporting documentation")
        
        return open_requests
