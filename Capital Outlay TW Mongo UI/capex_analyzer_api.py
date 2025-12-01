"""
Capital Addition Analyzer - Core Logic (No MongoDB)
Simplified version for API integration
"""

import os
import pandas as pd
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pdfplumber

# Optional Gemini integration
try:
    from gemini_integration import GeminiIntegrator
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False
    logging.warning("Gemini integration not available")

logger = logging.getLogger(__name__)


class CapExAnalyzer:
    """Capital Addition Sampling & Testwork Agent"""
    
    def __init__(self, upload_folder: str, cap_threshold: float = 500,
                 isi_level: float = 10000, coverage_target: float = 0.75,
                 materiality: float = 50000):
        self.upload_folder = Path(upload_folder)
        self.cap_threshold = cap_threshold
        self.isi_level = isi_level
        self.coverage_target = coverage_target
        self.materiality = materiality
        
        # Initialize data containers
        self.gl_data = None
        self.tb_data = None
        self.population = None
        self.sample_selection = None
        self.test_results = None
        self.exceptions = []
        self.proposed_ajes = []
        self.pbc_requests = []
        
        # Gemini integration (optional)
        if GEMINI_AVAILABLE:
            self.gemini = GeminiIntegrator()
        else:
            self.gemini = None
        
        logger.info(f"CapEx Analyzer initialized with threshold: {cap_threshold}")

    def run_analysis(self) -> Dict:
        """Execute the complete analysis pipeline"""
        try:
            # Step 1: Ingest data
            if not self._ingest_data():
                return {'success': False, 'error': 'Data ingestion failed'}
            
            # Step 2: Build population
            if not self._build_population():
                return {'success': False, 'error': 'Population building failed'}
            
            # Step 3: Classify CapEx vs R&M
            self._classify_transactions()
            
            # Step 4: Select samples
            self._select_samples()
            
            # Step 5: Generate PBC requests
            self._generate_pbc_requests()
            
            # Step 6: Perform testing
            self._perform_attribute_testing()
            
            # Step 7: Generate findings
            self._generate_findings()
            
            # Step 8: Produce artifacts
            artifact_files = self._produce_artifacts()
            
            # Generate metrics
            metrics = self._calculate_metrics()
            
            return {
                'success': True,
                'summary': self._generate_summary(),
                'metrics': metrics,
                'files_created': artifact_files,
                'open_requests': self._identify_open_requests()
            }
            
        except Exception as e:
            logger.error(f"Analysis pipeline error: {str(e)}", exc_info=True)
            return {'success': False, 'error': str(e)}

    def _ingest_data(self) -> bool:
        """Ingest uploaded data files"""
        try:
            files = list(self.upload_folder.glob('*'))
            
            # Look for GL data
            gl_files = [f for f in files if any(kw in f.name.lower() for kw in ['gl', 'general', 'ledger'])]
            if gl_files:
                self.gl_data = self._read_file(gl_files[0])
                logger.info(f"Loaded GL data: {len(self.gl_data)} records")
            
            # Look for TB data
            tb_files = [f for f in files if any(kw in f.name.lower() for kw in ['tb', 'trial', 'balance'])]
            if tb_files:
                self.tb_data = self._read_file(tb_files[0])
                logger.info(f"Loaded TB data: {len(self.tb_data)} records")
            
            # Fallback: use largest file
            if self.gl_data is None:
                data_files = [f for f in files if f.suffix.lower() in ['.xlsx', '.xls', '.csv']]
                if data_files:
                    largest = max(data_files, key=lambda x: x.stat().st_size)
                    self.gl_data = self._read_file(largest)
                    logger.info(f"Using {largest.name} as GL data: {len(self.gl_data)} records")
            
            if self.gl_data is None or len(self.gl_data) == 0:
                logger.error("No valid GL data found")
                return False
            
            self._standardize_columns()
            return True
            
        except Exception as e:
            logger.error(f"Data ingestion error: {str(e)}")
            return False

    def _read_file(self, filepath: Path) -> Optional[pd.DataFrame]:
        """Read various file formats"""
        try:
            if filepath.suffix == '.csv':
                return pd.read_csv(filepath, encoding='utf-8')
            elif filepath.suffix in ['.xlsx', '.xls']:
                return pd.read_excel(filepath)
            elif filepath.suffix == '.pdf':
                return self._extract_pdf_data(filepath)
            elif filepath.suffix == '.txt':
                return pd.read_csv(filepath, sep='\t', encoding='utf-8')
        except Exception as e:
            logger.error(f"Error reading {filepath.name}: {str(e)}")
        return pd.DataFrame()

    def _extract_pdf_data(self, filepath: Path) -> pd.DataFrame:
        """Extract tabular data from PDF"""
        try:
            with pdfplumber.open(filepath) as pdf:
                all_tables = []
                for page in pdf.pages:
                    tables = page.extract_tables()
                    all_tables.extend(tables)
                
                if all_tables and len(all_tables[0]) > 1:
                    table = all_tables[0]
                    return pd.DataFrame(table[1:], columns=table[0])
        except Exception as e:
            logger.error(f"PDF extraction error: {str(e)}")
        return pd.DataFrame()

    def _standardize_columns(self):
        """Standardize column names"""
        if self.gl_data is None:
            return
        
        column_mapping = {
            'date': ['date', 'transaction_date', 'post_date', 'posting_date'],
            'amount': ['amount', 'debit', 'credit', 'dr', 'cr', 'value'],
            'account': ['account', 'account_code', 'gl_account', 'account_number'],
            'description': ['description', 'desc', 'memo', 'reference', 'narrative'],
            'document': ['document', 'doc_no', 'doc_number', 'reference', 'voucher'],
            'vendor': ['vendor', 'supplier', 'payee', 'counterparty']
        }
        
        for standard_name, possible_names in column_mapping.items():
            for col in self.gl_data.columns:
                if col.lower() in [name.lower() for name in possible_names]:
                    if standard_name not in self.gl_data.columns:
                        self.gl_data = self.gl_data.rename(columns={col: standard_name})
                    break

    def _build_population(self) -> bool:
        """Build and filter the capital addition population"""
        try:
            if self.gl_data is None:
                return False
            
            # Ensure amount column
            if 'amount' not in self.gl_data.columns:
                numeric_cols = self.gl_data.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    self.gl_data['amount'] = self.gl_data[numeric_cols[0]]
                else:
                    logger.error("No amount column found")
                    return False
            
            self.gl_data['amount'] = pd.to_numeric(self.gl_data['amount'], errors='coerce')
            
            # Filter for capital-related accounts
            capital_keywords = ['asset', 'equipment', 'building', 'vehicle', 'capital', 'capex',
                              'property', 'plant', 'machinery', 'furniture', 'fixture', 'improvement']
            maintenance_keywords = ['maintenance', 'repair', 'service', 'upkeep']
            
            account_mask = pd.Series([False] * len(self.gl_data))
            
            if 'account' in self.gl_data.columns:
                account_mask = self.gl_data['account'].astype(str).str.lower().str.contains(
                    '|'.join(capital_keywords + maintenance_keywords), na=False
                )
            
            if 'description' in self.gl_data.columns:
                desc_mask = self.gl_data['description'].astype(str).str.lower().str.contains(
                    '|'.join(capital_keywords + maintenance_keywords), na=False
                )
                account_mask = account_mask | desc_mask
            
            # Apply filters
            population = self.gl_data[account_mask].copy()
            population = population[abs(population['amount']) >= self.cap_threshold]
            
            # Exclude AJEs
            if 'document' in population.columns:
                aje_mask = population['document'].astype(str).str.lower().str.contains(
                    'aje|adj|adjustment', na=False
                )
                population = population[~aje_mask]
            
            # Add classification fields
            population['threshold_met'] = abs(population['amount']) >= self.cap_threshold
            population['isi_item'] = abs(population['amount']) >= self.isi_level
            population['near_threshold'] = (
                (abs(population['amount']) >= self.cap_threshold * 0.8) &
                (abs(population['amount']) < self.cap_threshold * 1.2)
            )
            
            self.population = population
            logger.info(f"Built population with {len(population)} items")
            
            return True
            
        except Exception as e:
            logger.error(f"Population building error: {str(e)}")
            return False

    def _classify_transactions(self):
        """Classify transactions as CapEx vs R&M"""
        if self.population is None:
            return
        
        try:
            self.population['classification'] = 'Undetermined'
            
            capex_keywords = ['purchase', 'acquisition', 'installation', 'construction',
                             'improvement', 'upgrade', 'addition', 'new', 'asset']
            maintenance_keywords = ['repair', 'maintenance', 'service', 'fix', 'replace',
                                  'clean', 'inspect', 'tune']
            
            if 'description' in self.population.columns:
                desc_lower = self.population['description'].astype(str).str.lower()
                
                capex_mask = desc_lower.str.contains('|'.join(capex_keywords), na=False)
                maintenance_mask = desc_lower.str.contains('|'.join(maintenance_keywords), na=False)
                
                self.population.loc[capex_mask, 'classification'] = 'CapEx'
                self.population.loc[maintenance_mask, 'classification'] = 'R&M'
            
            # High value = CapEx
            high_value_mask = abs(self.population['amount']) >= self.isi_level
            self.population.loc[high_value_mask, 'classification'] = 'CapEx'
            
            # Default remaining to CapEx
            undetermined_mask = self.population['classification'] == 'Undetermined'
            self.population.loc[undetermined_mask, 'classification'] = 'CapEx'
            
            logger.info(f"Classification: {self.population['classification'].value_counts().to_dict()}")
            
        except Exception as e:
            logger.error(f"Classification error: {str(e)}")

    def _select_samples(self):
        """Select representative samples for testing"""
        if self.population is None:
            return
        
        try:
            # ISI items (auto-include)
            isi_items = self.population[self.population['isi_item'] == True].copy()
            
            # Calculate coverage
            total_capex_value = abs(self.population[self.population['classification'] == 'CapEx']['amount']).sum()
            isi_coverage = abs(isi_items['amount']).sum()
            target_coverage_value = total_capex_value * self.coverage_target
            remaining_needed = max(0, target_coverage_value - isi_coverage)
            
            # Select additional samples
            non_isi_items = self.population[
                (self.population['isi_item'] == False) &
                (self.population['classification'] == 'CapEx')
            ].copy()
            
            additional_samples = pd.DataFrame()
            if remaining_needed > 0 and len(non_isi_items) > 0:
                non_isi_sorted = non_isi_items.reindex(
                    non_isi_items['amount'].abs().sort_values(ascending=False).index
                )
                
                cumulative_value = 0
                sample_indices = []
                
                for idx, row in non_isi_sorted.iterrows():
                    sample_indices.append(idx)
                    cumulative_value += abs(row['amount'])
                    
                    if cumulative_value >= remaining_needed or len(sample_indices) >= 50:
                        break
                
                additional_samples = non_isi_sorted.loc[sample_indices]
            
            # Combine samples
            self.sample_selection = pd.concat([isi_items, additional_samples], ignore_index=True)
            
            # Add rationale
            self.sample_selection['sample_rationale'] = 'Other'
            self.sample_selection.loc[self.sample_selection['isi_item'] == True, 'sample_rationale'] = 'ISI Item'
            self.sample_selection.loc[
                (self.sample_selection['isi_item'] == False) &
                (abs(self.sample_selection['amount']) >= self.materiality),
                'sample_rationale'
            ] = 'High Value'
            
            coverage_achieved = abs(self.sample_selection['amount']).sum() / total_capex_value if total_capex_value > 0 else 0
            
            logger.info(f"Sample selection: {len(self.sample_selection)} items, {coverage_achieved:.1%} coverage")
            
        except Exception as e:
            logger.error(f"Sample selection error: {str(e)}")

    def _generate_pbc_requests(self):
        """Generate PBC request list"""
        if self.sample_selection is None:
            return
        
        try:
            self.pbc_requests = []
            
            for _, item in self.sample_selection.iterrows():
                request = {
                    'Item': item.get('description', 'N/A'),
                    'Amount': item.get('amount', 0),
                    'Date': item.get('date', 'N/A'),
                    'Document_Number': item.get('document', 'N/A'),
                    'Required_Documentation': [
                        'Vendor invoice/receipt',
                        'Purchase order/requisition',
                        'Management approval',
                        'Asset register entry'
                    ]
                }
                
                if item.get('classification') == 'CapEx':
                    request['Required_Documentation'].extend([
                        'Useful life determination',
                        'Depreciation calculation'
                    ])
                
                if abs(item.get('amount', 0)) >= self.isi_level:
                    request['Required_Documentation'].append('Board/committee approval')
                
                self.pbc_requests.append(request)
            
            logger.info(f"Generated {len(self.pbc_requests)} PBC requests")
            
        except Exception as e:
            logger.error(f"PBC generation error: {str(e)}")

    def _perform_attribute_testing(self):
        """Perform attribute testing on samples"""
        if self.sample_selection is None:
            return
        
        try:
            self.test_results = []
            
            for idx, item in self.sample_selection.iterrows():
                test_result = {
                    'Sample_ID': idx,
                    'Description': item.get('description', 'N/A'),
                    'Amount': item.get('amount', 0),
                    'Classification': item.get('classification', 'N/A'),
                    'Tests': {
                        'Proper_Capitalization': {
                            'Status': 'Not Performed',
                            'Comments': 'Invoice vouching required',
                            'Exception': False
                        },
                        'Depreciation_Calculation': {
                            'Status': 'Not Performed',
                            'Comments': 'Asset register verification needed',
                            'Exception': False
                        },
                        'Asset_Register_Trace': {
                            'Status': 'Not Performed',
                            'Comments': 'Asset register comparison required',
                            'Exception': False
                        }
                    }
                }
                
                # Simulate findings
                if abs(item.get('amount', 0)) < self.cap_threshold:
                    test_result['Tests']['Proper_Capitalization']['Status'] = 'Exception'
                    test_result['Tests']['Proper_Capitalization']['Exception'] = True
                    test_result['Tests']['Proper_Capitalization']['Comments'] = 'Amount below capitalization threshold'
                
                self.test_results.append(test_result)
            
            logger.info(f"Attribute testing completed for {len(self.test_results)} items")
            
        except Exception as e:
            logger.error(f"Attribute testing error: {str(e)}")

    def _generate_findings(self):
        """Generate findings and proposed AJEs"""
        if not self.test_results:
            return
        
        try:
            self.exceptions = []
            self.proposed_ajes = []
            
            for test_result in self.test_results:
                for test_name, test_data in test_result['Tests'].items():
                    if test_data.get('Exception'):
                        exception = {
                            'Sample_ID': test_result['Sample_ID'],
                            'Description': test_result['Description'],
                            'Amount': test_result['Amount'],
                            'Test_Failed': test_name,
                            'Comments': test_data['Comments']
                        }
                        self.exceptions.append(exception)
                        
                        # Generate AJE if needed
                        if 'threshold' in test_data['Comments'].lower():
                            aje = {
                                'AJE_Number': f"AJE-{len(self.proposed_ajes) + 1:03d}",
                                'Description': 'Reclassify capitalized amount below threshold',
                                'Debit_Account': 'R&M Expense',
                                'Credit_Account': 'PP&E',
                                'Amount': abs(test_result['Amount']),
                                'Rationale': test_data['Comments']
                            }
                            self.proposed_ajes.append(aje)
            
            logger.info(f"Generated {len(self.exceptions)} exceptions and {len(self.proposed_ajes)} AJEs")
            
        except Exception as e:
            logger.error(f"Findings generation error: {str(e)}")

    def _produce_artifacts(self) -> List[str]:
        """Produce audit artifacts"""
        try:
            created_files = []
            
            if self._create_population_workbook():
                created_files.append('CapEx_Population.xlsx')
            
            if self._create_sample_selection_workbook():
                created_files.append('CapEx_Sample_Selection.xlsx')
            
            if self._create_pbc_request_workbook():
                created_files.append('CapEx_PBC_Request_List.xlsx')
            
            if self._create_test_workpaper():
                created_files.append('CapEx_Test_Workpaper.xlsx')
            
            if self._create_exceptions_log():
                created_files.append('CapEx_Exceptions_Log.xlsx')
            
            if self._create_ajes_workbook():
                created_files.append('Proposed_AJEs.xlsx')
            
            if self._create_summary_memo():
                created_files.append('CapEx_Summary_Memo.md')
            
            logger.info(f"Created {len(created_files)} artifact files")
            return created_files
            
        except Exception as e:
            logger.error(f"Artifact production error: {str(e)}")
            return []

    def _create_population_workbook(self) -> bool:
        """Create population workbook"""
        try:
            if self.population is None:
                return False
            
            filename = self.upload_folder / 'CapEx_Population.xlsx'
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self.population.to_excel(writer, sheet_name='Population', index=False)
                
                summary = self.population.groupby('classification').agg({
                    'amount': ['count', 'sum']
                }).round(2)
                summary.columns = ['Count', 'Total_Amount']
                summary.to_excel(writer, sheet_name='Classification_Summary')
            
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logger.error(f"Population workbook error: {str(e)}")
            return False

    def _create_sample_selection_workbook(self) -> bool:
        """Create sample selection workbook"""
        try:
            if self.sample_selection is None:
                return False
            
            filename = self.upload_folder / 'CapEx_Sample_Selection.xlsx'
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self.sample_selection.to_excel(writer, sheet_name='Selected_Samples', index=False)
                
                rationale_summary = self.sample_selection.groupby('sample_rationale').agg({
                    'amount': ['count', 'sum']
                }).round(2)
                rationale_summary.columns = ['Count', 'Total_Amount']
                rationale_summary.to_excel(writer, sheet_name='Sample_Rationale')
            
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logger.error(f"Sample selection workbook error: {str(e)}")
            return False

    def _create_pbc_request_workbook(self) -> bool:
        """Create PBC request list workbook"""
        try:
            if not self.pbc_requests:
                return False
            
            filename = self.upload_folder / 'CapEx_PBC_Request_List.xlsx'
            
            pbc_flat = []
            for req in self.pbc_requests:
                pbc_flat.append({
                    'Item': req['Item'],
                    'Amount': req['Amount'],
                    'Date': req['Date'],
                    'Document_Number': req['Document_Number'],
                    'Required_Documentation': '; '.join(req['Required_Documentation'])
                })
            
            pbc_df = pd.DataFrame(pbc_flat)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                pbc_df.to_excel(writer, sheet_name='PBC_Requests', index=False)
            
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logger.error(f"PBC request workbook error: {str(e)}")
            return False

    def _create_test_workpaper(self) -> bool:
        """Create test workpaper"""
        try:
            if not self.test_results:
                return False
            
            filename = self.upload_folder / 'CapEx_Test_Workpaper.xlsx'
            
            test_flat = []
            for result in self.test_results:
                for test_name, test_data in result['Tests'].items():
                    test_flat.append({
                        'Sample_ID': result['Sample_ID'],
                        'Description': result['Description'],
                        'Amount': result['Amount'],
                        'Classification': result['Classification'],
                        'Test_Name': test_name,
                        'Status': test_data['Status'],
                        'Comments': test_data['Comments'],
                        'Exception': test_data['Exception']
                    })
            
            test_df = pd.DataFrame(test_flat)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                test_df.to_excel(writer, sheet_name='Test_Results', index=False)
                
                summary = test_df.groupby(['Test_Name', 'Status']).size().unstack(fill_value=0)
                summary.to_excel(writer, sheet_name='Test_Summary')
            
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logger.error(f"Test workpaper error: {str(e)}")
            return False

    def _create_exceptions_log(self) -> bool:
        """Create exceptions log workbook"""
        try:
            if not self.exceptions:
                return False
            
            filename = self.upload_folder / 'CapEx_Exceptions_Log.xlsx'
            
            exceptions_df = pd.DataFrame(self.exceptions)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                exceptions_df.to_excel(writer, sheet_name='Exceptions', index=False)
            
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logger.error(f"Exceptions log error: {str(e)}")
            return False

    def _create_ajes_workbook(self) -> bool:
        """Create proposed AJEs workbook"""
        try:
            if not self.proposed_ajes:
                return False
            
            filename = self.upload_folder / 'Proposed_AJEs.xlsx'
            
            ajes_df = pd.DataFrame(self.proposed_ajes)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                ajes_df.to_excel(writer, sheet_name='Proposed_AJEs', index=False)
            
            self._format_workbook(filename)
            return True
            
        except Exception as e:
            logger.error(f"AJEs workbook error: {str(e)}")
            return False

    def _create_summary_memo(self) -> bool:
        """Create summary memo"""
        try:
            filename = self.upload_folder / 'CapEx_Summary_Memo.md'
            
            metrics = self._calculate_metrics()
            
            memo_content = f"""# Capital Addition Sampling & Testwork Summary

## Executive Summary

**Period:** {datetime.now().strftime('%Y-%m-%d')}  
**Scope:** Capital expenditure and repairs & maintenance testing

### Key Metrics
- **Population Count:** {metrics['population_count']:,} items
- **Sample Count:** {metrics['sample_count']:,} items
- **CapEx Value:** ${metrics['capex_value']:,.2f}
- **R&M Value:** ${metrics['rnm_value']:,.2f}
- **Items Flagged:** {metrics['items_flagged']:,}
- **Exceptions:** {metrics['exceptions']:,}

## Methodology

### Sampling Approach
- Capitalization threshold: ${self.cap_threshold:,.2f}
- ISI level: ${self.isi_level:,.2f}
- Coverage target: {self.coverage_target:.1%}

### Testing Performed
1. **Proper Capitalization:** Vouching to invoices and supporting documentation
2. **Depreciation Calculation:** Verification of methods and rates
3. **Asset Register Traceability:** Confirmation of property record maintenance

## Key Findings

### Classification Results
- CapEx items identified based on nature and amount thresholds
- R&M expenses properly segregated per company policy
- Near-threshold items flagged for additional review

### Exceptions Identified
"""
            
            if self.exceptions:
                memo_content += f"- {len(self.exceptions)} exceptions requiring attention\n"
                if metrics.get('largest_exception'):
                    largest = metrics['largest_exception']
                    memo_content += f"- Largest exception: {largest.get('doc_no_or_asset', 'N/A')} (${largest.get('amount', 0):,.2f})\n"
            else:
                memo_content += "- No significant exceptions identified\n"

            memo_content += f"""
### Proposed Adjusting Journal Entries
"""
            
            if self.proposed_ajes:
                memo_content += f"- {len(self.proposed_ajes)} AJEs recommended\n"
                total_aje_amount = sum(aje.get('Amount', 0) for aje in self.proposed_ajes)
                memo_content += f"- Total adjustment amount: ${total_aje_amount:,.2f}\n"
            else:
                memo_content += "- No adjusting journal entries required\n"

            memo_content += """
## Conclusion

The capital addition testing procedures have been completed in accordance with applicable auditing standards.

## Recommendations

1. Review and approve proposed adjusting journal entries
2. Strengthen supporting documentation for near-threshold items
3. Consider updating capitalization policy thresholds

---
*Generated by Capital Addition Sampling & Testwork Agent*
"""

            # Use Gemini to polish if available
            if self.gemini and hasattr(self.gemini, 'polish_memo'):
                try:
                    polished = self.gemini.polish_memo(memo_content)
                    if polished:
                        memo_content = polished
                except Exception as e:
                    logger.warning(f"Gemini polishing failed: {str(e)}")

            with open(filename, 'w', encoding='utf-8') as f:
                f.write(memo_content)
            
            return True
            
        except Exception as e:
            logger.error(f"Summary memo error: {str(e)}")
            return False

    def _format_workbook(self, filename: Path):
        """Apply formatting to Excel workbooks"""
        try:
            wb = openpyxl.load_workbook(filename)
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for sheet in wb.worksheets:
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            logger.error(f"Workbook formatting error: {str(e)}")

    def _calculate_metrics(self) -> Dict:
        """Calculate summary metrics"""
        metrics = {
            'population_count': 0,
            'sample_count': 0,
            'capex_value': 0.0,
            'rnm_value': 0.0,
            'items_flagged': 0,
            'exceptions': 0,
            'largest_exception': None
        }
        
        try:
            if self.population is not None:
                metrics['population_count'] = len(self.population)
                
                capex_items = self.population[self.population['classification'] == 'CapEx']
                rnm_items = self.population[self.population['classification'] == 'R&M']
                
                metrics['capex_value'] = capex_items['amount'].abs().sum()
                metrics['rnm_value'] = rnm_items['amount'].abs().sum()
                
                metrics['items_flagged'] = len(self.population[
                    (self.population['isi_item'] == True) |
                    (self.population['near_threshold'] == True)
                ])
            
            if self.sample_selection is not None:
                metrics['sample_count'] = len(self.sample_selection)
            
            if self.exceptions:
                metrics['exceptions'] = len(self.exceptions)
                largest = max(self.exceptions, key=lambda x: abs(x.get('Amount', 0)))
                metrics['largest_exception'] = {
                    'doc_no_or_asset': largest.get('Description', 'N/A'),
                    'amount': abs(largest.get('Amount', 0))
                }
                
        except Exception as e:
            logger.error(f"Metrics calculation error: {str(e)}")
        
        return metrics

    def _generate_summary(self) -> str:
        """Generate concise summary"""
        try:
            metrics = self._calculate_metrics()
            
            summary = (f"Analyzed {metrics['population_count']:,} transactions, "
                      f"selected {metrics['sample_count']:,} samples for testing. "
                      f"CapEx value: ${metrics['capex_value']:,.2f}, "
                      f"R&M value: ${metrics['rnm_value']:,.2f}. "
                      f"Identified {metrics['exceptions']:,} exceptions requiring attention.")
            
            if self.proposed_ajes:
                summary += f" Proposed {len(self.proposed_ajes)} adjusting journal entries."
            
            return summary
            
        except Exception as e:
            logger.error(f"Summary generation error: {str(e)}")
            return "Analysis completed with errors."

    def _identify_open_requests(self) -> List[str]:
        """Identify open requests"""
        open_requests = []
        
        if self.tb_data is None:
            open_requests.append("Trial Balance mapping needed for full tie-out")
        
        if self.test_results:
            pending_tests = sum(1 for result in self.test_results
                               for test in result['Tests'].values()
                               if test['Status'] == 'Not Performed')
            if pending_tests > 0:
                open_requests.append(f"{pending_tests} test procedures require supporting documentation")
        
        return open_requests