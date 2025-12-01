# import os
# from flask import Flask, render_template, request, jsonify, send_file, session
# import pandas as pd
# import json
# from werkzeug.utils import secure_filename
# import google.generativeai as genai
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# import numbers

# app = Flask(__name__)
# app.secret_key = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')

# google_api_key = os.environ.get('GOOGLE_API_KEY')
# if google_api_key:
#     genai.configure(api_key=google_api_key)

# UPLOAD_FOLDER = 'uploads'
# ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

# if not os.path.exists(UPLOAD_FOLDER):
#     os.makedirs(UPLOAD_FOLDER)

# app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# def allowed_file(filename):
#     return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# def sanitize_excel_string(s):
#     """Sanitize string for Excel to prevent formula injection"""
#     if isinstance(s, str) and s and s[0] in ('=', '+', '-', '@'):
#         return "'" + s
#     return s

# def read_budget_file(filepath):
#     if filepath.endswith('.csv'):
#         df = pd.read_csv(filepath)
#     else:
#         df = pd.read_excel(filepath)
#     return df

# def df_to_json_safe(df):
#     """Convert DataFrame to JSON-safe dict, replacing NaN with None"""
#     df_copy = df.copy()
#     df_copy = df_copy.fillna('')
#     return df_copy.to_dict('records')

# def export_budget_to_excel(df, filepath, client_name='', user_name='', budget_period=''):
#     """Export budget DataFrame to Excel with header information"""
#     from openpyxl import Workbook
    
#     client_name = sanitize_excel_string(client_name)
#     user_name = sanitize_excel_string(user_name)
#     budget_period = sanitize_excel_string(budget_period)
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Proposed Budget"
    
#     current_row = 1
    
#     if client_name or user_name or budget_period:
#         if client_name:
#             ws.cell(row=current_row, column=1, value="Client Name:")
#             ws.cell(row=current_row, column=2, value=client_name)
#             ws.cell(row=current_row, column=1).font = Font(bold=True)
#             current_row += 1
        
#         if user_name:
#             ws.cell(row=current_row, column=1, value="Prepared By:")
#             ws.cell(row=current_row, column=2, value=user_name)
#             ws.cell(row=current_row, column=1).font = Font(bold=True)
#             current_row += 1
        
#         if budget_period:
#             ws.cell(row=current_row, column=1, value="Budget Period:")
#             ws.cell(row=current_row, column=2, value=budget_period)
#             ws.cell(row=current_row, column=1).font = Font(bold=True)
#             current_row += 1
        
#         current_row += 1
    
#     header_row = current_row
#     for col_idx, column_name in enumerate(df.columns, start=1):
#         cell = ws.cell(row=header_row, column=col_idx, value=column_name)
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
#         cell.alignment = Alignment(horizontal='center')
    
#     for row_idx, row_data in enumerate(df.itertuples(index=False), start=header_row+1):
#         for col_idx, value in enumerate(row_data, start=1):
#             sanitized_value = sanitize_excel_string(value) if isinstance(value, str) else value
#             cell = ws.cell(row=row_idx, column=col_idx, value=sanitized_value)
#             if df.columns[col_idx-1] != 'GL Account details' and isinstance(value, numbers.Number):
#                 cell.number_format = '#,##0.00'
            
#             if row_idx == header_row + len(df) and df.iloc[-1]['GL Account details'] == 'TOTAL':
#                 cell.font = Font(bold=True)
#                 cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    
#     for col in ws.columns:
#         max_length = 0
#         column = col[0].column_letter
#         for cell in col:
#             try:
#                 cell_value_str = str(cell.value) if cell.value is not None else ''
#                 if len(cell_value_str) > max_length:
#                     max_length = len(cell_value_str)
#             except:
#                 pass
#         adjusted_width = max(min(max_length + 2, 50), 12)
#         ws.column_dimensions[column].width = adjusted_width
    
#     wb.save(filepath)
#     return filepath

# def clean_numeric_column(series):
#     if pd.api.types.is_numeric_dtype(series):
#         return series
    
#     cleaned = series.astype(str).str.replace(',', '').str.replace('$', '').str.strip()
#     cleaned = cleaned.replace(['-', '', 'nan', 'None'], '0')
#     return pd.to_numeric(cleaned, errors='coerce').fillna(0)

# def find_column(df, possible_names):
#     for name in possible_names:
#         for col in df.columns:
#             if name.lower() in col.lower():
#                 return col
#     return None

# def analyze_gl_transactions(df):
#     try:
#         gl_account_col = find_column(df, ['Head', 'GL Account', 'Account', 'GL Account details', 'Program Name', 'Program', 'Category', 'Description', 'Item'])
#         amount_col = find_column(df, ['Amount', 'Debit', 'Expense', 'Total', 'Cost'])
        
#         if not gl_account_col or not amount_col:
#             print(f"Could not find required columns. GL Account column: {gl_account_col}, Amount column: {amount_col}")
#             return None
        
#         df_copy = df.copy()
#         df_copy[amount_col] = clean_numeric_column(df_copy[amount_col])
        
#         summary = df_copy.groupby(gl_account_col)[amount_col].sum().reset_index()
#         summary.columns = ['GL Account details', 'Actual Expenses']
        
#         summary['Actual Expenses'] = summary['Actual Expenses'].round(2)
        
#         return summary
#     except Exception as e:
#         print(f"Error analyzing GL transactions: {e}")
#         return None

# def prepare_budget(prior_year_df=None, gl_df=None, inflation_rate=0.0, planned_adjustments='', manual_data=None):
#     result = None
    
#     if prior_year_df is not None:
#         gl_account_col = find_column(prior_year_df, ['GL Account details', 'GL Account', 'Head', 'Account', 'Program Name', 'Program', 'Project', 'Department'])
#         budget_col = find_column(prior_year_df, ['Budget', 'Prior Year Budget', 'Amount', 'Budgeted'])
        
#         result = pd.DataFrame()
        
#         if gl_account_col:
#             result['GL Account details'] = prior_year_df[gl_account_col]
#         else:
#             result['GL Account details'] = 'Item ' + (prior_year_df.index + 1).astype(str)
        
#         if budget_col:
#             result['Prior Year Budget'] = clean_numeric_column(prior_year_df[budget_col])
#         else:
#             result['Prior Year Budget'] = 0
        
#         if gl_df is not None:
#             gl_summary = analyze_gl_transactions(gl_df)
#             if gl_summary is not None:
#                 result = result.merge(gl_summary, on='GL Account details', how='left')
#                 result['Actual Expenses'] = result['Actual Expenses'].fillna(0)
                
#                 result['Carryover'] = (result['Prior Year Budget'] - result['Actual Expenses']).clip(lower=0)
                
#                 result['Proposed Budget'] = result.apply(
#                     lambda row: (max(row['Actual Expenses'] * 1.05, row['Prior Year Budget']) + row['Carryover']) * (1 + inflation_rate / 100),
#                     axis=1
#                 )
#             else:
#                 result['Actual Expenses'] = 0
#                 result['Carryover'] = 0
#                 result['Proposed Budget'] = result['Prior Year Budget'] * (1 + inflation_rate / 100)
#         else:
#             result['Actual Expenses'] = 0
#             result['Carryover'] = 0
#             result['Proposed Budget'] = result['Prior Year Budget'] * (1 + inflation_rate / 100)
    
#     elif gl_df is not None:
#         gl_summary = analyze_gl_transactions(gl_df)
#         if gl_summary is not None:
#             result = gl_summary.copy()
#             result['Prior Year Budget'] = 0
#             result['Carryover'] = 0
#             result['Proposed Budget'] = result['Actual Expenses'] * (1 + inflation_rate / 100) * 1.05
    
#     elif manual_data is not None:
#         result = pd.DataFrame(manual_data)
#         if 'Proposed Budget' not in result.columns and 'Amount' in result.columns:
#             result['Amount'] = clean_numeric_column(result['Amount'])
#             result['Proposed Budget'] = result['Amount'] * (1 + inflation_rate / 100)
        
#         if 'GL Account details' not in result.columns:
#             result['GL Account details'] = 'General Fund'
#         if 'Prior Year Budget' not in result.columns:
#             result['Prior Year Budget'] = 0
#         if 'Actual Expenses' not in result.columns:
#             result['Actual Expenses'] = 0
#         if 'Carryover' not in result.columns:
#             result['Carryover'] = 0
    
#     if result is not None:
#         result['Reclass Budget'] = 0.0
#         result['Final Proposed Budget'] = result['Proposed Budget']
        
#         result = result[['GL Account details', 'Prior Year Budget', 'Actual Expenses', 'Carryover', 'Proposed Budget', 'Reclass Budget', 'Final Proposed Budget']]
        
#         total_row = pd.DataFrame([{
#             'GL Account details': 'TOTAL',
#             'Prior Year Budget': result['Prior Year Budget'].sum(),
#             'Actual Expenses': result['Actual Expenses'].sum(),
#             'Carryover': result['Carryover'].sum(),
#             'Proposed Budget': result['Proposed Budget'].sum(),
#             'Reclass Budget': 0.0,
#             'Final Proposed Budget': result['Proposed Budget'].sum()
#         }])
        
#         result = pd.concat([result, total_row], ignore_index=True)
    
#     return result

# def allocate_lump_sum_budget(prior_year_df, total_funding, inflation_rate=0.0):
#     """
#     Allocate a lump sum budget proportionally based on prior year allocations
#     """
#     gl_account_col = find_column(prior_year_df, ['GL Account details', 'GL Account', 'Head', 'Account', 'Program Name', 'Program', 'Project', 'Department'])
#     budget_col = find_column(prior_year_df, ['Budget', 'Prior Year Budget', 'Amount', 'Budgeted'])
    
#     result = pd.DataFrame()
    
#     if gl_account_col:
#         result['GL Account details'] = prior_year_df[gl_account_col]
#     else:
#         result['GL Account details'] = 'Item ' + (prior_year_df.index + 1).astype(str)
    
#     if budget_col:
#         result['Prior Year Budget'] = clean_numeric_column(prior_year_df[budget_col])
#     else:
#         result['Prior Year Budget'] = 0
    
#     total_prior = result['Prior Year Budget'].sum()
    
#     if total_prior > 0:
#         result['Proposed Budget'] = ((result['Prior Year Budget'] / total_prior) * total_funding).round(2)
        
#         allocated_sum = result['Proposed Budget'].sum()
#         residual = total_funding - allocated_sum
#         if residual != 0 and len(result) > 0:
#             largest_idx = result['Proposed Budget'].idxmax()
#             result.at[largest_idx, 'Proposed Budget'] += residual
#             result['Proposed Budget'] = result['Proposed Budget'].round(2)
#     else:
#         num_items = len(result)
#         result['Proposed Budget'] = (total_funding / num_items) if num_items > 0 else 0
        
#         allocated_sum = result['Proposed Budget'].sum()
#         residual = total_funding - allocated_sum
#         if residual != 0 and len(result) > 0:
#             result.at[result.index[-1], 'Proposed Budget'] += residual
#             result['Proposed Budget'] = result['Proposed Budget'].round(2)
    
#     result['Actual Expenses'] = 0
#     result['Carryover'] = 0
#     result['Reclass Budget'] = 0.0
#     result['Final Proposed Budget'] = result['Proposed Budget']
    
#     result = result[['GL Account details', 'Prior Year Budget', 'Actual Expenses', 'Carryover', 'Proposed Budget', 'Reclass Budget', 'Final Proposed Budget']]
    
#     total_row = pd.DataFrame([{
#         'GL Account details': 'TOTAL',
#         'Prior Year Budget': result['Prior Year Budget'].sum(),
#         'Actual Expenses': result['Actual Expenses'].sum(),
#         'Carryover': result['Carryover'].sum(),
#         'Proposed Budget': result['Proposed Budget'].sum(),
#         'Reclass Budget': 0.0,
#         'Final Proposed Budget': result['Proposed Budget'].sum()
#     }])
    
#     result = pd.concat([result, total_row], ignore_index=True)
    
#     return result

# def get_ai_budget_recommendations(prior_year_df, gl_df, inflation_rate, planned_adjustments):
#     context = f"Inflation Rate: {inflation_rate}%\nPlanned Adjustments: {planned_adjustments}\n\n"
    
#     if prior_year_df is not None:
#         context += f"Prior Year Budget:\n{prior_year_df.to_string(index=False, max_rows=20)}\n\n"
    
#     if gl_df is not None:
#         gl_summary = analyze_gl_transactions(gl_df)
#         if gl_summary is not None:
#             context += f"GL Transaction Summary:\n{gl_summary.to_string(index=False, max_rows=20)}\n\n"
    
#     prompt = f"""You are an AI-powered Budget Preparation Assistant. Your goal is to help create accurate, detailed, and user-friendly proposed budgets.

# Context:
# {context}

# Based on the provided data, please:

# 1. **Analyze spending patterns**: Identify trends from prior year budget and actual expenses
# 2. **Recommend budget allocations**: Suggest appropriate budget amounts for each GL account
# 3. **Highlight opportunities**: Identify areas for potential savings or reallocation
# 4. **Provide justification**: Explain your reasoning for budget recommendations
# 5. **Consider adjustments**: Factor in inflation ({inflation_rate}%) and planned changes

# Keep your response structured, actionable, and focused on helping prepare an accurate proposed budget."""

#     try:
#         api_key = os.environ.get('GOOGLE_API_KEY')
#         if not api_key:
#             return "Google API key not configured. Please set GOOGLE_API_KEY environment variable."
        
#         model_obj = genai.GenerativeModel('gemini-1.5-flash')
#         response = model_obj.generate_content(prompt)
#         return response.text
#     except Exception as e:
#         return f"AI analysis unavailable: {str(e)}"

# @app.route('/')
# def index():
#     return render_template('index.html')

# @app.route('/upload-prior-budget', methods=['POST'])
# def upload_prior_budget():
#     if 'file' not in request.files:
#         return jsonify({'error': 'No file provided'}), 400
    
#     file = request.files['file']
#     if not file.filename or file.filename == '':
#         return jsonify({'error': 'No file selected'}), 400
    
#     if file and allowed_file(file.filename):
#         filename = secure_filename(file.filename)
#         filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"prior_{filename}")
#         file.save(filepath)
        
#         session['prior_budget_file'] = f"prior_{filename}"
        
#         df = read_budget_file(filepath)
#         preview_data = df_to_json_safe(df.head(10))
#         columns = df.columns.tolist()
        
#         return jsonify({
#             'success': True,
#             'filename': filename,
#             'preview': preview_data,
#             'columns': columns,
#             'total_rows': len(df)
#         })
    
#     return jsonify({'error': 'Invalid file type. Please upload CSV or Excel file'}), 400

# @app.route('/upload-gl-data', methods=['POST'])
# def upload_gl_data():
#     if 'file' not in request.files:
#         return jsonify({'error': 'No file provided'}), 400
    
#     file = request.files['file']
#     if not file.filename or file.filename == '':
#         return jsonify({'error': 'No file selected'}), 400
    
#     if file and allowed_file(file.filename):
#         filename = secure_filename(file.filename)
#         filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"gl_{filename}")
#         file.save(filepath)
        
#         session['gl_data_file'] = f"gl_{filename}"
        
#         df = read_budget_file(filepath)
#         preview_data = df_to_json_safe(df.head(10))
#         columns = df.columns.tolist()
        
#         return jsonify({
#             'success': True,
#             'filename': filename,
#             'preview': preview_data,
#             'columns': columns,
#             'total_rows': len(df)
#         })
    
#     return jsonify({'error': 'Invalid file type. Please upload CSV or Excel file'}), 400

# @app.route('/prepare-budget', methods=['POST'])
# def prepare_budget_route():
#     data = request.json or {}
    
#     client_name = data.get('clientName', '')
#     user_name = data.get('userName', '')
#     budget_period = data.get('budgetPeriod', '')
    
#     session['client_name'] = client_name
#     session['user_name'] = user_name
#     session['budget_period'] = budget_period
    
#     prior_budget_file = session.get('prior_budget_file')
#     gl_data_file = session.get('gl_data_file')
    
#     prior_year_df = None
#     gl_df = None
    
#     if prior_budget_file:
#         try:
#             filepath = os.path.join(app.config['UPLOAD_FOLDER'], prior_budget_file)
#             prior_year_df = read_budget_file(filepath)
#         except Exception as e:
#             return jsonify({'error': f'Error reading prior year budget: {str(e)}'}), 400
    
#     if gl_data_file:
#         try:
#             filepath = os.path.join(app.config['UPLOAD_FOLDER'], gl_data_file)
#             gl_df = read_budget_file(filepath)
#         except Exception as e:
#             return jsonify({'error': f'Error reading GL data: {str(e)}'}), 400
    
#     try:
#         inflation_rate = float(data.get('inflationRate', 0))
#     except (ValueError, TypeError):
#         return jsonify({'error': 'Invalid inflation rate'}), 400
    
#     planned_adjustments = data.get('plannedAdjustments', '')
#     manual_data = data.get('manualData')
    
#     if prior_year_df is None and gl_df is None and not manual_data:
#         return jsonify({'error': 'Please upload at least one file or provide manual data'}), 400
    
#     proposed_budget_df = prepare_budget(prior_year_df, gl_df, inflation_rate, planned_adjustments, manual_data)
    
#     if proposed_budget_df is None:
#         return jsonify({'error': 'Unable to prepare budget from provided data'}), 400
    
#     ai_recommendations = get_ai_budget_recommendations(prior_year_df, gl_df, inflation_rate, planned_adjustments)
    
#     result_filename = f"proposed_budget_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#     result_filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
    
#     client_name = session.get('client_name', '')
#     export_budget_to_excel(proposed_budget_df, result_filepath, client_name, user_name, budget_period)
    
#     session['result_file'] = result_filename
    
#     total_budget = float(proposed_budget_df[proposed_budget_df['GL Account details'] != 'TOTAL']['Proposed Budget'].sum()) if len(proposed_budget_df) > 0 else 0
    
#     return jsonify({
#         'success': True,
#         'data': df_to_json_safe(proposed_budget_df),
#         'columns': proposed_budget_df.columns.tolist(),
#         'recommendations': ai_recommendations,
#         'downloadFilename': result_filename,
#         'totalBudget': total_budget
#     })

# @app.route('/prepare-budget-lump-sum', methods=['POST'])
# def prepare_budget_lump_sum():
#     data = request.json or {}
    
#     client_name = data.get('clientName', '')
#     user_name = data.get('userName', '')
#     budget_period = data.get('budgetPeriod', '')
    
#     session['client_name'] = client_name
#     session['user_name'] = user_name
#     session['budget_period'] = budget_period
    
#     prior_budget_file = session.get('prior_budget_file')
    
#     if not prior_budget_file:
#         return jsonify({'error': 'Please upload a prior year budget file first'}), 400
    
#     try:
#         total_funding = float(data.get('totalFunding', 0))
#         if total_funding <= 0:
#             return jsonify({'error': 'Total funding must be greater than zero'}), 400
#     except (ValueError, TypeError):
#         return jsonify({'error': 'Invalid total funding amount'}), 400
    
#     try:
#         inflation_rate = float(data.get('inflationRate', 0))
#     except (ValueError, TypeError):
#         return jsonify({'error': 'Invalid inflation rate'}), 400
    
#     try:
#         filepath = os.path.join(app.config['UPLOAD_FOLDER'], prior_budget_file)
#         prior_year_df = read_budget_file(filepath)
#     except Exception as e:
#         return jsonify({'error': f'Error reading prior year budget: {str(e)}'}), 400
    
#     proposed_budget_df = allocate_lump_sum_budget(prior_year_df, total_funding, inflation_rate)
    
#     if proposed_budget_df is None:
#         return jsonify({'error': 'Unable to allocate budget'}), 400
    
#     budget_period = data.get('budgetPeriod', 'Next Fiscal Year')
    
#     context = f"Total Funding: ${total_funding:,.2f}\nBudget Period: {budget_period}\n\n"
#     context += f"Budget Allocation:\n{proposed_budget_df.to_string(index=False, max_rows=20)}"
    
#     prompt = f"""You are an AI-powered Budget Preparation Assistant. 

# Context:
# {context}

# This budget was prepared by proportionally allocating a lump sum of ${total_funding:,.2f} based on prior year spending patterns for {budget_period}.

# Please provide:
# 1. **Budget narrative**: Brief summary of the allocation approach and key highlights
# 2. **Recommendations**: Suggest any adjustments or areas that may need attention
# 3. **Key observations**: Identify significant allocations or areas requiring attention

# Keep your response concise and actionable."""
    
#     ai_recommendations = ""
#     try:
#         api_key = os.environ.get('GOOGLE_API_KEY')
#         if api_key:
#             model_obj = genai.GenerativeModel('gemini-1.5-flash')
#             response = model_obj.generate_content(prompt)
#             ai_recommendations = response.text
#     except Exception as e:
#         ai_recommendations = f"AI recommendations unavailable: {str(e)}"
    
#     result_filename = f"lumpsum_budget_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#     result_filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
    
#     client_name = session.get('client_name', '')
#     export_budget_to_excel(proposed_budget_df, result_filepath, client_name, user_name, budget_period)
    
#     session['result_file'] = result_filename
    
#     total_budget = float(proposed_budget_df[proposed_budget_df['GL Account details'] != 'TOTAL']['Proposed Budget'].sum()) if len(proposed_budget_df) > 0 else 0
    
#     return jsonify({
#         'success': True,
#         'data': df_to_json_safe(proposed_budget_df),
#         'columns': proposed_budget_df.columns.tolist(),
#         'recommendations': ai_recommendations,
#         'downloadFilename': result_filename,
#         'totalBudget': total_budget,
#         'budgetPeriod': budget_period
#     })

# @app.route('/update-budget-item', methods=['POST'])
# def update_budget_item():
#     data = request.json or {}
#     result_file = session.get('result_file')
    
#     if not result_file:
#         return jsonify({'error': 'No active budget to update'}), 400
    
#     try:
#         filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_file)
#         df = pd.read_excel(filepath)
        
#         row_index = int(data.get('rowIndex', -1))
#         if row_index < 0 or row_index >= len(df):
#             return jsonify({'error': 'Invalid row index'}), 400
        
#         updates = data.get('updates', {})
#         for col, value in updates.items():
#             if col in df.columns:
#                 if col in ['Prior Year Budget', 'Actual Expenses', 'Proposed Budget', 'Carryover']:
#                     df.at[row_index, col] = float(value)
#                 else:
#                     df.at[row_index, col] = value
        
#         if 'Prior Year Budget' in updates or 'Actual Expenses' in updates or 'Proposed Budget' in updates:
#             if 'Prior Year Budget' in df.columns and 'Actual Expenses' in df.columns and 'Proposed Budget' in df.columns:
#                 prior_budget = df.at[row_index, 'Prior Year Budget']
#                 actual = df.at[row_index, 'Actual Expenses']
#                 proposed = df.at[row_index, 'Proposed Budget']
                
#                 if prior_budget > 0:
#                     df.at[row_index, 'Variance'] = actual - prior_budget
#                     df.at[row_index, 'Variance %'] = round((df.at[row_index, 'Variance'] / prior_budget) * 100, 2)
#                 elif proposed > 0 and actual >= 0:
#                     df.at[row_index, 'Variance'] = proposed - actual
#                     df.at[row_index, 'Variance %'] = round((df.at[row_index, 'Variance'] / proposed) * 100, 2) if proposed != 0 else 0
#                 else:
#                     df.at[row_index, 'Variance'] = 0
#                     df.at[row_index, 'Variance %'] = 0
                
#                 if abs(df.at[row_index, 'Variance %']) > 20:
#                     df.at[row_index, 'Flag'] = '⚠️ High Variance'
#                 else:
#                     df.at[row_index, 'Flag'] = ''
        
#         client_name = session.get('client_name', '')
#         user_name = session.get('user_name', '')
#         budget_period = session.get('budget_period', '')
#         export_budget_to_excel(df, filepath, client_name, user_name, budget_period)
        
#         return jsonify({
#             'success': True,
#             'data': df_to_json_safe(df),
#             'columns': df.columns.tolist(),
#             'totalBudget': float(df['Proposed Budget'].sum()) if 'Proposed Budget' in df.columns else 0
#         })
#     except Exception as e:
#         return jsonify({'error': f'Error updating budget: {str(e)}'}), 400

# @app.route('/update-budget-data', methods=['POST'])
# def update_budget_data():
#     """Update budget data with reclass changes before download"""
#     data = request.json or {}
#     budget_data = data.get('budgetData', [])
#     result_file = session.get('result_file')
    
#     if not result_file:
#         return jsonify({'error': 'No active budget to update'}), 400
    
#     try:
#         df = pd.DataFrame(budget_data)
        
#         non_total_mask = df['GL Account details'] != 'TOTAL'
        
#         df.loc[non_total_mask, 'Reclass Budget'] = pd.to_numeric(df.loc[non_total_mask, 'Reclass Budget'], errors='coerce').fillna(0)
        
#         reclass_total = df.loc[non_total_mask, 'Reclass Budget'].sum()
#         if abs(reclass_total) > 0.01:
#             return jsonify({
#                 'error': f'Reclass Budget must sum to zero. Current total: ${reclass_total:.2f}'
#             }), 400
        
#         df['Final Proposed Budget'] = df['Proposed Budget'] + df['Reclass Budget']
        
#         total_idx = df[df['GL Account details'] == 'TOTAL'].index
#         if len(total_idx) > 0:
#             idx = total_idx[0]
#             df.at[idx, 'Prior Year Budget'] = df.loc[non_total_mask, 'Prior Year Budget'].sum()
#             df.at[idx, 'Actual Expenses'] = df.loc[non_total_mask, 'Actual Expenses'].sum()
#             df.at[idx, 'Carryover'] = df.loc[non_total_mask, 'Carryover'].sum()
#             df.at[idx, 'Proposed Budget'] = df.loc[non_total_mask, 'Proposed Budget'].sum()
#             df.at[idx, 'Reclass Budget'] = df.loc[non_total_mask, 'Reclass Budget'].sum()
#             df.at[idx, 'Final Proposed Budget'] = df.loc[non_total_mask, 'Final Proposed Budget'].sum()
        
#         df = df[['GL Account details', 'Prior Year Budget', 'Actual Expenses', 'Carryover', 'Proposed Budget', 'Reclass Budget', 'Final Proposed Budget']]
        
#         client_name = session.get('client_name', '')
#         user_name = session.get('user_name', '')
#         budget_period = session.get('budget_period', '')
        
#         result_filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_file)
#         export_budget_to_excel(df, result_filepath, client_name, user_name, budget_period)
        
#         return jsonify({
#             'success': True,
#             'message': 'Budget data updated successfully'
#         })
#     except Exception as e:
#         return jsonify({'error': f'Failed to update budget: {str(e)}'}), 400

# @app.route('/download/<filename>')
# def download(filename):
#     safe_filename = secure_filename(filename)
#     if not allowed_file(safe_filename) and not safe_filename.endswith('.xlsx'):
#         return jsonify({'error': 'Invalid file type'}), 400
    
#     filepath = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], safe_filename))
#     upload_dir = os.path.abspath(app.config['UPLOAD_FOLDER'])
    
#     if not filepath.startswith(upload_dir):
#         return jsonify({'error': 'Invalid file path'}), 403
    
#     if not os.path.exists(filepath):
#         return jsonify({'error': 'File not found'}), 404
    
#     return send_file(filepath, as_attachment=True)


# @app.route('/transfer-budget', methods=['POST'])
# def transfer_budget():
#     data = request.json or {}
#     result_file = session.get('result_file')
    
#     if not result_file:
#         return jsonify({'error': 'No active budget to modify'}), 400
    
#     try:
#         filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_file)
#         df = pd.read_excel(filepath)
        
#         from_index = int(data.get('fromIndex', -1))
#         to_index = int(data.get('toIndex', -1))
#         amount = float(data.get('amount', 0))
        
#         if from_index < 0 or from_index >= len(df) or to_index < 0 or to_index >= len(df):
#             return jsonify({'error': 'Invalid row indices'}), 400
        
#         if df.loc[from_index, 'GL Account details'] == 'TOTAL' or df.loc[to_index, 'GL Account details'] == 'TOTAL':
#             return jsonify({'error': 'Cannot transfer from or to TOTAL row'}), 400
        
#         if amount <= 0:
#             return jsonify({'error': 'Transfer amount must be positive'}), 400
        
#         current_from_budget = float(df.loc[from_index, 'Proposed Budget'])
        
#         if current_from_budget < amount:
#             return jsonify({'error': 'Insufficient budget in source line item'}), 400
        
#         df.loc[from_index, 'Proposed Budget'] = current_from_budget - amount
#         df.loc[to_index, 'Proposed Budget'] = float(df.loc[to_index, 'Proposed Budget']) + amount
        
#         total_index = df[df['GL Account details'] == 'TOTAL'].index
#         if len(total_index) > 0:
#             total_idx = total_index[0]
#             df.loc[total_idx, 'Proposed Budget'] = df[df['GL Account details'] != 'TOTAL']['Proposed Budget'].sum()
        
#         df.to_excel(filepath, index=False)
        
#         return jsonify({
#             'success': True,
#             'data': df_to_json_safe(df),
#             'message': f'Transferred ${amount:,.2f} successfully'
#         })
    
#     except Exception as e:
#         return jsonify({'error': str(e)}), 400

# @app.route('/clear-session', methods=['POST'])
# def clear_session():
#     session.clear()
#     return jsonify({'success': True})

# if __name__ == '__main__':
#     app.run(host='0.0.0.0', port=5000, debug=False)
































import os
from flask import Flask, render_template, request, jsonify, send_file, session
from werkzeug.utils import secure_filename
import pandas as pd
import json
from datetime import datetime
from bson import ObjectId
import google.generativeai as genai

from openpyxl.styles import Font, Alignment, PatternFill

import numbers

# MongoDB imports
from pymongo import MongoClient
import gridfs

from db_config import get_database, get_gridfs

from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')

# Configure Google AI
google_api_key = os.environ.get('GOOGLE_API_KEY')
if google_api_key:
    genai.configure(api_key=google_api_key)

# MongoDB setup
db = get_database()
fs = get_gridfs()

# Collections
budgets_collection = db['budgets']
clients_collection = db['clients']
users_collection = db['users']
audit_log_collection = db['audit_log']

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}




def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def sanitize_excel_string(s):
    """Sanitize string for Excel to prevent formula injection"""
    if isinstance(s, str) and s and s[0] in ('=', '+', '-', '@'):
        return "'" + s
    return s

def read_budget_file_from_gridfs(file_id):
    """Read budget file from GridFS"""
    try:
        file_data = fs.get(ObjectId(file_id))
        filename = file_data.filename
        
        if filename.endswith('.csv'):
            df = pd.read_csv(file_data)
        else:
            df = pd.read_excel(file_data)
        return df
    except Exception as e:
        print(f"Error reading file from GridFS: {e}")
        return None

def save_file_to_gridfs(file, file_type, budget_id=None):
    """Save uploaded file to GridFS"""
    try:
        file_id = fs.put(
            file,
            filename=secure_filename(file.filename),
            metadata={
                'file_type': file_type,
                'budget_id': str(budget_id) if budget_id else None,
                'upload_date': datetime.utcnow()
            }
        )
        return str(file_id)
    except Exception as e:
        print(f"Error saving file to GridFS: {e}")
        return None

def df_to_json_safe(df):
    """Convert DataFrame to JSON-safe dict, replacing NaN with None"""
    df_copy = df.copy()
    df_copy = df_copy.fillna('')
    return df_copy.to_dict('records')

def export_budget_to_excel(df, filepath, client_name='', user_name='', budget_period=''):
    """Export budget DataFrame to Excel with header information"""
    from openpyxl import Workbook
    
    client_name = sanitize_excel_string(client_name)
    user_name = sanitize_excel_string(user_name)
    budget_period = sanitize_excel_string(budget_period)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposed Budget"
    
    current_row = 1
    
    if client_name or user_name or budget_period:
        if client_name:
            ws.cell(row=current_row, column=1, value="Client Name:")
            ws.cell(row=current_row, column=2, value=client_name)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        
        if user_name:
            ws.cell(row=current_row, column=1, value="Prepared By:")
            ws.cell(row=current_row, column=2, value=user_name)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        
        if budget_period:
            ws.cell(row=current_row, column=1, value="Budget Period:")
            ws.cell(row=current_row, column=2, value=budget_period)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        
        current_row += 1
    
    header_row = current_row
    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=header_row+1):
        for col_idx, value in enumerate(row_data, start=1):
            sanitized_value = sanitize_excel_string(value) if isinstance(value, str) else value
            cell = ws.cell(row=row_idx, column=col_idx, value=sanitized_value)
            if df.columns[col_idx-1] != 'GL Account details' and isinstance(value, numbers.Number):
                cell.number_format = '#,##0.00'
            
            if row_idx == header_row + len(df) and df.iloc[-1]['GL Account details'] == 'TOTAL':
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_value_str = str(cell.value) if cell.value is not None else ''
                if len(cell_value_str) > max_length:
                    max_length = len(cell_value_str)
            except:
                pass
        adjusted_width = max(min(max_length + 2, 50), 12)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filepath)
    return filepath

def clean_numeric_column(series):
    if pd.api.types.is_numeric_dtype(series):
        return series
    
    cleaned = series.astype(str).str.replace(',', '').str.replace('$', '').str.strip()
    cleaned = cleaned.replace(['-', '', 'nan', 'None'], '0')
    return pd.to_numeric(cleaned, errors='coerce').fillna(0)

def find_column(df, possible_names):
    for name in possible_names:
        for col in df.columns:
            if name.lower() in col.lower():
                return col
    return None

def analyze_gl_transactions(df):
    try:
        gl_account_col = find_column(df, ['Head', 'GL Account', 'Account', 'GL Account details', 'Program Name', 'Program', 'Category', 'Description', 'Item'])
        amount_col = find_column(df, ['Amount', 'Debit', 'Expense', 'Total', 'Cost'])
        
        if not gl_account_col or not amount_col:
            print(f"Could not find required columns. GL Account column: {gl_account_col}, Amount column: {amount_col}")
            return None
        
        df_copy = df.copy()
        df_copy[amount_col] = clean_numeric_column(df_copy[amount_col])
        
        summary = df_copy.groupby(gl_account_col)[amount_col].sum().reset_index()
        summary.columns = ['GL Account details', 'Actual Expenses']
        
        summary['Actual Expenses'] = summary['Actual Expenses'].round(2)
        
        return summary
    except Exception as e:
        print(f"Error analyzing GL transactions: {e}")
        return None

def prepare_budget(prior_year_df=None, gl_df=None, inflation_rate=0.0, planned_adjustments='', manual_data=None):
    result = None
    
    if prior_year_df is not None:
        gl_account_col = find_column(prior_year_df, ['GL Account details', 'GL Account', 'Head', 'Account', 'Program Name', 'Program', 'Project', 'Department'])
        budget_col = find_column(prior_year_df, ['Budget', 'Prior Year Budget', 'Amount', 'Budgeted'])
        
        result = pd.DataFrame()
        
        if gl_account_col:
            result['GL Account details'] = prior_year_df[gl_account_col]
        else:
            result['GL Account details'] = 'Item ' + (prior_year_df.index + 1).astype(str)
        
        if budget_col:
            result['Prior Year Budget'] = clean_numeric_column(prior_year_df[budget_col])
        else:
            result['Prior Year Budget'] = 0
        
        if gl_df is not None:
            gl_summary = analyze_gl_transactions(gl_df)
            if gl_summary is not None:
                result = result.merge(gl_summary, on='GL Account details', how='left')
                result['Actual Expenses'] = result['Actual Expenses'].fillna(0)
                
                result['Carryover'] = (result['Prior Year Budget'] - result['Actual Expenses']).clip(lower=0)
                
                result['Proposed Budget'] = result.apply(
                    lambda row: (max(row['Actual Expenses'] * 1.05, row['Prior Year Budget']) + row['Carryover']) * (1 + inflation_rate / 100),
                    axis=1
                )
            else:
                result['Actual Expenses'] = 0
                result['Carryover'] = 0
                result['Proposed Budget'] = result['Prior Year Budget'] * (1 + inflation_rate / 100)
        else:
            result['Actual Expenses'] = 0
            result['Carryover'] = 0
            result['Proposed Budget'] = result['Prior Year Budget'] * (1 + inflation_rate / 100)
    
    elif gl_df is not None:
        gl_summary = analyze_gl_transactions(gl_df)
        if gl_summary is not None:
            result = gl_summary.copy()
            result['Prior Year Budget'] = 0
            result['Carryover'] = 0
            result['Proposed Budget'] = result['Actual Expenses'] * (1 + inflation_rate / 100) * 1.05
    
    elif manual_data is not None:
        result = pd.DataFrame(manual_data)
        if 'Proposed Budget' not in result.columns and 'Amount' in result.columns:
            result['Amount'] = clean_numeric_column(result['Amount'])
            result['Proposed Budget'] = result['Amount'] * (1 + inflation_rate / 100)
        
        if 'GL Account details' not in result.columns:
            result['GL Account details'] = 'General Fund'
        if 'Prior Year Budget' not in result.columns:
            result['Prior Year Budget'] = 0
        if 'Actual Expenses' not in result.columns:
            result['Actual Expenses'] = 0
        if 'Carryover' not in result.columns:
            result['Carryover'] = 0
    
    if result is not None:
        result['Reclass Budget'] = 0.0
        result['Final Proposed Budget'] = result['Proposed Budget']
        
        result = result[['GL Account details', 'Prior Year Budget', 'Actual Expenses', 'Carryover', 'Proposed Budget', 'Reclass Budget', 'Final Proposed Budget']]
        
        total_row = pd.DataFrame([{
            'GL Account details': 'TOTAL',
            'Prior Year Budget': result['Prior Year Budget'].sum(),
            'Actual Expenses': result['Actual Expenses'].sum(),
            'Carryover': result['Carryover'].sum(),
            'Proposed Budget': result['Proposed Budget'].sum(),
            'Reclass Budget': 0.0,
            'Final Proposed Budget': result['Proposed Budget'].sum()
        }])
        
        result = pd.concat([result, total_row], ignore_index=True)
    
    return result

def allocate_lump_sum_budget(prior_year_df, total_funding, inflation_rate=0.0):

    gl_account_col = find_column(prior_year_df, ['GL Account details', 'GL Account', 'Head', 'Account', 'Program Name', 'Program', 'Project', 'Department'])
    budget_col = find_column(prior_year_df, ['Budget', 'Prior Year Budget', 'Amount', 'Budgeted'])
    
    result = pd.DataFrame()
    
    if gl_account_col:
        result['GL Account details'] = prior_year_df[gl_account_col]
    else:
        result['GL Account details'] = 'Item ' + (prior_year_df.index + 1).astype(str)
    
    if budget_col:
        result['Prior Year Budget'] = clean_numeric_column(prior_year_df[budget_col])
    else:
        result['Prior Year Budget'] = 0
    
    total_prior = result['Prior Year Budget'].sum()
    
    if total_prior > 0:
        result['Proposed Budget'] = ((result['Prior Year Budget'] / total_prior) * total_funding).round(2)
        
        allocated_sum = result['Proposed Budget'].sum()
        residual = total_funding - allocated_sum
        if residual != 0 and len(result) > 0:
            largest_idx = result['Proposed Budget'].idxmax()
            result.at[largest_idx, 'Proposed Budget'] += residual
            result['Proposed Budget'] = result['Proposed Budget'].round(2)
    else:
        num_items = len(result)
        result['Proposed Budget'] = (total_funding / num_items) if num_items > 0 else 0
        
        allocated_sum = result['Proposed Budget'].sum()
        residual = total_funding - allocated_sum
        if residual != 0 and len(result) > 0:
            result.at[result.index[-1], 'Proposed Budget'] += residual
            result['Proposed Budget'] = result['Proposed Budget'].round(2)
    
    result['Actual Expenses'] = 0
    result['Carryover'] = 0
    result['Reclass Budget'] = 0.0
    result['Final Proposed Budget'] = result['Proposed Budget']
    
    result = result[['GL Account details', 'Prior Year Budget', 'Actual Expenses', 'Carryover', 'Proposed Budget', 'Reclass Budget', 'Final Proposed Budget']]
    
    total_row = pd.DataFrame([{
        'GL Account details': 'TOTAL',
        'Prior Year Budget': result['Prior Year Budget'].sum(),
        'Actual Expenses': result['Actual Expenses'].sum(),
        'Carryover': result['Carryover'].sum(),
        'Proposed Budget': result['Proposed Budget'].sum(),
        'Reclass Budget': 0.0,
        'Final Proposed Budget': result['Proposed Budget'].sum()
    }])
    
    result = pd.concat([result, total_row], ignore_index=True)
    
    return result

def get_ai_budget_recommendations(prior_year_df, gl_df, inflation_rate, planned_adjustments):
    context = f"Inflation Rate: {inflation_rate}%\nPlanned Adjustments: {planned_adjustments}\n\n"
    
    if prior_year_df is not None:
        context += f"Prior Year Budget:\n{prior_year_df.to_string(index=False, max_rows=20)}\n\n"
    
    if gl_df is not None:
        gl_summary = analyze_gl_transactions(gl_df)
        if gl_summary is not None:
            context += f"GL Transaction Summary:\n{gl_summary.to_string(index=False, max_rows=20)}\n\n"
    
    prompt = f"""You are an AI-powered Budget Preparation Assistant. Your goal is to help create accurate, detailed, and user-friendly proposed budgets.

Context:
{context}

Based on the provided data, please:

1. **Analyze spending patterns**: Identify trends from prior year budget and actual expenses
2. **Recommend budget allocations**: Suggest appropriate budget amounts for each GL account
3. **Highlight opportunities**: Identify areas for potential savings or reallocation
4. **Provide justification**: Explain your reasoning for budget recommendations
5. **Consider adjustments**: Factor in inflation ({inflation_rate}%) and planned changes

Keep your response structured, actionable, and focused on helping prepare an accurate proposed budget."""

    try:
        api_key = os.environ.get('GOOGLE_API_KEY')
        if not api_key:
            return "Google API key not configured. Please set GOOGLE_API_KEY environment variable."
        
        model_obj = genai.GenerativeModel('models/gemini-2.5-flash')
        response = model_obj.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"AI analysis unavailable: {str(e)}"

def log_audit(budget_id, action, user, changes=None):
    """Log actions to audit trail"""
    audit_log_collection.insert_one({
        'budget_id': ObjectId(budget_id),
        'action': action,
        'user': user,
        'changes': changes,
        'timestamp': datetime.utcnow()
    })

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload-prior-budget', methods=['POST'])
def upload_prior_budget():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if not file.filename or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        file_id = save_file_to_gridfs(file, 'prior_budget')
        
        if not file_id:
            return jsonify({'error': 'Failed to save file'}), 500
        
        session['prior_budget_file_id'] = file_id
        
        df = read_budget_file_from_gridfs(file_id)
        preview_data = df_to_json_safe(df.head(10))
        columns = df.columns.tolist()
        
        return jsonify({
            'success': True,
            'filename': file.filename,
            'preview': preview_data,
            'columns': columns,
            'total_rows': len(df)
        })
    
    return jsonify({'error': 'Invalid file type. Please upload CSV or Excel file'}), 400

@app.route('/upload-gl-data', methods=['POST'])
def upload_gl_data():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if not file.filename or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        file_id = save_file_to_gridfs(file, 'gl_data')
        
        if not file_id:
            return jsonify({'error': 'Failed to save file'}), 500
        
        session['gl_data_file_id'] = file_id
        
        df = read_budget_file_from_gridfs(file_id)
        preview_data = df_to_json_safe(df.head(10))
        columns = df.columns.tolist()
        
        return jsonify({
            'success': True,
            'filename': file.filename,
            'preview': preview_data,
            'columns': columns,
            'total_rows': len(df)
        })
    
    return jsonify({'error': 'Invalid file type. Please upload CSV or Excel file'}), 400

@app.route('/prepare-budget', methods=['POST'])
def prepare_budget_route():
    data = request.json or {}
    
    client_name = data.get('clientName', '')
    user_name = data.get('userName', '')
    budget_period = data.get('budgetPeriod', '')
    
    prior_budget_file_id = session.get('prior_budget_file_id')
    gl_data_file_id = session.get('gl_data_file_id')
    
    prior_year_df = None
    gl_df = None
    
    if prior_budget_file_id:
        prior_year_df = read_budget_file_from_gridfs(prior_budget_file_id)
    
    if gl_data_file_id:
        gl_df = read_budget_file_from_gridfs(gl_data_file_id)
    
    try:
        inflation_rate = float(data.get('inflationRate', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid inflation rate'}), 400
    
    planned_adjustments = data.get('plannedAdjustments', '')
    manual_data = data.get('manualData')
    
    if prior_year_df is None and gl_df is None and not manual_data:
        return jsonify({'error': 'Please upload at least one file or provide manual data'}), 400
    
    proposed_budget_df = prepare_budget(prior_year_df, gl_df, inflation_rate, planned_adjustments, manual_data)
    
    if proposed_budget_df is None:
        return jsonify({'error': 'Unable to prepare budget from provided data'}), 400
    
    ai_recommendations = get_ai_budget_recommendations(prior_year_df, gl_df, inflation_rate, planned_adjustments)
    
    # Save to MongoDB
    line_items = []
    for _, row in proposed_budget_df[proposed_budget_df['GL Account details'] != 'TOTAL'].iterrows():
        line_items.append({
            'gl_account': row['GL Account details'],
            'prior_year_budget': float(row['Prior Year Budget']),
            'actual_expenses': float(row['Actual Expenses']),
            'carryover': float(row['Carryover']),
            'proposed_budget': float(row['Proposed Budget']),
            'reclass_budget': float(row['Reclass Budget']),
            'final_proposed_budget': float(row['Final Proposed Budget'])
        })
    
    total_row = proposed_budget_df[proposed_budget_df['GL Account details'] == 'TOTAL'].iloc[0]
    
    budget_doc = {
        'client_name': client_name,
        'prepared_by': user_name,
        'budget_period': budget_period,
        'method': 'detailed',
        'inflation_rate': inflation_rate,
        'planned_adjustments': planned_adjustments,
        'status': 'draft',
        'created_at': datetime.utcnow(),
        'updated_at': datetime.utcnow(),
        'line_items': line_items,
        'totals': {
            'prior_year_budget': float(total_row['Prior Year Budget']),
            'actual_expenses': float(total_row['Actual Expenses']),
            'carryover': float(total_row['Carryover']),
            'proposed_budget': float(total_row['Proposed Budget']),
            'final_proposed_budget': float(total_row['Final Proposed Budget'])
        },
        'ai_recommendations': ai_recommendations,
        'prior_budget_file_id': prior_budget_file_id,
        'gl_data_file_id': gl_data_file_id
    }
    
    result = budgets_collection.insert_one(budget_doc)
    budget_id = str(result.inserted_id)
    
    session['current_budget_id'] = budget_id
    
    log_audit(budget_id, 'created', user_name)
    
    total_budget = float(proposed_budget_df[proposed_budget_df['GL Account details'] != 'TOTAL']['Proposed Budget'].sum())
    
    return jsonify({
        'success': True,
        'budget_id': budget_id,
        'data': df_to_json_safe(proposed_budget_df),
        'columns': proposed_budget_df.columns.tolist(),
        'recommendations': ai_recommendations,

        'totalBudget': total_budget
    })

@app.route('/prepare-budget-lump-sum', methods=['POST'])
def prepare_budget_lump_sum():
    data = request.json or {}
    
    client_name = data.get('clientName', '')
    user_name = data.get('userName', '')
    budget_period = data.get('budgetPeriod', '')
    
    prior_budget_file_id = session.get('prior_budget_file_id')
    
    if not prior_budget_file_id:
        return jsonify({'error': 'Please upload a prior year budget file first'}), 400
    
    try:
        total_funding = float(data.get('totalFunding', 0))
        if total_funding <= 0:
            return jsonify({'error': 'Total funding must be greater than zero'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid total funding amount'}), 400
    
    try:
        inflation_rate = float(data.get('inflationRate', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid inflation rate'}), 400
    
    prior_year_df = read_budget_file_from_gridfs(prior_budget_file_id)
    
    if prior_year_df is None:
        return jsonify({'error': 'Error reading prior year budget'}), 400
    
    proposed_budget_df = allocate_lump_sum_budget(prior_year_df, total_funding, inflation_rate)
    
    if proposed_budget_df is None:
        return jsonify({'error': 'Unable to allocate budget'}), 400
    


    context = f"Total Funding: ${total_funding:,.2f}\nBudget Period: {budget_period}\n\n"
    context += f"Budget Allocation:\n{proposed_budget_df.to_string(index=False, max_rows=20)}"
    
    prompt = f"""You are an AI-powered Budget Preparation Assistant. 

Context:
{context}

This budget was prepared by proportionally allocating a lump sum of ${total_funding:,.2f} based on prior year spending patterns for {budget_period}.

Please provide:
1. **Budget narrative**: Brief summary of the allocation approach and key highlights
2. **Recommendations**: Suggest any adjustments or areas that may need attention
3. **Key observations**: Identify significant allocations or areas requiring attention

Keep your response concise and actionable."""
    
    ai_recommendations = ""
    try:
        api_key = os.environ.get('GOOGLE_API_KEY')
        if api_key:
            model_obj = genai.GenerativeModel('models/gemini-2.5-flash')
            response = model_obj.generate_content(prompt)
            ai_recommendations = response.text
    except Exception as e:
        ai_recommendations = f"AI recommendations unavailable: {str(e)}"
    
    # Save to MongoDB
    line_items = []
    for _, row in proposed_budget_df[proposed_budget_df['GL Account details'] != 'TOTAL'].iterrows():
        line_items.append({
            'gl_account': row['GL Account details'],
            'prior_year_budget': float(row['Prior Year Budget']),
            'actual_expenses': float(row['Actual Expenses']),
            'carryover': float(row['Carryover']),
            'proposed_budget': float(row['Proposed Budget']),
            'reclass_budget': float(row['Reclass Budget']),
            'final_proposed_budget': float(row['Final Proposed Budget'])
        })
    
    total_row = proposed_budget_df[proposed_budget_df['GL Account details'] == 'TOTAL'].iloc[0]
    
    budget_doc = {
        'client_name': client_name,
        'prepared_by': user_name,
        'budget_period': budget_period,
        'method': 'lump_sum',
        'total_funding': total_funding,
        'inflation_rate': inflation_rate,
        'status': 'draft',
        'created_at': datetime.utcnow(),
        'updated_at': datetime.utcnow(),
        'line_items': line_items,
        'totals': {
            'prior_year_budget': float(total_row['Prior Year Budget']),
            'actual_expenses': float(total_row['Actual Expenses']),
            'carryover': float(total_row['Carryover']),
            'proposed_budget': float(total_row['Proposed Budget']),
            'final_proposed_budget': float(total_row['Final Proposed Budget'])
        },
        'ai_recommendations': ai_recommendations,
        'prior_budget_file_id': prior_budget_file_id
    }
    
    result = budgets_collection.insert_one(budget_doc)
    budget_id = str(result.inserted_id)
    
    session['current_budget_id'] = budget_id
    
    log_audit(budget_id, 'created', user_name)
    
    total_budget = float(proposed_budget_df[proposed_budget_df['GL Account details'] != 'TOTAL']['Proposed Budget'].sum())
    
    return jsonify({
        'success': True,
        'budget_id': budget_id,
        'data': df_to_json_safe(proposed_budget_df),
        'columns': proposed_budget_df.columns.tolist(),
        'recommendations': ai_recommendations,

        'totalBudget': total_budget,
        'budgetPeriod': budget_period
    })



@app.route('/update-budget-data', methods=['POST'])
def update_budget_data():
    """Update budget data with reclass changes"""
    data = request.json or {}
    budget_data = data.get('budgetData', [])
    budget_id = session.get('current_budget_id')
    
    if not budget_id:
        return jsonify({'error': 'No active budget to update'}), 400
    
    try:
        df = pd.DataFrame(budget_data)
        
        non_total_mask = df['GL Account details'] != 'TOTAL'
        
        df.loc[non_total_mask, 'Reclass Budget'] = pd.to_numeric(df.loc[non_total_mask, 'Reclass Budget'], errors='coerce').fillna(0)
        
        reclass_total = df.loc[non_total_mask, 'Reclass Budget'].sum()
        if abs(reclass_total) > 0.01:
            return jsonify({
                'error': f'Reclass Budget must sum to zero. Current total: ${reclass_total:.2f}'
            }), 400
        
        df['Final Proposed Budget'] = df['Proposed Budget'] + df['Reclass Budget']
        
        # Update MongoDB
        line_items = []
        for _, row in df[df['GL Account details'] != 'TOTAL'].iterrows():
            line_items.append({
                'gl_account': row['GL Account details'],
                'prior_year_budget': float(row['Prior Year Budget']),
                'actual_expenses': float(row['Actual Expenses']),
                'carryover': float(row['Carryover']),
                'proposed_budget': float(row['Proposed Budget']),
                'reclass_budget': float(row['Reclass Budget']),
                'final_proposed_budget': float(row['Final Proposed Budget'])
            })
        
        total_row = df[df['GL Account details'] == 'TOTAL'].iloc[0]
        
        budgets_collection.update_one(
            {'_id': ObjectId(budget_id)},
            {
                '$set': {
                    'line_items': line_items,
                    'totals': {
                        'prior_year_budget': float(total_row['Prior Year Budget']),
                        'actual_expenses': float(total_row['Actual Expenses']),
                        'carryover': float(total_row['Carryover']),
                        'proposed_budget': float(total_row['Proposed Budget']),
                        'final_proposed_budget': float(total_row['Final Proposed Budget'])
                    },
                    'updated_at': datetime.utcnow()
                }
            }
        )
        
        budget_doc = budgets_collection.find_one({'_id': ObjectId(budget_id)})
        log_audit(budget_id, 'updated', budget_doc.get('prepared_by', 'Unknown'), {'reclass_changes': True})
        
        return jsonify({
            'success': True,
            'message': 'Budget data updated successfully'
        })
    except Exception as e:
        return jsonify({'error': f'Failed to update budget: {str(e)}'}), 400

@app.route('/download-budget/<budget_id>')
def download_budget(budget_id):
    try:
        budget_doc = budgets_collection.find_one({'_id': ObjectId(budget_id)})
        
        if not budget_doc:
            return jsonify({'error': 'Budget not found'}), 404
        
        # Reconstruct DataFrame
        line_items = budget_doc.get('line_items', [])
        df_data = []
        
        for item in line_items:
            df_data.append({
                'GL Account details': item['gl_account'],
                'Prior Year Budget': item['prior_year_budget'],
                'Actual Expenses': item['actual_expenses'],
                'Carryover': item['carryover'],
                'Proposed Budget': item['proposed_budget'],
                'Reclass Budget': item['reclass_budget'],
                'Final Proposed Budget': item['final_proposed_budget']
            })
        
        totals = budget_doc.get('totals', {})
        df_data.append({
            'GL Account details': 'TOTAL',
            'Prior Year Budget': totals.get('prior_year_budget', 0),
            'Actual Expenses': totals.get('actual_expenses', 0),
            'Carryover': totals.get('carryover', 0),
            'Proposed Budget': totals.get('proposed_budget', 0),
            'Reclass Budget': 0,
            'Final Proposed Budget': totals.get('final_proposed_budget', 0)
        })
        
        df = pd.DataFrame(df_data)
        
        # Generate Excel file in temp location
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        filepath = temp_file.name
        
        export_budget_to_excel(
            df, 
            filepath,
            budget_doc.get('client_name', ''),
            budget_doc.get('prepared_by', ''),
            budget_doc.get('budget_period', '')
        )
        
        # Update status to finalized
        budgets_collection.update_one(
            {'_id': ObjectId(budget_id)},
            {'$set': {'status': 'finalized'}}
        )
        
        log_audit(budget_id, 'downloaded', budget_doc.get('prepared_by', 'Unknown'))
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f"budget_{budget_doc.get('budget_period', 'export')}_{budget_id[:8]}.xlsx"
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/list-budgets', methods=['GET'])
def list_budgets():
    """List all budgets with filtering options"""
    client_name = request.args.get('client_name')
    budget_period = request.args.get('budget_period')
    status = request.args.get('status')
    
    query = {}
    if client_name:
        query['client_name'] = {'$regex': client_name, '$options': 'i'}
    if budget_period:
        query['budget_period'] = {'$regex': budget_period, '$options': 'i'}
    if status:
        query['status'] = status
    
    budgets = list(budgets_collection.find(query).sort('created_at', -1).limit(50))
    
    for budget in budgets:
        budget['_id'] = str(budget['_id'])
        budget['created_at'] = budget['created_at'].isoformat()
        budget['updated_at'] = budget['updated_at'].isoformat()
    
    return jsonify({'success': True, 'budgets': budgets})

@app.route('/clear-session', methods=['POST'])
def clear_session():
    session.clear()
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5016, debug=True)

    