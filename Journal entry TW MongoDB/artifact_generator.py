import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ArtifactGenerator:
    """Generates all audit artifacts and workpapers"""
    
    def __init__(self):
        self.test_results = {
            'passed': 0,
            'exceptions': 0,
            'largest_exception': {}
        }
    
    def generate_all_artifacts(self, population, samples, parameters, output_dir):
        """
        Generate all required audit artifacts
        
        Returns: List of generated file names
        """
        generated_files = []
        
        try:
            # 1. JE Population workbook
            pop_file = self._generate_population_workbook(population, output_dir)
            if pop_file:
                generated_files.append(pop_file)
            
            # 2. Sample Selection workbook
            sample_file = self._generate_sample_selection_workbook(samples, output_dir)
            if sample_file:
                generated_files.append(sample_file)
            
            # 3. PBC Request List
            pbc_file = self._generate_pbc_request_list(samples, output_dir)
            if pbc_file:
                generated_files.append(pbc_file)
            
            # 4. Test Workpaper
            test_file = self._generate_test_workpaper(samples, output_dir)
            if test_file:
                generated_files.append(test_file)
            
            # 5. Exceptions Log
            exceptions_file = self._generate_exceptions_log(samples, output_dir)
            if exceptions_file:
                generated_files.append(exceptions_file)
            
            # 6. Proposed AJEs
            aje_file = self._generate_proposed_ajes(samples, output_dir)
            if aje_file:
                generated_files.append(aje_file)
            
            logger.info(f"Generated {len(generated_files)} artifact files")
            
        except Exception as e:
            logger.error(f"Error generating artifacts: {str(e)}", exc_info=True)
        
        return generated_files
    
    def _generate_population_workbook(self, population, output_dir):
        """Generate JE_Population.xlsx"""
        try:
            filename = 'JE_Population.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            # Select key columns for population summary
            columns_to_include = [
                'je_id', 'date', 'doc_no', 'account', 'description', 
                'debit', 'credit', 'net', 'user', 'risk_score', 'risk_category'
            ]
            
            export_df = population[[col for col in columns_to_include if col in population.columns]].copy()
            
            # Create workbook with formatting
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "JE Population"
            
            # Add headers
            for r in dataframe_to_rows(export_df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            self._format_excel_headers(ws)
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            self._add_population_summary(summary_ws, population)
            
            wb.save(filepath)
            logger.info(f"Generated population workbook: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating population workbook: {str(e)}", exc_info=True)
            return None
    
    def _generate_sample_selection_workbook(self, samples, output_dir):
        """Generate JE_Sample_Selection.xlsx"""
        try:
            filename = 'JE_Sample_Selection.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            if samples.empty:
                logger.warning("No samples to export")
                return None
            
            # Select columns for sample selection
            columns_to_include = [
                'sample_id', 'je_id', 'date', 'account', 'description',
                'net', 'risk_category', 'selection_method', 'selection_rationale'
            ]
            
            export_df = samples[[col for col in columns_to_include if col in samples.columns]].copy()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sample Selection"
            
            for r in dataframe_to_rows(export_df, index=False, header=True):
                ws.append(r)
            
            self._format_excel_headers(ws)
            wb.save(filepath)
            
            logger.info(f"Generated sample selection workbook: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating sample selection workbook: {str(e)}", exc_info=True)
            return None
    
    def _generate_pbc_request_list(self, samples, output_dir):
        """Generate JE_Request_List.xlsx"""
        try:
            filename = 'JE_Request_List.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            if samples.empty:
                return None
            
            # Create PBC request template
            pbc_data = []
            for _, row in samples.iterrows():
                pbc_data.append({
                    'Sample_ID': row.get('sample_id', ''),
                    'JE_ID': row.get('je_id', ''),
                    'Date': row.get('date', ''),
                    'Amount': row.get('net', 0),
                    'Description': row.get('description', ''),
                    'Supporting_Documents_Requested': 'Invoice, contract, memo, or other supporting documentation',
                    'Approval_Evidence_Requested': 'Email approval, signed form, or system approval log',
                    'Business_Justification_Required': 'Business reason and necessity for the journal entry',
                    'Reversal_Details_Required': 'If reversing entry, provide original entry details',
                    'Status': 'Pending',
                    'Date_Requested': datetime.now().strftime('%Y-%m-%d'),
                    'Date_Received': '',
                    'Comments': ''
                })
            
            pbc_df = pd.DataFrame(pbc_data)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PBC Requests"
            
            for r in dataframe_to_rows(pbc_df, index=False, header=True):
                ws.append(r)
            
            self._format_excel_headers(ws)
            wb.save(filepath)
            
            logger.info(f"Generated PBC request list: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating PBC request list: {str(e)}", exc_info=True)
            return None
    
    def _generate_test_workpaper(self, samples, output_dir):
        """Generate JE_Test_Workpaper.xlsx"""
        try:
            filename = 'JE_Test_Workpaper.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            if samples.empty:
                return None
            
            # Create test workpaper template
            test_data = []
            for _, row in samples.iterrows():
                test_data.append({
                    'Sample_ID': row.get('sample_id', ''),
                    'JE_ID': row.get('je_id', ''),
                    'Date': row.get('date', ''),
                    'Amount': row.get('net', 0),
                    'Description': row.get('description', ''),
                    'Test_1_Necessary_Reasonable': '',  # Pass/Fail/NA
                    'Test_1_Comments': '',
                    'Test_2_Supporting_Documentation': '',
                    'Test_2_Comments': '',
                    'Test_3_Amount_Agreement': '',
                    'Test_3_Comments': '',
                    'Test_4_Correct_Period': '',
                    'Test_4_Comments': '',
                    'Test_5_Reviewed_Approved': '',
                    'Test_5_Comments': '',
                    'Overall_Result': '',
                    'Exception_Amount': 0,
                    'Exception_Nature': '',
                    'Proposed_AJE_Required': '',
                    'Tested_By': '',
                    'Test_Date': '',
                    'Reviewed_By': '',
                    'Review_Date': ''
                })
            
            test_df = pd.DataFrame(test_data)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "JE Test Results"
            
            for r in dataframe_to_rows(test_df, index=False, header=True):
                ws.append(r)
            
            self._format_excel_headers(ws)
            
            # Add instructions sheet
            instructions_ws = wb.create_sheet("Test Instructions")
            self._add_test_instructions(instructions_ws)
            
            wb.save(filepath)
            
            logger.info(f"Generated test workpaper: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating test workpaper: {str(e)}", exc_info=True)
            return None
    
    def _generate_exceptions_log(self, samples, output_dir):
        """Generate JE_Exceptions_Log.xlsx"""
        try:
            filename = 'JE_Exceptions_Log.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            # Create empty exceptions log template
            exceptions_data = [{
                'Exception_ID': '',
                'Sample_ID': '',
                'JE_ID': '',
                'Exception_Type': '',
                'Exception_Description': '',
                'Amount_Impact': 0,
                'Account_Affected': '',
                'Risk_Level': '',
                'Root_Cause': '',
                'Management_Response': '',
                'Proposed_AJE_Ref': '',
                'Status': 'Open',
                'Follow_up_Required': '',
                'Date_Identified': datetime.now().strftime('%Y-%m-%d'),
                'Date_Resolved': '',
                'Resolved_By': ''
            }]
            
            exceptions_df = pd.DataFrame(exceptions_data)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Exceptions Log"
            
            for r in dataframe_to_rows(exceptions_df, index=False, header=True):
                ws.append(r)
            
            self._format_excel_headers(ws)
            wb.save(filepath)
            
            logger.info(f"Generated exceptions log: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating exceptions log: {str(e)}", exc_info=True)
            return None
    
    def _generate_proposed_ajes(self, samples, output_dir):
        """Generate Proposed_AJEs.xlsx"""
        try:
            filename = 'Proposed_AJEs.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            # Create empty AJE template
            aje_data = [{
                'AJE_ID': '',
                'Related_JE_ID': '',
                'Related_Sample_ID': '',
                'AJE_Description': '',
                'Account_Number': '',
                'Account_Name': '',
                'Debit_Amount': 0,
                'Credit_Amount': 0,
                'Supporting_Reference': '',
                'Rationale': '',
                'Impact_Assessment': '',
                'Management_Agreement': '',
                'Status': 'Proposed',
                'Date_Proposed': datetime.now().strftime('%Y-%m-%d'),
                'Proposed_By': '',
                'Date_Reviewed': '',
                'Reviewed_By': ''
            }]
            
            aje_df = pd.DataFrame(aje_data)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Proposed AJEs"
            
            for r in dataframe_to_rows(aje_df, index=False, header=True):
                ws.append(r)
            
            self._format_excel_headers(ws)
            
            # Add summary sheet
            summary_ws = wb.create_sheet("AJE Summary")
            self._add_aje_summary(summary_ws)
            
            wb.save(filepath)
            
            logger.info(f"Generated proposed AJEs: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating proposed AJEs: {str(e)}", exc_info=True)
            return None
    
    def _format_excel_headers(self, ws):
        """Apply formatting to Excel headers"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format first row as headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_population_summary(self, ws, population):
        """Add population summary to worksheet"""
        summary_data = [
            ['Population Summary', ''],
            ['Total Journal Entries', len(population)],
            ['Risk Categories:', ''],
            ['  Critical Risk', len(population[population['risk_category'] == 'Critical'])],
            ['  High Risk', len(population[population['risk_category'] == 'High Risk'])],
            ['  Medium Risk', len(population[population['risk_category'] == 'Medium Risk'])],
            ['  Low Risk', len(population[population['risk_category'] == 'Low Risk'])],
            ['', ''],
            ['Amount Summary:', ''],
            ['  Total Net Amount', population['net'].sum() if 'net' in population.columns else 0],
            ['  Average Entry', population['net'].mean() if 'net' in population.columns else 0],
            ['  Largest Entry', population['net'].max() if 'net' in population.columns else 0],
            ['  Smallest Entry', population['net'].min() if 'net' in population.columns else 0]
        ]
        
        for row_data in summary_data:
            ws.append(row_data)
    
    def _add_test_instructions(self, ws):
        """Add testing instructions to worksheet"""
        instructions = [
            ['Journal Entry Testing Instructions', ''],
            ['', ''],
            ['Test 1: Necessary and Reasonable', 'Evaluate if the JE is necessary and conforms to accounting principles'],
            ['Test 2: Supporting Documentation', 'Verify adequate supporting documentation exists'],
            ['Test 3: Amount Agreement', 'Ensure amounts agree to supporting documentation'],
            ['Test 4: Correct Period', 'Verify JE relates to the correct fiscal period'],
            ['Test 5: Reviewed and Approved', 'Confirm proper review and approval per policy'],
            ['', ''],
            ['Results:', ''],
            ['Pass', 'Test criterion satisfied with no exceptions'],
            ['Fail', 'Test criterion not satisfied - exception noted'],
            ['N/A', 'Test criterion not applicable to this entry'],
        ]
        
        for row_data in instructions:
            ws.append(row_data)
    
    def _add_aje_summary(self, ws):
        """Add AJE summary to worksheet"""
        summary_data = [
            ['Proposed Adjusting Journal Entries Summary', ''],
            ['', ''],
            ['Total Proposed AJEs', '0'],
            ['Total Debit Amount', '0'],
            ['Total Credit Amount', '0'],
            ['', ''],
            ['Status Summary:', ''],
            ['  Proposed', '0'],
            ['  Accepted', '0'],
            ['  Rejected', '0'],
            ['  Under Review', '0']
        ]
        
        for row_data in summary_data:
            ws.append(row_data)
    
    def get_test_results(self):
        """Return current test results for reporting"""
        return self.test_results