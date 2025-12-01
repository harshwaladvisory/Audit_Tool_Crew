import os
import logging
from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import openpyxl
from scraper import CharityStatusScraper
import uuid
import threading
import time
import requests
import json
import mongoengine
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "fallback-secret-key-for-development")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# MongoDB Configuration
MONGODB_URI = os.environ.get("MONGODB_URI", "mongodb://localhost:27017/rrf_status_tracker")

# Connect to MongoDB
try:
    mongoengine.connect(host=MONGODB_URI, alias='default')
    logger.info(f"Successfully connected to MongoDB at {MONGODB_URI.split('@')[1] if '@' in MONGODB_URI else MONGODB_URI}")
except Exception as e:
    logger.error(f"Failed to connect to MongoDB: {str(e)}")
    # Continue anyway - app will work without MongoDB

# Import models after MongoDB connection
try:
    from models import CharityCheck, ProcessingTask
    logger.info("Successfully imported MongoDB models")
except ImportError as e:
    logger.warning(f"Could not import models: {e}. MongoDB features will be disabled.")
    CharityCheck = None
    ProcessingTask = None

# Configuration
UPLOAD_FOLDER = 'uploads'
DOWNLOAD_FOLDER = 'downloads'
ALLOWED_EXTENSIONS = {'xlsx'}

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

# Monday.com Configuration (can be moved to .env)
MONDAY_API_KEY = os.environ.get("MONDAY_API_KEY", "eyJhbGciOiJIUzI1NiJ9.eyJ0aWQiOjUzMDEzMzYyOSwiYWFpIjoxMSwidWlkIjo3NzU2NzA3MSwiaWFkIjoiMjAyNS0wNi0yNFQwNTozODowMy4wMDBaIiwicGVyIjoibWU6d3JpdGUiLCJhY3RpZCI6MzAxMDczODMsInJnbiI6ImFwc2UyIn0.GiRrh9ziAaKHXZAIIjnZOtfwq3_Ozfo11O9CJUvovZg")
MONDAY_BOARD_ID = os.environ.get("MONDAY_BOARD_ID", "2032522534")
MONDAY_COLUMN_ID = os.environ.get("MONDAY_COLUMN_ID", "project_status")

# Global dictionary to track processing status
processing_status = {}

def update_monday_status(item_id, status_text):
    url = "https://api.monday.com/v2"
    
    headers = {
        "Authorization": MONDAY_API_KEY,
        "Content-Type": "application/json"
    }

    query = """
    mutation ($item_id: Int!, $column_values: JSON!) {
        change_multiple_column_values(item_id: $item_id, board_id: %s, column_values: $column_values) {
            id
        }
    }
    """ % MONDAY_BOARD_ID

    variables = {
        "item_id": int(item_id),
        "column_values": json.dumps({
            MONDAY_COLUMN_ID: status_text
        })
    }

    response = requests.post(url, json={'query': query, 'variables': variables}, headers=headers)
    
    if response.status_code != 200:
        logger.error(f"Failed to update Monday.com: {response.text}")
    else:
        logger.info(f"Updated Monday.com item {item_id} with status: {status_text}")


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_excel_file(file_path):
    """Validate that the Excel file has the required EIN Number column"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Check if first row contains headers
        headers = [cell.value for cell in sheet[1]]
        
        # Look for EIN Number column (case insensitive)
        ein_column = None
        for i, header in enumerate(headers):
            if header and 'ein' in str(header).lower() and 'number' in str(header).lower():
                ein_column = i + 1  # openpyxl uses 1-based indexing
                break
        
        if ein_column is None:
            return False, "Excel file must contain an 'EIN Number' column"
        
        return True, ein_column
        
    except Exception as e:
        logger.error(f"Error validating Excel file: {str(e)}")
        return False, f"Error reading Excel file: {str(e)}"

def find_item_id_by_ein(ein):
    url = "https://api.monday.com/v2"
    headers = {
        "Authorization": MONDAY_API_KEY,
        "Content-Type": "application/json"
    }

    query = """
    query ($board_id: Int!, $column_id: String!, $value: String!) {
      items_by_column_values(board_id: $board_id, column_id: $column_id, column_value: $value) {
        id
        name
      }
    }
    """

    variables = {
        "board_id": int(MONDAY_BOARD_ID),
        "column_id": "name",
        "value": ein
    }

    response = requests.post(url, headers=headers, json={"query": query, "variables": variables})

    try:
        data = response.json()
        logger.info(f"🔎 EIN '{ein}' query response: {data}")
        items = data.get("data", {}).get("items_by_column_values", [])
        if items:
            return items[0]["id"]
        else:
            logger.warning(f"⚠️ No match found in Monday for EIN: {ein}")
            return None
    except Exception as e:
        logger.error(f"❌ Error parsing response for EIN {ein}: {e}")
        return None


def process_excel_file(file_path, task_id):
    """Process the Excel file and update status"""
    try:
        processing_status[task_id] = {
            'status': 'processing',
            'progress': 0,
            'total': 0,
            'current_ein': '',
            'message': 'Starting processing...',
            'completed': False,
            'error': None,
            'output_file': None
        }
        
        # Create ProcessingTask in MongoDB
        if ProcessingTask:
            db_task = ProcessingTask(
                task_id=task_id,
                file_name=os.path.basename(file_path),
                status='processing'
            )
            db_task.save()
        
        # Load workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Find EIN column
        headers = [cell.value for cell in sheet[1]]
        ein_column = None

        for i, header in enumerate(headers):
            header_lower = str(header).lower() if header else ""
            if 'ein' in header_lower and 'number' in header_lower:
                ein_column = i + 1
                break

        if ein_column is None:
            processing_status[task_id]['error'] = "EIN Number column not found"
            if ProcessingTask:
                db_task.status = 'error'
                db_task.error_message = "EIN Number column not found"
                db_task.save()
            return
        
        # Add new column header
        rrf_column = len(headers) + 1
        sheet.cell(row=1, column=rrf_column, value='RRF-1 Status')
        
        # Get all EINs to process
        eins_to_process = []
        for row in range(2, sheet.max_row + 1):
            ein_cell = sheet.cell(row=row, column=ein_column)
            if ein_cell.value:
                eins_to_process.append((row, str(ein_cell.value).strip()))
        
        total_eins = len(eins_to_process)
        processing_status[task_id]['total'] = total_eins
        
        # Update MongoDB task
        if ProcessingTask:
            db_task.total_eins = total_eins
            db_task.save()
        
        if total_eins == 0:
            processing_status[task_id]['error'] = "No EIN numbers found in the file"
            if ProcessingTask:
                db_task.status = 'error'
                db_task.error_message = "No EIN numbers found"
                db_task.save()
            return
        
        # Initialize scraper
        scraper = CharityStatusScraper()
        
        try:
            # Process each EIN
            for i, (row, ein) in enumerate(eins_to_process):
                processing_status[task_id]['progress'] = i
                processing_status[task_id]['current_ein'] = ein
                processing_status[task_id]['message'] = f'Processing EIN {ein} ({i+1}/{total_eins})'
                
                logger.info(f"Processing EIN: {ein}")
                
                # Clean EIN (digits only)
                clean_ein = ''.join(filter(str.isdigit, ein))
                
                if len(clean_ein) != 9:
                    status = "Invalid EIN Format"
                else:
                    try:
                        status = scraper.get_charity_status(clean_ein)
                    except Exception as e:
                        logger.error(f"Error processing EIN {clean_ein}: {str(e)}")
                        status = f"Error: {str(e)}"
                
                # Update Excel file
                sheet.cell(row=row, column=rrf_column, value=status)

                # Get Monday.com item_id
                item_id = find_item_id_by_ein(clean_ein)
                monday_updated = False

                if item_id and item_id.isdigit():
                    try:
                        update_monday_status(item_id=int(item_id), status_text=status)
                        monday_updated = True
                    except Exception as e:
                        logger.error(f"Failed to update Monday.com for item {item_id}: {str(e)}")
                else:
                    logger.warning(f"Invalid or missing item ID in row {row}")
                
                # Save to MongoDB
                if CharityCheck:
                    try:
                        charity_check = CharityCheck(
                            ein=clean_ein,
                            status=status,
                            check_date=datetime.utcnow(),
                            monday_item_id=str(item_id) if item_id else None,
                            monday_updated=monday_updated,
                            file_name=os.path.basename(file_path),
                            task_id=task_id
                        )
                        charity_check.save()
                        logger.info(f"Saved charity check to MongoDB: {clean_ein} - {status}")
                    except Exception as e:
                        logger.error(f"Error saving to MongoDB: {e}")
                
                # Update MongoDB task progress
                if ProcessingTask:
                    db_task.processed_eins = i + 1
                    db_task.save()
                
                # Small delay to avoid overwhelming the website
                time.sleep(1)
            
            # Save updated file
            output_filename = f"clients_updated_{task_id}.xlsx"
            output_path = os.path.join(DOWNLOAD_FOLDER, output_filename)
            workbook.save(output_path)
            
            processing_status[task_id]['status'] = 'completed'
            processing_status[task_id]['progress'] = total_eins
            processing_status[task_id]['message'] = 'Processing completed successfully!'
            processing_status[task_id]['completed'] = True
            processing_status[task_id]['output_file'] = output_filename
            
            # Update MongoDB task
            if ProcessingTask:
                db_task.status = 'completed'
                db_task.completed_at = datetime.utcnow()
                db_task.save()
            
        finally:
            scraper.close()
            
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        processing_status[task_id]['error'] = str(e)
        processing_status[task_id]['status'] = 'error'
        
        # Update MongoDB task
        if ProcessingTask:
            try:
                db_task.status = 'error'
                db_task.error_message = str(e)
                db_task.completed_at = datetime.utcnow()
                db_task.save()
            except:
                pass

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file selected'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        task_id = str(uuid.uuid4())
        file_path = os.path.join(UPLOAD_FOLDER, f"{task_id}_{filename}")
        file.save(file_path)
        
        # Validate file structure
        is_valid, result = validate_excel_file(file_path)
        if not is_valid:
            os.remove(file_path)
            return jsonify({'error': result}), 400
        
        # Start processing in background thread
        thread = threading.Thread(target=process_excel_file, args=(file_path, task_id))
        thread.daemon = True
        thread.start()
        
        return jsonify({'task_id': task_id})
    
    return jsonify({'error': 'Invalid file format. Please upload an .xlsx file'}), 400

@app.route('/progress/<task_id>')
def get_progress(task_id):
    if task_id not in processing_status:
        return jsonify({'error': 'Task not found'}), 404
    
    return jsonify(processing_status[task_id])

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(DOWNLOAD_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name='clients_updated.xlsx')
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}")
        return jsonify({'error': 'Error downloading file'}), 500

@app.route('/progress_page/<task_id>')
def progress_page(task_id):
    if task_id not in processing_status:
        return redirect(url_for('index'))
    return render_template('progress.html', task_id=task_id)

# API endpoint to get recent checks from MongoDB
@app.route('/api/recent-checks')
def get_recent_checks():
    if not CharityCheck:
        return jsonify({'error': 'MongoDB not configured'}), 503
    
    try:
        limit = request.args.get('limit', 50, type=int)
        checks = CharityCheck.objects().order_by('-check_date').limit(limit)
        
        results = []
        for check in checks:
            results.append({
                'ein': check.ein,
                'status': check.status,
                'check_date': check.check_date.isoformat(),
                'monday_updated': check.monday_updated
            })
        
        return jsonify({'checks': results})
    except Exception as e:
        logger.error(f"Error fetching recent checks: {e}")
        return jsonify({'error': str(e)}), 500

# API endpoint to get task history from MongoDB
@app.route('/api/tasks')
def get_tasks():
    if not ProcessingTask:
        return jsonify({'error': 'MongoDB not configured'}), 503
    
    try:
        limit = request.args.get('limit', 20, type=int)
        tasks = ProcessingTask.objects().order_by('-started_at').limit(limit)
        
        results = []
        for task in tasks:
            results.append({
                'task_id': task.task_id,
                'file_name': task.file_name,
                'status': task.status,
                'total_eins': task.total_eins,
                'processed_eins': task.processed_eins,
                'started_at': task.started_at.isoformat(),
                'completed_at': task.completed_at.isoformat() if task.completed_at else None
            })
        
        return jsonify({'tasks': results})
    except Exception as e:
        logger.error(f"Error fetching tasks: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8515, debug=True)