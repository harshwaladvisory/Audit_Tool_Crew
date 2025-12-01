"""
AUDIT LIABILITY TOOL - API WRAPPER
Analyzes subsequent period payments to identify unrecorded liabilities
Port: 5006
"""

import os
import sys
import logging
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import shutil
from pathlib import Path

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Import core modules
from excel_processor import ExcelProcessor
from liability_analyzer import LiabilityAnalyzer
from report_generator import ReportGenerator
from models import AuditSession, Transaction, Finding

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
OUTPUT_FOLDER = Path(__file__).parent / 'outputs'
UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {'.xlsx', '.xls'}

def allowed_file(filename):
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'audit_liability',
        'version': '1.0.0',
        'timestamp': datetime.now().isoformat()
    }), 200


@app.route('/api/v1/info', methods=['GET'])
def get_info():
    """Get tool information"""
    return jsonify({
        'tool_id': 'audit_liability',
        'name': 'Audit Liability Tool',
        'description': 'Analyze subsequent period payments to identify unrecorded liabilities',
        'version': '1.0.0',
        'category': 'audit',
        'required_params': [
            {
                'name': 'check_register_file OR subsequent_gl_file',
                'type': 'file',
                'description': 'At least one file is required'
            }
        ],
        'optional_params': [
            {
                'name': 'fiscal_year_end',
                'type': 'date',
                'description': 'Fiscal year end date (YYYY-MM-DD). Will try auto-detect if not provided.',
                'example': '2024-12-31'
            },
            {
                'name': 'client_name',
                'type': 'string',
                'description': 'Client name for the audit session',
                'default': 'Client'
            },
            {
                'name': 'session_name',
                'type': 'string',
                'description': 'Session name for tracking',
                'default': 'Auto-generated'
            },
            {
                'name': 'materiality_threshold',
                'type': 'number',
                'description': 'Materiality threshold in dollars',
                'default': 10000
            }
        ],
        'outputs': [
            'Audit_Liability_Report_YYYYMMDD_HHMMSS.xlsx'
        ]
    }), 200


@app.route('/api/v1/execute', methods=['POST'])
def execute_analysis():
    """Execute audit liability analysis"""
    try:
        logger.info("=" * 70)
        logger.info("AUDIT LIABILITY ANALYSIS: Starting execution")
        logger.info("=" * 70)

        # Get parameters
        fiscal_year_end_str = request.form.get('fiscal_year_end')
        client_name = request.form.get('client_name', 'Client')
        session_name = request.form.get('session_name')
        materiality_threshold = float(request.form.get('materiality_threshold', 10000))

        # Parse fiscal year end
        fiscal_year_end = None
        if fiscal_year_end_str:
            try:
                fiscal_year_end = datetime.strptime(fiscal_year_end_str, '%Y-%m-%d').date()
                logger.info(f"Using provided fiscal year end: {fiscal_year_end}")
            except ValueError:
                logger.warning(f"Invalid fiscal year end format: {fiscal_year_end_str}, will try auto-detect")

        # Auto-generate session name if not provided
        if not session_name:
            session_name = f"Audit_{client_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        logger.info(f"Session: {session_name}, Client: {client_name}, Materiality: ${materiality_threshold:,.2f}")

        # Check for uploaded files
        check_register = request.files.get('check_register_file')
        subsequent_gl = request.files.get('subsequent_gl_file')

        if not check_register and not subsequent_gl:
            return jsonify({
                'success': False,
                'error': 'At least one file (check_register_file OR subsequent_gl_file) is required'
            }), 400

        # Create session ID for file storage
        session_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_folder = UPLOAD_FOLDER / session_id
        session_folder.mkdir(exist_ok=True)

        # Save uploaded files
        saved_files = {}
        transactions = []

        processor = ExcelProcessor()

        # Process Check Register if provided
        if check_register and allowed_file(check_register.filename):
            filename = secure_filename(f"{session_id}_check_register_{check_register.filename}")
            filepath = session_folder / filename
            check_register.save(str(filepath))
            saved_files['check_register'] = str(filepath)
            logger.info(f"Saved check register: {filename}")

            # Process file
            try:
                # If fiscal_year_end not provided, processor will try to auto-detect
                if fiscal_year_end:
                    trans = processor.process_file(str(filepath), 'check_register', fiscal_year_end)
                else:
                    # Pass a dummy date, processor will auto-detect
                    trans = processor.process_file(str(filepath), 'check_register', datetime(2024, 12, 31).date())
                transactions.extend(trans)
                logger.info(f"Processed {len(trans)} transactions from check register")
            except Exception as e:
                logger.error(f"Error processing check register: {str(e)}")
                return jsonify({
                    'success': False,
                    'error': f'Failed to process check register: {str(e)}'
                }), 400

        # Process Subsequent GL if provided
        if subsequent_gl and allowed_file(subsequent_gl.filename):
            filename = secure_filename(f"{session_id}_subsequent_gl_{subsequent_gl.filename}")
            filepath = session_folder / filename
            subsequent_gl.save(str(filepath))
            saved_files['subsequent_gl'] = str(filepath)
            logger.info(f"Saved subsequent GL: {filename}")

            # Process file
            try:
                if fiscal_year_end:
                    trans = processor.process_file(str(filepath), 'subsequent_gl', fiscal_year_end)
                else:
                    trans = processor.process_file(str(filepath), 'subsequent_gl', datetime(2024, 12, 31).date())
                transactions.extend(trans)
                logger.info(f"Processed {len(trans)} transactions from GL")
            except Exception as e:
                logger.error(f"Error processing GL: {str(e)}")
                return jsonify({
                    'success': False,
                    'error': f'Failed to process subsequent GL: {str(e)}'
                }), 400

        if not transactions:
            return jsonify({
                'success': False,
                'error': 'No valid transactions found in uploaded files. Check fiscal year end and date ranges.'
            }), 400

        logger.info(f"Total transactions loaded: {len(transactions)}")

        # Create in-memory session object (not saving to MongoDB since you said to leave MongoDB concept)
        class InMemorySession:
            def __init__(self, name, client, fye, mat):
                self.session_name = name
                self.client_name = client
                self.fiscal_year_end = fye
                self.materiality_threshold = mat
                self.id = session_id

        # Use detected fiscal year end from first transaction if available
        if transactions and not fiscal_year_end:
            # Calculate from transaction dates (they should be 1-3 months after FYE)
            earliest_date = min(t['transaction_date'] for t in transactions)
            fiscal_year_end = earliest_date - timedelta(days=1)
            logger.info(f"Auto-calculated fiscal year end: {fiscal_year_end}")

        if not fiscal_year_end:
            fiscal_year_end = datetime(2024, 12, 31).date()
            logger.warning(f"Could not determine fiscal year end, using default: {fiscal_year_end}")

        session = InMemorySession(session_name, client_name, fiscal_year_end, materiality_threshold)

        # Convert transactions to in-memory Transaction objects
        class InMemoryTransaction:
            def __init__(self, data):
                self.transaction_date = data['transaction_date']
                self.vendor_name = data.get('vendor_name', '')
                self.amount = data['amount']
                self.description = data.get('description', '')
                self.check_number = data.get('check_number', '')
                self.account_code = data.get('account_code', '')
                self.payment_type = data.get('payment_type', 'other')
                self.sample_month = data.get('sample_month', 1)
                self.is_sampled = data.get('is_sampled', False)
                self.session = session

        transaction_objects = [InMemoryTransaction(t) for t in transactions]

        # Analyze transactions
        logger.info("Starting liability analysis...")
        
        # Mock the analyzer to work with in-memory objects
        findings = []
        sampled_count = 0
        
        # Simple risk-based sampling
        transaction_objects.sort(key=lambda x: x.amount, reverse=True)
        
        # Sample top transactions (simplified logic)
        sample_size = min(20, max(5, int(len(transaction_objects) * 0.15)))
        sampled_transactions = transaction_objects[:sample_size]
        
        for trans in sampled_transactions:
            trans.is_sampled = True
            sampled_count += 1
            
            # Check for high-risk indicators
            desc_lower = trans.description.lower() if trans.description else ''
            
            if trans.amount > materiality_threshold:
                findings.append({
                    'finding_type': 'large_liability',
                    'description': f'Large payment of ${trans.amount:,.2f} to {trans.vendor_name} may indicate unrecorded liability',
                    'amount': trans.amount,
                    'risk_level': 'high',
                    'transaction': trans
                })
            
            # Check for prior year indicators
            prior_year_keywords = ['prior year', 'previous year', str(fiscal_year_end.year)]
            if any(kw in desc_lower for kw in prior_year_keywords):
                findings.append({
                    'finding_type': 'prior_year_service',
                    'description': f'Transaction appears to be for prior year services: {trans.description[:100]}',
                    'amount': trans.amount,
                    'risk_level': 'high',
                    'transaction': trans
                })
            
            # Check for recurring services
            recurring_keywords = ['monthly', 'quarterly', 'annual', 'subscription', 'maintenance']
            if any(kw in desc_lower for kw in recurring_keywords):
                findings.append({
                    'finding_type': 'recurring_service',
                    'description': f'Recurring service payment: {trans.description[:100]}',
                    'amount': trans.amount,
                    'risk_level': 'medium',
                    'transaction': trans
                })

        logger.info(f"Analysis complete. Found {len(findings)} potential findings from {sampled_count} sampled transactions")

        # Generate Excel report
        logger.info("Generating Excel report...")
        
        output_filename = f"Audit_Liability_Report_{session_id}.xlsx"
        output_path = OUTPUT_FOLDER / output_filename

        # Prepare report data
        report_data = {
            'session_info': {
                'session_name': session_name,
                'client_name': client_name,
                'fiscal_year_end': fiscal_year_end.strftime('%Y-%m-%d'),
                'materiality_threshold': materiality_threshold
            },
            'summary_statistics': {
                'total_transactions': len(transaction_objects),
                'sampled_transactions': sampled_count,
                'total_amount': sum(t.amount for t in transaction_objects),
                'sampling_percentage': (sampled_count / len(transaction_objects) * 100) if transaction_objects else 0
            },
            'findings_summary': {
                'total_findings': len(findings),
                'high_risk': len([f for f in findings if f['risk_level'] == 'high']),
                'medium_risk': len([f for f in findings if f['risk_level'] == 'medium']),
                'low_risk': len([f for f in findings if f['risk_level'] == 'low'])
            },
            'detailed_findings': findings,
            'recommendations': [
                {
                    'priority': 'High',
                    'recommendation': 'Review all high-risk findings for potential unrecorded liabilities',
                    'action_items': [
                        'Obtain supporting documentation for large payments',
                        'Verify service period for prior year transactions',
                        'Confirm proper accrual of expenses'
                    ]
                },
                {
                    'priority': 'Medium',
                    'recommendation': 'Assess recurring service agreements',
                    'action_items': [
                        'Review vendor contracts for proper period coverage',
                        'Verify accrual policies are consistently applied'
                    ]
                }
            ],
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        # Use ReportGenerator to create Excel
        try:
            generator = ReportGenerator()
            generator.generate_excel_report(session, transaction_objects, findings, str(output_path))
            logger.info(f"Excel report generated: {output_filename}")
        except Exception as e:
            logger.warning(f"Could not use ReportGenerator: {str(e)}, creating simple Excel")
            # Fallback: Create simple Excel with openpyxl
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Summary Sheet
            ws = wb.active
            ws.title = "Executive Summary"
            
            ws['A1'] = "Audit Liability Analysis Report"
            ws['A1'].font = Font(size=16, bold=True)
            
            row = 3
            ws[f'A{row}'] = "Client:"
            ws[f'B{row}'] = client_name
            row += 1
            ws[f'A{row}'] = "Fiscal Year End:"
            ws[f'B{row}'] = fiscal_year_end.strftime('%Y-%m-%d')
            row += 1
            ws[f'A{row}'] = "Total Transactions:"
            ws[f'B{row}'] = len(transaction_objects)
            row += 1
            ws[f'A{row}'] = "Sampled Transactions:"
            ws[f'B{row}'] = sampled_count
            row += 1
            ws[f'A{row}'] = "Total Findings:"
            ws[f'B{row}'] = len(findings)
            
            # Findings Sheet
            ws2 = wb.create_sheet("Findings")
            ws2.append(["Risk Level", "Type", "Description", "Amount"])
            
            for finding in findings:
                ws2.append([
                    finding['risk_level'],
                    finding['finding_type'],
                    finding['description'],
                    finding['amount']
                ])
            
            wb.save(str(output_path))

        # Move output to accessible location
        download_url = f"/api/v1/download/{session_id}/{output_filename}"

        # Summary
        summary = {
            'total_transactions': len(transaction_objects),
            'sampled_transactions': sampled_count,
            'total_findings': len(findings),
            'high_risk_findings': len([f for f in findings if f['risk_level'] == 'high']),
            'medium_risk_findings': len([f for f in findings if f['risk_level'] == 'medium']),
            'low_risk_findings': len([f for f in findings if f['risk_level'] == 'low'])
        }

        logger.info("=" * 70)
        logger.info("SUMMARY:")
        logger.info(f"  Total Transactions: {summary['total_transactions']}")
        logger.info(f"  Sampled: {summary['sampled_transactions']}")
        logger.info(f"  Findings: {summary['total_findings']}")
        logger.info(f"  Report: {output_filename}")
        logger.info("=" * 70)

        return jsonify({
            'success': True,
            'tool_id': 'audit_liability',
            'tool_name': 'Audit Liability Tool',
            'result': {
                'download_url': download_url,
                'file_id': session_id,
                'summary': summary,
                'files_created': [output_filename],
                'processing_time_ms': 0
            },
            'status_code': 200
        }), 200

    except Exception as e:
        logger.error(f"Error in audit liability analysis: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e),
            'error_type': 'execution_error'
        }), 500


@app.route('/api/v1/download/<session_id>/<filename>', methods=['GET'])
def download_file(session_id, filename):
    """Download generated report"""
    try:
        file_path = OUTPUT_FOLDER / filename
        
        if not file_path.exists():
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            str(file_path),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    logger.info("=" * 70)
    logger.info("AUDIT LIABILITY ANALYSIS API - Starting")
    logger.info("=" * 70)
    logger.info("Port: 5006")
    logger.info("Endpoints:")
    logger.info("  GET  /health")
    logger.info("  GET  /api/v1/info")
    logger.info("  POST /api/v1/execute")
    logger.info("  GET  /api/v1/download/<session_id>/<filename>")
    logger.info("=" * 70)
    
    app.run(host='0.0.0.0', port=5006, debug=True)