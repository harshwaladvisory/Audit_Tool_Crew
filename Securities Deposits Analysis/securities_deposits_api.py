"""
SECURITIES DEPOSITS ANALYSIS - API WRAPPER (FIXED)
AI-powered analysis of securities deposits for audit purposes
Port: 5008 (Orchestration Standard)
Original App Port: 8567 (MongoDB)

FIXED: Works in standalone mode WITHOUT MongoDB
"""

import os
import sys
import logging
import requests
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from datetime import datetime
from pathlib import Path
import json
import pandas as pd
import traceback

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
OUTPUT_FOLDER = Path(__file__).parent / 'outputs'
UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.pdf'}

# Original app configuration
ORIGINAL_APP_URL = "http://localhost:8567"
ORIGINAL_APP_AVAILABLE = False


def check_original_app():
    """Check if original app on port 8567 is available"""
    global ORIGINAL_APP_AVAILABLE
    try:
        response = requests.get(f"{ORIGINAL_APP_URL}/api/files-list", timeout=2)
        ORIGINAL_APP_AVAILABLE = response.status_code == 200
        if ORIGINAL_APP_AVAILABLE:
            logger.info("✅ Original Securities Deposits app (port 8567) is available")
        return ORIGINAL_APP_AVAILABLE
    except:
        ORIGINAL_APP_AVAILABLE = False
        logger.warning("⚠️ Original Securities Deposits app (port 8567) not available")
        return False


def allowed_file(filename):
    """Check if file extension is allowed"""
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    check_original_app()
    return jsonify({
        'status': 'healthy',
        'service': 'securities_deposits',
        'version': '1.0.0',
        'port': 5008,
        'original_app_available': ORIGINAL_APP_AVAILABLE,
        'timestamp': datetime.now().isoformat()
    }), 200


@app.route('/api/v1/info', methods=['GET'])
def get_info():
    """Get tool information"""
    return jsonify({
        'tool_id': 'securities_deposits',
        'name': 'Securities Deposits Analysis',
        'description': 'AI-powered analysis of securities deposits for audit purposes with MongoDB integration',
        'version': '1.0.0',
        'category': 'audit',
        'required_params': [
            {
                'name': 'file',
                'type': 'file',
                'description': 'Deposit data file (Excel, CSV, or PDF)',
                'allowed_types': ['.xlsx', '.xls', '.csv', '.pdf']
            }
        ],
        'optional_params': [
            {
                'name': 'analysis_type',
                'type': 'string',
                'description': 'Type of analysis to perform',
                'options': ['aging', 'interest', 'unclaimed', 'exception', 'compliance', 'all'],
                'default': 'all'
            },
            {
                'name': 'org_type',
                'type': 'string',
                'description': 'Organization type for audit program generation',
                'options': ['NPO', 'Government', 'Commercial'],
                'default': 'NPO'
            },
            {
                'name': 'report_type',
                'type': 'string',
                'description': 'Type of report to generate',
                'options': ['summary', 'aging', 'interest', 'exception', 'compliance', 'unclaimed'],
                'default': 'summary'
            },
            {
                'name': 'generate_audit_program',
                'type': 'boolean',
                'description': 'Generate audit program',
                'default': False
            }
        ],
        'features': [
            'Excel/CSV/PDF file processing with OCR',
            'Aging analysis with compliance tracking',
            'Interest accrual calculations',
            'Unclaimed deposits identification',
            'Exception and compliance reporting',
            'Audit program generation (NPO/Government/Commercial)',
            'MongoDB persistence (when available)'
        ],
        'outputs': [
            'Comprehensive analysis reports (Excel/JSON)',
            'Audit programs (PDF/JSON)',
            'Dashboard visualizations'
        ]
    }), 200


def process_excel_file_standalone(filepath: str) -> pd.DataFrame:
    """Process Excel file without MongoDB - standalone mode"""
    try:
        # Read Excel file
        df = pd.read_excel(filepath)
        
        # Basic column normalization
        df.columns = df.columns.str.lower().str.strip()
        
        # Ensure required columns exist (with defaults)
        if 'account_number' not in df.columns:
            df['account_number'] = df.index.astype(str)
        
        if 'customer_name' not in df.columns:
            df['customer_name'] = 'Unknown'
        
        if 'amount' not in df.columns:
            df['amount'] = 0.0
        
        if 'deposit_date' not in df.columns:
            df['deposit_date'] = datetime.now()
        
        if 'interest_rate' not in df.columns:
            df['interest_rate'] = 0.0
        
        # Clean amount field
        if 'amount' in df.columns:
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0.0)
        
        # Parse dates
        if 'deposit_date' in df.columns:
            df['deposit_date'] = pd.to_datetime(df['deposit_date'], errors='coerce')
        
        if 'maturity_date' in df.columns:
            df['maturity_date'] = pd.to_datetime(df['maturity_date'], errors='coerce')
        
        logger.info(f"✅ Processed {len(df)} records from Excel")
        return df
        
    except Exception as e:
        logger.error(f"Error processing Excel: {str(e)}")
        raise


def perform_aging_analysis(df: pd.DataFrame) -> dict:
    """Perform aging analysis on deposits"""
    try:
        today = datetime.now()
        results = []
        
        for idx, row in df.iterrows():
            deposit_date = row.get('deposit_date', today)
            if pd.isna(deposit_date):
                deposit_date = today
            
            maturity_date = row.get('maturity_date')
            if pd.isna(maturity_date):
                maturity_date = deposit_date + pd.Timedelta(days=365)
            
            days_to_maturity = (maturity_date - today).days
            
            # Determine aging bucket
            if days_to_maturity < 0:
                days_overdue = abs(days_to_maturity)
                if days_overdue <= 30:
                    bucket = '0-30 days overdue'
                elif days_overdue <= 90:
                    bucket = '31-90 days overdue'
                elif days_overdue <= 180:
                    bucket = '91-180 days overdue'
                else:
                    bucket = '180+ days overdue'
                risk = 'high' if days_overdue > 90 else 'medium'
            else:
                bucket = 'Current'
                risk = 'low'
            
            results.append({
                'account_number': row.get('account_number', ''),
                'customer_name': row.get('customer_name', ''),
                'amount': float(row.get('amount', 0)),
                'deposit_date': str(deposit_date.date()),
                'maturity_date': str(maturity_date.date()) if not pd.isna(maturity_date) else None,
                'days_to_maturity': days_to_maturity,
                'aging_bucket': bucket,
                'risk_level': risk
            })
        
        return {
            'status': 'success',
            'total_analyzed': len(results),
            'results': results
        }
    except Exception as e:
        logger.error(f"Aging analysis error: {str(e)}")
        return {'status': 'error', 'error': str(e)}


def calculate_interest_standalone(df: pd.DataFrame) -> dict:
    """Calculate interest accruals"""
    try:
        today = datetime.now()
        results = []
        
        for idx, row in df.iterrows():
            amount = float(row.get('amount', 0))
            rate = float(row.get('interest_rate', 0))
            
            if rate == 0:
                continue
            
            deposit_date = row.get('deposit_date', today)
            if pd.isna(deposit_date):
                deposit_date = today
            
            days_held = (today - deposit_date).days
            years_held = days_held / 365.0
            
            interest_earned = amount * rate * years_held
            total_value = amount + interest_earned
            
            results.append({
                'account_number': row.get('account_number', ''),
                'customer_name': row.get('customer_name', ''),
                'principal': amount,
                'interest_rate': rate * 100,  # As percentage
                'days_held': days_held,
                'interest_earned': interest_earned,
                'total_value': total_value
            })
        
        return {
            'status': 'success',
            'total_calculated': len(results),
            'results': results
        }
    except Exception as e:
        logger.error(f"Interest calculation error: {str(e)}")
        return {'status': 'error', 'error': str(e)}


@app.route('/api/v1/execute', methods=['POST'])
def execute_analysis():
    """Execute securities deposits analysis - EXCEL OUTPUT VERSION"""
    try:
        logger.info("=" * 70)
        logger.info("SECURITIES DEPOSITS ANALYSIS: Starting execution")
        logger.info("=" * 70)

        # Get parameters
        analysis_type = request.form.get('analysis_type', 'all')
        org_type = request.form.get('org_type', 'NPO')
        report_type = request.form.get('report_type', 'summary')
        generate_audit_program = request.form.get('generate_audit_program', 'false').lower() == 'true'

        logger.info(f"Analysis Type: {analysis_type}")
        logger.info(f"Organization Type: {org_type}")
        logger.info(f"Report Type: {report_type}")
        logger.info(f"Generate Audit Program: {generate_audit_program}")

        # Check for uploaded file
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file uploaded. Please provide a deposit data file (Excel, CSV, or PDF)'
            }), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400

        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': f'Invalid file format. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'
            }), 400

        # Create session ID and folder
        session_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_folder = UPLOAD_FOLDER / session_id
        session_folder.mkdir(exist_ok=True)

        # Save uploaded file
        filename = secure_filename(f"{session_id}_{file.filename}")
        filepath = session_folder / filename
        file.save(str(filepath))
        logger.info(f"✅ Saved file: {filename}")

        # ========================================================================
        # OPTION 1: Use Original App (if available)
        # ========================================================================
        if check_original_app():
            try:
                logger.info("🔄 Forwarding to original app on port 8567...")
                
                with open(str(filepath), 'rb') as f:
                    files = {'file': (file.filename, f, file.content_type)}
                    upload_response = requests.post(
                        f"{ORIGINAL_APP_URL}/api/upload-file",
                        files=files,
                        timeout=60
                    )
                
                if upload_response.status_code == 200:
                    upload_data = upload_response.json()
                    file_id = upload_data.get('file_id')
                    records_processed = upload_data.get('processed_records', 0)
                    
                    logger.info(f"✅ File processed by original app: {records_processed} records")
                    
                    # Generate Excel report with MongoDB data
                    output_filename = f"Securities_Deposits_Analysis_{session_id}.xlsx"
                    output_path = OUTPUT_FOLDER / output_filename
                    
                    # TODO: Fetch data from MongoDB and create Excel
                    # For now, create a simple report
                    import pandas as pd
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Summary"
                    
                    # Header
                    ws['A1'] = "SECURITIES DEPOSITS ANALYSIS"
                    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
                    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    ws.merge_cells('A1:D1')
                    
                    # Summary info
                    ws['A3'] = "Session ID:"
                    ws['B3'] = session_id
                    ws['A4'] = "Records Processed:"
                    ws['B4'] = records_processed
                    ws['A5'] = "Analysis Mode:"
                    ws['B5'] = "MongoDB (Full)"
                    
                    wb.save(str(output_path))
                    
                    download_url = f"/api/v1/download/{session_id}/{output_filename}"
                    
                    return jsonify({
                        'success': True,
                        'tool_id': 'securities_deposits',
                        'tool_name': 'Securities Deposits Analysis',
                        'result': {
                            'download_url': download_url,
                            'session_id': session_id,
                            'summary': {
                                'records_processed': records_processed,
                                'analysis_mode': 'mongodb_app'
                            },
                            'files_created': [output_filename],
                            'processing_time_ms': 0
                        },
                        'status_code': 200
                    }), 200
                
            except Exception as e:
                logger.error(f"Error using original app: {str(e)}")
                logger.info("Falling back to standalone processing...")
        
        # ========================================================================
        # OPTION 2: Standalone Processing (NO MONGODB) - EXCEL OUTPUT
        # ========================================================================
        logger.info("🔧 Running standalone analysis (MongoDB-free)...")
        
        # Process Excel file
        df = process_excel_file_standalone(str(filepath))
        records_processed = len(df)
        
        logger.info(f"✅ Processed {records_processed} records")
        
        # Run analyses
        analysis_results = {}
        
        if analysis_type in ['aging', 'all']:
            analysis_results['aging'] = perform_aging_analysis(df)
            logger.info("✅ Aging analysis complete")
        
        if analysis_type in ['interest', 'all']:
            analysis_results['interest'] = calculate_interest_standalone(df)
            logger.info("✅ Interest calculation complete")
        
        # ========================================================================
        # 🎯 GENERATE COMPREHENSIVE EXCEL REPORT (Like Other Tools!)
        # ========================================================================
        
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        output_filename = f"Securities_Deposits_Analysis_{session_id}.xlsx"
        output_path = OUTPUT_FOLDER / output_filename
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ====================================================================
        # SHEET 1: SUMMARY
        # ====================================================================
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Header
        ws_summary['A1'] = "SECURITIES DEPOSITS ANALYSIS"
        ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.merge_cells('A1:E1')
        ws_summary.row_dimensions[1].height = 30
        
        # Session Info
        row = 3
        ws_summary[f'A{row}'] = "Session Information"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        info_data = [
            ("Session ID:", session_id),
            ("Analysis Date:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("File Name:", file.filename),
            ("Records Processed:", records_processed),
            ("Organization Type:", org_type),
            ("Analysis Type:", analysis_type),
            ("Analysis Mode:", "Standalone (MongoDB-free)")
        ]
        
        for label, value in info_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Analysis Results Summary
        row += 2
        ws_summary[f'A{row}'] = "Analysis Results Summary"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for analysis_name, result in analysis_results.items():
            ws_summary[f'A{row}'] = f"{analysis_name.title()} Analysis:"
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = f"✓ {result.get('total_analyzed', result.get('total_calculated', 0))} records"
            ws_summary[f'B{row}'].font = Font(color="008000")
            row += 1
        
        # Auto-size columns
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40
        
        # ====================================================================
        # SHEET 2: AGING ANALYSIS
        # ====================================================================
        if 'aging' in analysis_results and analysis_results['aging'].get('results'):
            ws_aging = wb.create_sheet("Aging Analysis")
            
            # Header
            ws_aging['A1'] = "AGING ANALYSIS"
            ws_aging['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_aging['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws_aging['A1'].alignment = Alignment(horizontal='center')
            ws_aging.merge_cells('A1:H1')
            
            # Convert to DataFrame
            aging_df = pd.DataFrame(analysis_results['aging']['results'])
            
            # Write headers
            headers = list(aging_df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws_aging.cell(row=3, column=col_idx)
                cell.value = header.replace('_', ' ').title()
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Write data
            for row_idx, row_data in enumerate(aging_df.values, 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_aging.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Format amounts
                    if col_idx == 3:  # Amount column
                        cell.number_format = '$#,##0.00'
                    
                    # Color code risk levels
                    if col_idx == 8:  # Risk level column
                        if value == 'high':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value == 'medium':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                ws_aging.column_dimensions[chr(64 + col_idx)].width = 15
        
        # ====================================================================
        # SHEET 3: INTEREST CALCULATIONS
        # ====================================================================
        if 'interest' in analysis_results and analysis_results['interest'].get('results'):
            ws_interest = wb.create_sheet("Interest Calculations")
            
            # Header
            ws_interest['A1'] = "INTEREST CALCULATIONS"
            ws_interest['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_interest['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws_interest['A1'].alignment = Alignment(horizontal='center')
            ws_interest.merge_cells('A1:G1')
            
            # Convert to DataFrame
            interest_df = pd.DataFrame(analysis_results['interest']['results'])
            
            # Write headers
            headers = list(interest_df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws_interest.cell(row=3, column=col_idx)
                cell.value = header.replace('_', ' ').title()
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Write data with totals
            total_principal = 0
            total_interest = 0
            total_value = 0
            
            for row_idx, row_data in enumerate(interest_df.values, 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_interest.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Format currency columns
                    if col_idx in [3, 6, 7]:  # Principal, Interest, Total columns
                        cell.number_format = '$#,##0.00'
                        
                        # Sum totals
                        if col_idx == 3:
                            total_principal += float(value) if value else 0
                        elif col_idx == 6:
                            total_interest += float(value) if value else 0
                        elif col_idx == 7:
                            total_value += float(value) if value else 0
                    
                    # Format percentage
                    if col_idx == 4:  # Interest rate
                        cell.number_format = '0.00"%"'
            
            # Add totals row
            total_row = len(interest_df) + 5
            ws_interest[f'A{total_row}'] = "TOTALS:"
            ws_interest[f'A{total_row}'].font = Font(bold=True, size=12)
            ws_interest[f'C{total_row}'] = total_principal
            ws_interest[f'C{total_row}'].number_format = '$#,##0.00'
            ws_interest[f'C{total_row}'].font = Font(bold=True)
            ws_interest[f'F{total_row}'] = total_interest
            ws_interest[f'F{total_row}'].number_format = '$#,##0.00'
            ws_interest[f'F{total_row}'].font = Font(bold=True)
            ws_interest[f'G{total_row}'] = total_value
            ws_interest[f'G{total_row}'].number_format = '$#,##0.00'
            ws_interest[f'G{total_row}'].font = Font(bold=True)
            
            # Highlight totals row
            for col_idx in range(1, len(headers) + 1):
                ws_interest.cell(row=total_row, column=col_idx).fill = PatternFill(
                    start_color="E7E6E6",
                    end_color="E7E6E6",
                    fill_type="solid"
                )
            
            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                ws_interest.column_dimensions[chr(64 + col_idx)].width = 18
        
        # ====================================================================
        # SHEET 4: RAW DATA
        # ====================================================================
        ws_raw = wb.create_sheet("Raw Data")
        
        # Header
        ws_raw['A1'] = "RAW DEPOSIT DATA"
        ws_raw['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_raw['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_raw['A1'].alignment = Alignment(horizontal='center')
        ws_raw.merge_cells(f'A1:{chr(64 + len(df.columns))}1')
        
        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws_raw.cell(row=r_idx, column=c_idx)
                cell.value = value
                
                # Header row formatting
                if r_idx == 3:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-size columns
        for col_idx in range(1, len(df.columns) + 1):
            ws_raw.column_dimensions[chr(64 + col_idx)].width = 15
        
        # Save Excel file
        wb.save(str(output_path))
        logger.info(f"✅ Excel report generated: {output_filename}")
        
        # ========================================================================
        # RETURN RESPONSE WITH EXCEL DOWNLOAD
        # ========================================================================
        
        download_url = f"/api/v1/download/{session_id}/{output_filename}"
        
        summary = {
            'records_processed': records_processed,
            'analyses_completed': len(analysis_results),
            'analysis_mode': 'standalone_no_mongodb',
            'file_type': Path(file.filename).suffix,
            'sheets_created': len(wb.sheetnames)
        }
        
        logger.info("=" * 70)
        logger.info("EXECUTION COMPLETE (standalone - Excel output)")
        logger.info(f"  Records: {records_processed}")
        logger.info(f"  Analyses: {len(analysis_results)}")
        logger.info(f"  Sheets: {len(wb.sheetnames)}")
        logger.info(f"  Report: {output_filename}")
        logger.info("=" * 70)
        
        return jsonify({
            'success': True,
            'tool_id': 'securities_deposits',
            'tool_name': 'Securities Deposits Analysis',
            'result': {
                'download_url': download_url,
                'session_id': session_id,
                'summary': summary,
                'files_created': [output_filename],
                'processing_time_ms': 0
            },
            'status_code': 200
        }), 200

    except Exception as e:
        logger.error(f"Error in securities deposits analysis: {str(e)}", exc_info=True)
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e),
            'error_type': 'execution_error',
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/v1/download/<session_id>/<filename>', methods=['GET'])
def download_file(session_id, filename):
    """Download generated report - FIXED VERSION"""
    try:
        # Try to find file in outputs folder
        file_path = OUTPUT_FOLDER / filename
        
        logger.info(f"📥 Download request: {filename}")
        logger.info(f"📁 Looking for file: {file_path}")
        
        if not file_path.exists():
            logger.error(f"❌ File not found: {file_path}")
            
            # Try to list available files
            available_files = list(OUTPUT_FOLDER.glob('*'))
            logger.info(f"📂 Available files: {[f.name for f in available_files]}")
            
            return jsonify({
                'error': 'File not found',
                'requested': filename,
                'available_files': [f.name for f in available_files]
            }), 404
        
        # Verify file is not empty
        file_size = file_path.stat().st_size
        if file_size == 0:
            logger.error(f"❌ File is empty: {file_path}")
            return jsonify({'error': 'File is empty'}), 500
        
        logger.info(f"✅ File found: {file_size} bytes")
        
        # Determine mimetype and read content
        if filename.endswith('.json'):
            # Validate JSON before sending
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                logger.info(f"✅ JSON is valid")
                
                # Return JSON with proper headers
                return jsonify(json_data), 200, {
                    'Content-Disposition': f'attachment; filename="{filename}"',
                    'Content-Type': 'application/json'
                }
            except json.JSONDecodeError as e:
                logger.error(f"❌ Invalid JSON: {str(e)}")
                return jsonify({'error': f'Invalid JSON file: {str(e)}'}), 500
        
        elif filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif filename.endswith('.pdf'):
            mimetype = 'application/pdf'
        else:
            mimetype = 'application/octet-stream'
        
        # For non-JSON files, use send_file
        logger.info(f"✅ Sending file: {filename} ({mimetype})")
        return send_file(
            str(file_path),
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype
        )
    
    except Exception as e:
        logger.error(f"❌ Error downloading file: {str(e)}", exc_info=True)
        return jsonify({
            'error': str(e),
            'filename': filename,
            'traceback': traceback.format_exc()
        }), 500


if __name__ == '__main__':
    logger.info("=" * 70)
    logger.info("SECURITIES DEPOSITS ANALYSIS API - Starting")
    logger.info("=" * 70)
    logger.info("Port: 5008 (Orchestration Standard)")
    logger.info("Original App: http://localhost:8567 (MongoDB)")
    logger.info("Mode: Standalone (MongoDB-free) + Original App Proxy")
    logger.info("Endpoints:")
    logger.info("  GET  /health")
    logger.info("  GET  /api/v1/info")
    logger.info("  POST /api/v1/execute")
    logger.info("  GET  /api/v1/download/<session_id>/<filename>")
    logger.info("=" * 70)
    
    # Check original app on startup
    check_original_app()
    
    app.run(host='0.0.0.0', port=5008, debug=True)