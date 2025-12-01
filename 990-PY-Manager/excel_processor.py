import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.styles.colors import Color
import io
import streamlit as st

class ExcelProcessor:
    def __init__(self):
        self.required_sheets = ['SOR', 'SOFE', 'SOFP']
    
    def _apply_number_formatting(self, cell, value, is_mapping_column=False):
        """Apply comma-style number formatting for numerical values, except for 990 Mapping columns"""
        if value is not None and isinstance(value, (int, float)) and not isinstance(value, bool):
            if not is_mapping_column:
                # Apply accounting format without currency symbol (comma separator with 2 decimals)
                cell.number_format = '#,##0.00'
                return True
            # For mapping columns, keep General formatting (default)
        return False
    
    def process_file_pair(self, input_file, template_file):
        """Process a single input file with its corresponding template"""
        try:
            # Load both files
            input_wb = openpyxl.load_workbook(input_file, data_only=True)
            template_wb = openpyxl.load_workbook(template_file)
            
            # Validate that both files have required sheets
            if not self._validate_sheets(input_wb, template_wb):
                return None
            
            # Process each sheet
            for sheet_name in self.required_sheets:
                success = self._process_sheet(
                    input_wb[sheet_name], 
                    template_wb[sheet_name], 
                    sheet_name
                )
                if not success:
                    st.warning(f"⚠️ Issues encountered processing {sheet_name} sheet")
            
            # Save to buffer
            output_buffer = io.BytesIO()
            template_wb.save(output_buffer)
            output_buffer.seek(0)
            
            return output_buffer
            
        except Exception as e:
            st.error(f"Error processing files: {str(e)}")
            return None
    
    def _validate_sheets(self, input_wb, template_wb):
        """Validate that both workbooks have required sheets"""
        input_sheets = set(input_wb.sheetnames)
        template_sheets = set(template_wb.sheetnames)
        required_sheets_set = set(self.required_sheets)
        
        if not required_sheets_set.issubset(input_sheets):
            missing = required_sheets_set - input_sheets
            st.error(f"Input file missing sheets: {', '.join(missing)}")
            return False
        
        if not required_sheets_set.issubset(template_sheets):
            missing = required_sheets_set - template_sheets
            st.error(f"Template file missing sheets: {', '.join(missing)}")
            return False
        
        return True
    
    def _apply_accounting_format_to_total_columns(self, sheet, sheet_name):
        """Apply accounting format to TOTAL(CY) and TOTAL(LY) columns"""
        # Get headers from row 3
        headers = self._get_headers_from_row(sheet, 3)
        if not headers:
            return
        
        # Find TOTAL columns based on sheet type
        total_columns = []
        if sheet_name == 'SOR':
            # For SOR sheet: Amount (CY) and Amount (LY)
            total_cy_col = self._find_column_index(headers, 'Amount (CY)')
            total_ly_col = self._find_column_index(headers, 'Amount (LY)')
        elif sheet_name == 'SOFE':
            # For SOFE sheet: Total (CY) and Total (LY)
            total_cy_col = self._find_column_index(headers, 'Total (CY)')
            total_ly_col = self._find_column_index(headers, 'Total (LY)')
        else:
            return
        
        # Add found columns to the list
        if total_cy_col is not None:
            total_columns.append(total_cy_col)
        if total_ly_col is not None:
            total_columns.append(total_ly_col)
        
        # Apply accounting format to all cells in these columns (starting from row 4)
        for col in total_columns:
            for row in range(4, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    # Use proper accounting format
                    cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

    def _find_first_total_row(self, sheet):
        """Find the first row that contains a value starting with 'Total' in the first column (Particulars)"""
        for row in range(1, sheet.max_row + 1):
            # Only check the first column (column A - Particulars column)
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip().upper().startswith('TOTAL'):
                return row
        return None

    def _get_last_data_column(self, sheet):
        """Find the last column that contains data in the sheet"""
        last_col = 1
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    last_col = max(last_col, col)
        return last_col

    def _is_red_text_row(self, sheet, row_num, last_col):
        """Check if any cell in the row has red text color"""
        for col in range(1, last_col + 1):
            cell = sheet.cell(row=row_num, column=col)
            if cell.font and cell.font.color:
                # Check if color is red (various ways red can be represented)
                color = cell.font.color
                if hasattr(color, 'rgb') and color.rgb:
                    rgb_value = str(color.rgb).upper()
                    # Common red color values
                    if rgb_value in ['FFFF0000', 'FF0000', 'FFFF0000']:
                        return True
                elif hasattr(color, 'theme') and color.theme == 2:  # Theme red
                    return True
                elif hasattr(color, 'indexed') and color.indexed == 10:  # Indexed red
                    return True
        return False

    def _find_ranges_for_borders(self, sheet):
        """Find non-empty row blocks in a sheet after the first 'Total' row and return the corresponding ranges."""
        ranges = []
        start_row = None
        
        # Find the first row containing "Total"
        first_total_row = self._find_first_total_row(sheet)
        if first_total_row is None:
            return ranges  # No "Total" row found, return empty ranges
        
        # Get the last column with data dynamically
        last_data_col = self._get_last_data_column(sheet)
        last_col_letter = get_column_letter(last_data_col)
        
        # Start checking from the row after the first "Total" row
        start_check_row = first_total_row + 1
        
        # Iterate through rows to find non-empty rows after the "Total" row
        for row in range(start_check_row, sheet.max_row + 1):
            # Skip rows with red text
            if self._is_red_text_row(sheet, row, last_data_col):
                if start_row is not None:
                    # End current block before the red text row
                    end_row = row - 1
                    ranges.append(f"A{start_row}:{last_col_letter}{end_row}")
                    start_row = None
                continue
            
            # Check if row is empty (all cells in the row are None or empty)
            is_empty = all([
                sheet.cell(row=row, column=col).value is None or 
                str(sheet.cell(row=row, column=col).value).strip() == ""
                for col in range(1, last_data_col + 1)  # Check up to last data column
            ])
            
            if not is_empty:
                if start_row is None:
                    start_row = row  # Start a new block
            else:
                if start_row is not None:
                    # End of a block, create range
                    end_row = row - 1
                    ranges.append(f"A{start_row}:{last_col_letter}{end_row}")
                    start_row = None  # Reset the block
                
        # Handle case where the last block extends to the end of the sheet
        if start_row is not None:
            end_row = sheet.max_row
            ranges.append(f"A{start_row}:{last_col_letter}{end_row}")
        
        return ranges

    def _get_existing_border_style(self, sheet):
        """Get the existing border style from the sheet by sampling existing borders"""
        # Get the last data column for dynamic checking
        last_data_col = self._get_last_data_column(sheet)
        
        # Sample some cells to find existing border style
        for row in range(1, min(20, sheet.max_row + 1)):
            for col in range(1, min(last_data_col + 1, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.border and cell.border.left and cell.border.left.style:
                    return cell.border.left.style
        
        # Default to thin if no existing borders found
        return 'thin'

    def _apply_matching_outside_borders(self, sheet, ranges):
        """Apply borders matching existing border thickness only to the outside edges of given ranges."""
        # Get the existing border style from the sheet
        border_style = self._get_existing_border_style(sheet)
        matching_side = Side(style=border_style)
        
        for range_str in ranges:
            start_cell, end_cell = range_str.split(':')
            start_row, start_col = coordinate_to_tuple(start_cell)
            end_row, end_col = coordinate_to_tuple(end_cell)
            
            # Apply matching borders only to the outer edges of the range
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    # Get current border or create new one
                    current_border = cell.border
                    
                    # Determine which sides need matching borders
                    left_border = matching_side if col == start_col else current_border.left
                    right_border = matching_side if col == end_col else current_border.right
                    top_border = matching_side if row == start_row else current_border.top
                    bottom_border = matching_side if row == end_row else current_border.bottom
                    
                    # Apply the border
                    cell.border = Border(
                        left=left_border,
                        right=right_border,
                        top=top_border,
                        bottom=bottom_border,
                        diagonal=current_border.diagonal,
                        diagonal_direction=current_border.diagonal_direction,
                        outline=current_border.outline,
                        vertical=current_border.vertical,
                        horizontal=current_border.horizontal
                    )

    def _shift_column_c_border_to_d(self, sheet):
        """Shift the right border from column C to column D, and extend top/bottom borders to both columns"""
        for row in range(4, sheet.max_row + 1):
            cell_c = sheet.cell(row=row, column=3)  # Column C
            cell_d = sheet.cell(row=row, column=4)  # Column D
            
            # Check if column C has borders
            has_right_border = cell_c.border and cell_c.border.right and cell_c.border.right.style
            has_top_border = cell_c.border and cell_c.border.top and cell_c.border.top.style
            has_bottom_border = cell_c.border and cell_c.border.bottom and cell_c.border.bottom.style
            
            if has_right_border or has_top_border or has_bottom_border:
                # Store the borders from column C
                right_border_style = cell_c.border.right if has_right_border else None
                top_border_style = cell_c.border.top if has_top_border else None
                bottom_border_style = cell_c.border.bottom if has_bottom_border else None
                
                # Update column C: Keep left, top, bottom borders, remove only right border
                cell_c.border = Border(
                    left=cell_c.border.left,
                    right=None,  # Remove right border only
                    top=cell_c.border.top,     # Keep top border
                    bottom=cell_c.border.bottom, # Keep bottom border
                    diagonal=cell_c.border.diagonal,
                    diagonal_direction=cell_c.border.diagonal_direction,
                    outline=cell_c.border.outline,
                    vertical=cell_c.border.vertical,
                    horizontal=cell_c.border.horizontal
                )
                
                # Update column D: Add right border from C, top/bottom borders from C
                cell_d.border = Border(
                    left=cell_d.border.left if cell_d.border else None,  # Keep existing left border
                    right=right_border_style,  # Add right border from C to create outside border
                    top=top_border_style,      # Copy top border here
                    bottom=bottom_border_style, # Copy bottom border here
                    diagonal=cell_d.border.diagonal if cell_d.border else None,
                    diagonal_direction=cell_d.border.diagonal_direction if cell_d.border else None,
                    outline=cell_d.border.outline if cell_d.border else None,
                    vertical=cell_d.border.vertical if cell_d.border else None,
                    horizontal=cell_d.border.horizontal if cell_d.border else None
                )

    def _process_sheet(self, input_sheet, template_sheet, sheet_name):
        """Process a single sheet based on its type"""
        try:
            if sheet_name == 'SOR':
                return self._process_sor_sheet(input_sheet, template_sheet)
            elif sheet_name == 'SOFE':
                return self._process_sofe_sheet(input_sheet, template_sheet)
            elif sheet_name == 'SOFP':
                return self._process_sofp_sheet(input_sheet, template_sheet)
            return False
        except Exception as e:
            st.error(f"Error processing {sheet_name} sheet: {str(e)}")
            return False
    
    def _process_sor_sheet(self, input_sheet, template_sheet):
        """Process SOR sheet: headers on row 3"""
        # Find headers in row 3
        input_headers = self._get_headers_from_row(input_sheet, 3)
        template_headers = self._get_headers_from_row(template_sheet, 3)
        
        if not input_headers or not template_headers:
            st.error("Could not find headers in SOR sheet row 3")
            return False
        
        # Find column indices
        input_cols = self._find_column_indices(input_headers, ['Particulars', 'Amount (CY)', '990 Mapping (CY)'])
        template_cols = self._find_column_indices(template_headers, ['Particulars', 'Amount (LY)', '990 Mapping (LY)'])
        
        if not all(col is not None for col in input_cols.values()) or not all(col is not None for col in template_cols.values()):
            st.warning("Some required columns not found in SOR sheet")
            return False
        
        # Find the "TOTAL (As Per Audit Report)" boundary in both sheets
        input_total_row = self._find_total_audit_row(input_sheet, input_cols['Particulars'])
        template_total_row = self._find_total_audit_row(template_sheet, template_cols['Particulars'])
        
        column_mapping = {
            input_cols['Particulars']: template_cols['Particulars'],
            input_cols['Amount (CY)']: template_cols['Amount (LY)'],
            input_cols['990 Mapping (CY)']: template_cols['990 Mapping (LY)']
        }
        
        # Copy data with boundary respect
        self._copy_data_with_boundary(
            input_sheet, template_sheet, 4, column_mapping,
            input_total_row, template_total_row
        )
        
        # Also copy mapping to CY column if it exists
        template_mapping_cy_col = self._find_column_index(template_headers, '990 Mapping (CY)')
        if template_mapping_cy_col is not None:
            self._copy_mapping_with_boundary(
                input_sheet, template_sheet, 4,
                input_cols['990 Mapping (CY)'], template_mapping_cy_col,
                input_total_row, template_total_row
            )
        
        # Apply accounting format to TOTAL columns
        self._apply_accounting_format_to_total_columns(template_sheet, 'SOR')
        
        # Apply matching outside borders to non-empty row blocks after first "Total" row
        border_ranges = self._find_ranges_for_borders(template_sheet)
        self._apply_matching_outside_borders(template_sheet, border_ranges)
        
        return True
    
    def _process_sofe_sheet(self, input_sheet, template_sheet):
        """Process SOFE sheet: headers on row 3"""
        # Find headers in row 3
        input_headers = self._get_headers_from_row(input_sheet, 3)
        template_headers = self._get_headers_from_row(template_sheet, 3)
        
        if not input_headers or not template_headers:
            st.error("Could not find headers in SOFE sheet")
            return False
        
        # Find column indices
        input_cols = self._find_column_indices(input_headers, ['Particulars', 'Total (CY)', '990 Mapping (CY)'])
        template_cols = self._find_column_indices(template_headers, ['Particulars', 'Total (LY)', '990 Mapping (LY)'])
        
        if not all(col is not None for col in input_cols.values()) or not all(col is not None for col in template_cols.values()):
            st.warning("Some required columns not found in SOFE sheet")
            return False
        
        # Find the "TOTAL (As Per Audit Report)" boundary in both sheets
        input_total_row = self._find_total_audit_row(input_sheet, input_cols['Particulars'])
        template_total_row = self._find_total_audit_row(template_sheet, template_cols['Particulars'])
        
        column_mapping = {
            input_cols['Particulars']: template_cols['Particulars'],
            input_cols['Total (CY)']: template_cols['Total (LY)'],
            input_cols['990 Mapping (CY)']: template_cols['990 Mapping (LY)']
        }
        
        # Copy data with boundary respect
        self._copy_data_with_boundary(
            input_sheet, template_sheet, 4, column_mapping,
            input_total_row, template_total_row
        )
        
        # Also copy mapping to CY column if it exists
        template_mapping_cy_col = self._find_column_index(template_headers, '990 Mapping (CY)')
        if template_mapping_cy_col is not None:
            self._copy_mapping_with_boundary(
                input_sheet, template_sheet, 4,
                input_cols['990 Mapping (CY)'], template_mapping_cy_col,
                input_total_row, template_total_row
            )
        
        # Apply accounting format to TOTAL columns
        self._apply_accounting_format_to_total_columns(template_sheet, 'SOFE')
        
        # Apply matching outside borders to non-empty row blocks after first "Total" row
        border_ranges = self._find_ranges_for_borders(template_sheet)
        self._apply_matching_outside_borders(template_sheet, border_ranges)
        
        return True
    
    def _process_sofp_sheet(self, input_sheet, template_sheet):
        """Process SOFP sheet: headers on row 3, special column mapping, shift column C border to D"""
        # Find headers in row 3
        input_headers = self._get_headers_from_row(input_sheet, 3)
        template_headers = self._get_headers_from_row(template_sheet, 3)
        
        if not input_headers or not template_headers:
            st.error("Could not find headers in SOFP sheet")
            return False
        
        # Find Particulars column
        input_particulars_col = self._find_column_index(input_headers, 'Particulars')
        template_particulars_col = self._find_column_index(template_headers, 'Particulars')
        
        if input_particulars_col is None or template_particulars_col is None:
            st.error("Could not find Particulars column in SOFP sheet")
            return False
        
        # Find 990 Mapping columns
        input_mapping_col = self._find_column_index(input_headers, '990 Mapping (CY)')
        template_mapping_col = self._find_column_index(template_headers, '990 Mapping (LY)')
        
        # For SOFP: 2nd column after Particulars → 1st column after Particulars
        input_amount_col = input_particulars_col + 2  # 2nd column after Particulars
        template_amount_col = template_particulars_col + 1  # 1st column after Particulars
        
        # Build mapping
        column_mapping = {
            input_particulars_col: template_particulars_col,
            input_amount_col: template_amount_col
        }
        
        if input_mapping_col is not None and template_mapping_col is not None:
            column_mapping[input_mapping_col] = template_mapping_col
        
        # Also copy to CY mapping column if it exists
        if input_mapping_col is not None:
            template_mapping_cy_col = self._find_column_index(template_headers, '990 Mapping (CY)')
            if template_mapping_cy_col is not None:
                # Create a separate mapping entry for CY column
                self._copy_mapping_to_cy_column(input_sheet, template_sheet, 4, input_mapping_col, template_mapping_cy_col)
        
        # Copy data starting from row 4
        self._copy_data_rows(input_sheet, template_sheet, 4, column_mapping)
        
        # Apply accounting format to amount columns
        self._apply_accounting_format_to_total_columns(template_sheet, 'SOFP')
        
        # Shift the right border from column C to column D
        self._shift_column_c_border_to_d(template_sheet)
        
        return True
    
    def _get_headers_from_row(self, sheet, row_num):
        """Extract headers from a specific row"""
        headers = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col).value
            if cell_value:
                headers[str(cell_value).strip()] = col
        return headers
    
    def _find_column_indices(self, headers, column_names):
        """Find column indices for multiple column names"""
        result = {}
        for col_name in column_names:
            result[col_name] = self._find_column_index(headers, col_name)
        return result
    
    def _find_column_index(self, headers, column_name):
        """Find column index for a specific column name (case-insensitive)"""
        for header, col_index in headers.items():
            if column_name.lower() in header.lower():
                return col_index
        return None
    
    def _copy_data_rows(self, input_sheet, template_sheet, start_row, column_mapping):
        """Copy data rows based on column mapping - including blank rows with full formatting"""
        for row in range(start_row, input_sheet.max_row + 1):
            # Copy all rows (including blank ones) to maintain spacing
            self._copy_single_row_with_formatting(input_sheet, template_sheet, row, row, column_mapping)
    
    def _find_total_audit_row(self, sheet, particulars_col):
        """Find the row containing 'TOTAL (As Per Audit Report)' in the Particulars column"""
        if particulars_col is None:
            return None
        
        for row in range(4, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=particulars_col).value
            if cell_value and 'TOTAL' in str(cell_value).upper() and 'AUDIT REPORT' in str(cell_value).upper():
                return row
        return None
    
    def _copy_data_with_boundary(self, input_sheet, template_sheet, start_row, column_mapping, input_total_row, template_total_row):
        """Copy data while respecting the TOTAL (As Per Audit Report) boundary"""
        
        # If boundary not found in either sheet, fall back to regular copying
        if input_total_row is None or template_total_row is None:
            self._copy_data_rows(input_sheet, template_sheet, start_row, column_mapping)
            return
        
        # Copy data above the boundary (before TOTAL line) - including blank rows
        for row in range(start_row, input_total_row):
            self._copy_single_row_with_formatting(input_sheet, template_sheet, row, row, column_mapping)
        
        # Copy data below the boundary (after TOTAL line) - including all rows even blank ones
        input_after_total = input_total_row + 1
        template_after_total = template_total_row + 1
        
        # Copy all rows after TOTAL line to maintain spacing and formatting
        for i, input_row in enumerate(range(input_after_total, input_sheet.max_row + 1)):
            template_row = template_after_total + i
            self._copy_single_row_with_formatting(input_sheet, template_sheet, input_row, template_row, column_mapping)
    
    def _copy_single_row(self, input_sheet, template_sheet, input_row, template_row, column_mapping):
        """Copy a single row from input to template"""
        for input_col, template_col in column_mapping.items():
            if input_col is not None and template_col is not None:
                input_cell = input_sheet.cell(row=input_row, column=input_col)
                template_cell = template_sheet.cell(row=template_row, column=template_col)
                
                # Copy value
                template_cell.value = input_cell.value
                
                # Check if this is a mapping column
                is_mapping = template_col in [col for col in [
                    self._find_column_index(self._get_headers_from_row(template_sheet, 3), '990 Mapping (CY)'),
                    self._find_column_index(self._get_headers_from_row(template_sheet, 3), '990 Mapping (LY)')
                ] if col is not None]
                
                # Apply number formatting for numerical values
                self._apply_number_formatting(template_cell, input_cell.value, is_mapping)
                
                # Preserve formatting: Calibri, size 11, original color
                if input_cell.font:
                    template_cell.font = Font(
                        name='Calibri',
                        size=11,
                        color=input_cell.font.color if input_cell.font.color else '000000'
                    )
                else:
                    template_cell.font = Font(name='Calibri', size=11)
    
    def _copy_single_row_with_formatting(self, input_sheet, template_sheet, input_row, template_row, column_mapping):
        """Copy a single row from input to template with complete formatting preservation"""
        for input_col, template_col in column_mapping.items():
            if input_col is not None and template_col is not None:
                input_cell = input_sheet.cell(row=input_row, column=input_col)
                template_cell = template_sheet.cell(row=template_row, column=template_col)
                
                # Copy value (including None for blank cells)
                template_cell.value = input_cell.value
                
                # Check if this is a mapping column
                is_mapping = template_col in [col for col in [
                    self._find_column_index(self._get_headers_from_row(template_sheet, 3), '990 Mapping (CY)'),
                    self._find_column_index(self._get_headers_from_row(template_sheet, 3), '990 Mapping (LY)')
                ] if col is not None]
                
                # Apply number formatting for numerical values
                self._apply_number_formatting(template_cell, input_cell.value, is_mapping)
                
                # Preserve complete formatting
                if input_cell.font:
                    template_cell.font = Font(
                        name='Calibri',  # Keep Calibri as required
                        size=11,  # Keep size 11 as required
                        bold=input_cell.font.bold,  # Preserve bold
                        italic=input_cell.font.italic,  # Preserve italic
                        underline=input_cell.font.underline,  # Preserve underline
                        strike=input_cell.font.strike,  # Preserve strikethrough
                        color=input_cell.font.color if input_cell.font.color else '000000'  # Preserve color
                    )
                else:
                    template_cell.font = Font(name='Calibri', size=11)
                
                # Preserve cell fill (background color)
                if input_cell.fill and input_cell.fill.fill_type:
                    template_cell.fill = PatternFill(
                        fill_type=input_cell.fill.fill_type,
                        start_color=input_cell.fill.start_color,
                        end_color=input_cell.fill.end_color
                    )
                
                # Preserve borders
                if input_cell.border:
                    template_cell.border = Border(
                        left=input_cell.border.left,
                        right=input_cell.border.right,
                        top=input_cell.border.top,
                        bottom=input_cell.border.bottom,
                        diagonal=input_cell.border.diagonal,
                        diagonal_direction=input_cell.border.diagonal_direction,
                        outline=input_cell.border.outline,
                        vertical=input_cell.border.vertical,
                        horizontal=input_cell.border.horizontal
                    )
                
                # Preserve alignment
                if input_cell.alignment:
                    template_cell.alignment = Alignment(
                        horizontal=input_cell.alignment.horizontal,
                        vertical=input_cell.alignment.vertical,
                        text_rotation=input_cell.alignment.text_rotation,
                        wrap_text=input_cell.alignment.wrap_text,
                        shrink_to_fit=input_cell.alignment.shrink_to_fit,
                        indent=input_cell.alignment.indent
                    )
    
    def _copy_mapping_to_cy_column(self, input_sheet, template_sheet, start_row, input_col, template_col):
        """Copy mapping data to CY column separately"""
        for row in range(start_row, input_sheet.max_row + 1):
            input_cell = input_sheet.cell(row=row, column=input_col)
            template_cell = template_sheet.cell(row=row, column=template_col)
            
            # Copy value and formatting
            template_cell.value = input_cell.value
            
            # Apply number formatting for numerical values (mapping column = True)
            self._apply_number_formatting(template_cell, input_cell.value, True)
            
            # Preserve complete formatting
            if input_cell.font:
                template_cell.font = Font(
                    name='Calibri',
                    size=11,
                    bold=input_cell.font.bold,
                    italic=input_cell.font.italic,
                    underline=input_cell.font.underline,
                    strike=input_cell.font.strike,
                    color=input_cell.font.color if input_cell.font.color else '000000'
                )
            else:
                template_cell.font = Font(name='Calibri', size=11)
            
            # Preserve cell fill
            if input_cell.fill and input_cell.fill.fill_type:
                template_cell.fill = PatternFill(
                    fill_type=input_cell.fill.fill_type,
                    start_color=input_cell.fill.start_color,
                    end_color=input_cell.fill.end_color
                )
            
            # Preserve borders
            if input_cell.border:
                template_cell.border = Border(
                    left=input_cell.border.left,
                    right=input_cell.border.right,
                    top=input_cell.border.top,
                    bottom=input_cell.border.bottom,
                    diagonal=input_cell.border.diagonal,
                    diagonal_direction=input_cell.border.diagonal_direction,
                    outline=input_cell.border.outline,
                    vertical=input_cell.border.vertical,
                    horizontal=input_cell.border.horizontal
                )
            
            # Preserve alignment
            if input_cell.alignment:
                template_cell.alignment = Alignment(
                    horizontal=input_cell.alignment.horizontal,
                    vertical=input_cell.alignment.vertical,
                    text_rotation=input_cell.alignment.text_rotation,
                    wrap_text=input_cell.alignment.wrap_text,
                    shrink_to_fit=input_cell.alignment.shrink_to_fit,
                    indent=input_cell.alignment.indent
                )
    
    def _copy_mapping_with_boundary(self, input_sheet, template_sheet, start_row, input_col, template_col, input_total_row, template_total_row):
        """Copy mapping data to column while respecting boundary"""
        
        # If boundary not found in either sheet, fall back to regular copying
        if input_total_row is None or template_total_row is None:
            self._copy_mapping_to_cy_column(input_sheet, template_sheet, start_row, input_col, template_col)
            return
        
        # Copy data above the boundary (before TOTAL line)
        for row in range(start_row, input_total_row):
            self._copy_single_cell_with_formatting(input_sheet, template_sheet, row, row, input_col, template_col)
        
        # Copy data below the boundary (after TOTAL line)
        input_after_total = input_total_row + 1
        template_after_total = template_total_row + 1
        
        for i, input_row in enumerate(range(input_after_total, input_sheet.max_row + 1)):
            template_row = template_after_total + i
            self._copy_single_cell_with_formatting(input_sheet, template_sheet, input_row, template_row, input_col, template_col)
    
    def _copy_single_cell_with_formatting(self, input_sheet, template_sheet, input_row, template_row, input_col, template_col):
        """Copy a single cell with complete formatting"""
        input_cell = input_sheet.cell(row=input_row, column=input_col)
        template_cell = template_sheet.cell(row=template_row, column=template_col)
        
        # Copy value
        template_cell.value = input_cell.value
        
        # Apply number formatting for numerical values (mapping column = True)
        self._apply_number_formatting(template_cell, input_cell.value, True)
        
        # Preserve complete formatting
        if input_cell.font:
            template_cell.font = Font(
                name='Calibri',
                size=11,
                bold=input_cell.font.bold,
                italic=input_cell.font.italic,
                underline=input_cell.font.underline,
                strike=input_cell.font.strike,
                color=input_cell.font.color if input_cell.font.color else '000000'
            )
        else:
            template_cell.font = Font(name='Calibri', size=11)
        
        # Preserve cell fill
        if input_cell.fill and input_cell.fill.fill_type:
            template_cell.fill = PatternFill(
                fill_type=input_cell.fill.fill_type,
                start_color=input_cell.fill.start_color,
                end_color=input_cell.fill.end_color
            )
        
        # Preserve borders
        if input_cell.border:
            template_cell.border = Border(
                left=input_cell.border.left,
                right=input_cell.border.right,
                top=input_cell.border.top,
                bottom=input_cell.border.bottom,
                diagonal=input_cell.border.diagonal,
                diagonal_direction=input_cell.border.diagonal_direction,
                outline=input_cell.border.outline,
                vertical=input_cell.border.vertical,
                horizontal=input_cell.border.horizontal
            )
        
        # Preserve alignment
        if input_cell.alignment:
            template_cell.alignment = Alignment(
                horizontal=input_cell.alignment.horizontal,
                vertical=input_cell.alignment.vertical,
                text_rotation=input_cell.alignment.text_rotation,
                wrap_text=input_cell.alignment.wrap_text,
                shrink_to_fit=input_cell.alignment.shrink_to_fit,
                indent=input_cell.alignment.indent
            )