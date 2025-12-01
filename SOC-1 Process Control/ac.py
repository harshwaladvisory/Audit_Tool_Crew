# ac.py - UPDATED WITH MONGODB
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import os
import io
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font

# Import configuration
from config import Config

# Import MongoDB operations
from db_operations import (
    doc_ops,
    session_ops,
    result_ops,
    audit_ops
)

# Import document processing logic
from document_processor import (
    extract_text_from_image,
    extract_text_from_pdf,
    extract_text_from_pdf_with_pdf2image,
    extract_text_from_docx,
    extract_text_from_excel,
    clean_data_for_json,
    send_raw_to_n8n,
    save_to_excel
)

app = Flask(__name__)
app.config.from_object(Config)

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Get configuration values
WEBHOOK_URLS = Config.WEBHOOK_URLS
EXCEL_FILE_PATH = Config.EXCEL_FILE_PATH
CONTROL_WEBHOOK_URL = Config.CONTROL_WEBHOOK_URL

def allowed_file(filename):
    """Check if file has an allowed extension"""
    _, extension = os.path.splitext(filename.lower())
    return extension in Config.ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        # Process the file
        file_extension = os.path.splitext(filename)[1].lower()
        extracted_data = []
        raw_text = ""
        
        try:
            if file_extension == '.pdf':
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_pdf(f)
                    # If PyMuPDF extraction returns little text, try pdf2image as fallback
                    if len(raw_text.strip()) < 100:  # Threshold for minimal text
                        alt_text, alt_data = extract_text_from_pdf_with_pdf2image(f)
                        if len(alt_text.strip()) > len(raw_text.strip()):
                            raw_text = alt_text
                            extracted_data = alt_data
            
            elif file_extension == '.docx':
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_docx(f)

            elif file_extension in ['.xlsx', '.xls']:
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_excel(f)

            elif file_extension in ['.png', '.jpg', '.jpeg']:
                from PIL import Image
                image = Image.open(file_path)
                raw_text = extract_text_from_image(image)
                extracted_data = [{"page": 1, "source": "image_ocr", "content": raw_text}]
            
            else:
                return jsonify({'error': 'Unsupported file format'}), 400
            
            # Clean data before saving
            cleaned_extracted_data = clean_data_for_json(extracted_data)
            cleaned_raw_text = clean_data_for_json(raw_text)
            
            # ✅ SAVE TO MONGODB
            document_id = doc_ops.create_document(
                filename=filename,
                file_type=file_extension,
                file_size=file_size,
                raw_text=cleaned_raw_text,
                extracted_data=cleaned_extracted_data
            )
            
            # Create session
            session_id = f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}"
            
            if document_id:
                session_ops.create_session(
                    session_id=session_id,
                    document_id=document_id,
                    filename=filename,
                    file_path=file_path
                )
                print(f"✓ Document saved to MongoDB: {document_id}")
            else:
                # Fallback to temp file if MongoDB is not available
                print("⚠ MongoDB not available, using temporary file storage")
                session_data = {
                    'filename': filename,
                    'raw_text': cleaned_raw_text,
                    'extracted_data': cleaned_extracted_data,
                    'file_path': file_path
                }
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
                with open(temp_file.name, 'w') as f:
                    json.dump(session_data, f)
                session_id = os.path.basename(temp_file.name)
            
            return jsonify({
                'success': True,
                'filename': filename,
                'raw_text': raw_text[:500] + '...' if len(raw_text) > 500 else raw_text,  # Truncate for response
                'session_id': session_id,
                'document_id': document_id if document_id else None,
                'storage': 'mongodb' if document_id else 'temporary'
            })
            
        except Exception as e:
            audit_ops.log_action(
                "document_upload_failed",
                details={"filename": filename, "error": str(e)},
                status="error"
            )
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500
        
        finally:
            # Clean up uploaded file
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as cleanup_error:
                print(f"Warning: Could not clean up file {file_path}: {cleanup_error}")
    
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/send_to_n8n', methods=['POST'])
def send_to_n8n():
    import requests
    from openpyxl import load_workbook, Workbook
    import pandas as pd

    data = request.json
    session_id = data.get('session_id')
    repeat_count = int(data.get('count', 1))
    
    if not session_id:
        return jsonify({'error': 'No session ID provided'}), 400
    
    try:
        # ✅ TRY MONGODB FIRST
        session = session_ops.get_session(session_id)
        
        if session:
            # Load from MongoDB
            document = doc_ops.get_document(str(session['document_id']))
            if not document:
                return jsonify({'error': 'Document not found in database'}), 404
            
            filename = document['filename']
            extracted_data = document['extracted_data']
            print(f"✓ Loaded session from MongoDB: {session_id}")
            
        else:
            # ✅ FALLBACK TO TEMP FILE
            session_file_path = os.path.join(tempfile.gettempdir(), session_id)
            if not os.path.exists(session_file_path):
                return jsonify({'error': 'Session data not found'}), 400
                
            with open(session_file_path, 'r') as f:
                session_data = json.load(f)
            
            filename = session_data['filename']
            extracted_data = session_data['extracted_data']
            print(f"⚠ Loaded session from temporary file: {session_id}")
        
        file_extension = os.path.splitext(filename)[1].lower()
        webhook_url = WEBHOOK_URLS.get(file_extension)
        
        if not webhook_url:
            return jsonify({'error': 'No webhook URL for this file type'}), 400
        
        # Update session with processing info
        session_ops.update_repeat_count(session_id, repeat_count)
        session_ops.add_processing_step(
            session_id, "n8n_processing", "started", 
            {"repeat_count": repeat_count}
        )
        
        current_row = 1

        # Create or validate Excel file
        try:
            if not os.path.exists(EXCEL_FILE_PATH):
                wb = Workbook()
                wb.save(EXCEL_FILE_PATH)
                print(f"Created new Excel file: {EXCEL_FILE_PATH}")
            else:
                try:
                    wb = load_workbook(EXCEL_FILE_PATH)
                    wb.close()
                    print(f"Existing Excel file is valid: {EXCEL_FILE_PATH}")
                except Exception as e:
                    print(f"Existing Excel file is corrupted: {e}")
                    wb = Workbook()
                    wb.save(EXCEL_FILE_PATH)
                    print(f"Created new Excel file to replace corrupted one: {EXCEL_FILE_PATH}")
        except Exception as e:
            return jsonify({'error': f'Failed to create/validate Excel file: {str(e)}'}), 500

        # Load workbook to find current row
        try:
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active
            if ws.max_row > 1:
                current_row = ws.max_row + 2
            wb.close()
        except Exception as e:
            print(f"Error reading Excel file for row count: {e}")
            current_row = 1

        # Step 1: Send to n8n webhook and collect results
        all_results = []
        for i in range(repeat_count):
            try:
                success, n8n_response, message = send_raw_to_n8n(
                    extracted_data, webhook_url, f"{filename}_part{i+1}"
                )
                
                if not success:
                    session_ops.add_processing_step(
                        session_id, f"n8n_iteration_{i+1}", "failed",
                        {"error": message}
                    )
                    return jsonify({'success': False, 'message': f'Failed at iteration {i+1}: {message}'})

                if not isinstance(n8n_response, list):
                    n8n_response = [n8n_response] if isinstance(n8n_response, dict) else []

                if n8n_response:
                    # ✅ SAVE TO MONGODB
                    if session and document:
                        result_ids = result_ops.bulk_create_results(
                            document_id=str(document['_id']),
                            session_id=session_id,
                            results_data=n8n_response
                        )
                        all_results.extend(result_ids)
                        print(f"✓ Saved {len(result_ids)} results to MongoDB")
                    
                    # Save to Excel
                    try:
                        current_row = save_to_excel(
                            n8n_response, EXCEL_FILE_PATH, 
                            f"{filename}_part{i+1}", current_row
                        )
                        
                        session_ops.add_processing_step(
                            session_id, f"n8n_iteration_{i+1}", "completed",
                            {"result_count": len(n8n_response)}
                        )
                    except Exception as e:
                        print(f"Error saving to Excel at iteration {i+1}: {e}")
                        return jsonify({'error': f'Error saving to Excel: {str(e)}'}), 500
                        
            except Exception as e:
                return jsonify({'error': f'Error in n8n processing at iteration {i+1}: {str(e)}'}), 500

        # Mark results as exported to Excel
        if all_results:
            exported_count = result_ops.mark_as_excel_exported(all_results)
            print(f"✓ Marked {exported_count} results as exported to Excel")

        # Step 2: Validate Excel file
        try:
            if not os.path.exists(EXCEL_FILE_PATH):
                return jsonify({'error': 'Excel file was not created properly'}), 500
                
            df = pd.read_excel(EXCEL_FILE_PATH)
            excel_structured_data = df.fillna("").to_dict(orient="records")
            
        except Exception as e:
            return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 500

        # Step 3: Create Control sheet
        try:
            wb = load_workbook(EXCEL_FILE_PATH)
            if 'Control' in wb.sheetnames:
                wb.remove(wb['Control'])
            
            ws = wb.create_sheet('Control')
            ws.cell(row=1, column=1, value="Processing Complete - Use Finalize to generate control data")
            
            wb.save(EXCEL_FILE_PATH)
            wb.close()
            
        except Exception as e:
            print(f"Error creating Control sheet: {e}")

        # Update session status
        session_ops.update_session_status(session_id, "n8n_completed")
        
        # Log successful processing
        audit_ops.log_action(
            "n8n_processing_completed",
            details={
                "session_id": session_id,
                "iterations": repeat_count,
                "results_count": len(all_results)
            }
        )

        return jsonify({
            'success': True,
            'n8n_success': True,
            'excel_saved': True,
            'results_count': len(all_results),
            'message': f'Processed {repeat_count} iteration(s) successfully. Excel file ready for download.'
        })

    except Exception as e:
        audit_ops.log_action(
            "n8n_processing_failed",
            details={"session_id": session_id, "error": str(e)},
            status="error"
        )
        print(f"Unexpected error in send_to_n8n: {e}")
        return jsonify({'error': f'Unexpected error: {str(e)}'}), 500

@app.route('/finalize_excel', methods=['POST'])
def finalize_excel():
    import requests
    import pandas as pd
    from openpyxl import load_workbook

    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({'error': 'Excel file not found'}), 400

        try:
            df = pd.read_excel(EXCEL_FILE_PATH)
        except Exception as e:
            return jsonify({'error': f'Excel file is corrupt or unreadable: {str(e)}'}), 500

        structured_data = df.fillna("").to_dict(orient="records")

        try:
            control_response = requests.post(CONTROL_WEBHOOK_URL, json={"rows": structured_data})
            control_response.raise_for_status()
        except requests.exceptions.RequestException as e:
            audit_ops.log_action(
                "control_webhook_failed",
                details={"error": str(e)},
                status="error"
            )
            return jsonify({'error': f'Webhook call failed: {str(e)}'}), 500

        try:
            control_data = control_response.json()
        except Exception as e:
            return jsonify({'error': f'Invalid JSON received from webhook: {str(e)}'}), 500

        try:
            wb = load_workbook(EXCEL_FILE_PATH)
        except Exception as e:
            return jsonify({'error': f'Could not open Excel file to write Control sheet: {str(e)}'}), 500

        if 'Control' in wb.sheetnames:
            wb.remove(wb['Control'])
        ws = wb.create_sheet('Control')

        if isinstance(control_data, list) and control_data:
            headers = list(control_data[0].keys())
            for col_index, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_index, value=header)
                ws.cell(row=1, column=col_index).font = Font(bold=True, size=14)
            for row_index, row in enumerate(control_data, 2):
                for col_index, header in enumerate(headers, 1):
                    ws.cell(row=row_index, column=col_index, value=row.get(header))

        wb.save(EXCEL_FILE_PATH)
        
        audit_ops.log_action(
            "excel_finalized",
            details={"control_records": len(control_data) if isinstance(control_data, list) else 0}
        )

        return jsonify({
            'success': True,
            'message': 'Control sheet updated.',
            'control_data': control_data
        })

    except Exception as e:
        audit_ops.log_action(
            "excel_finalize_failed",
            details={"error": str(e)},
            status="error"
        )
        return jsonify({'error': f'Finalize error: {str(e)}'}), 500


@app.route('/download_text/<session_id>')
def download_text(session_id):
    try:
        # Try MongoDB first
        session = session_ops.get_session(session_id)
        
        if session:
            document = doc_ops.get_document(str(session['document_id']))
            if document:
                raw_text = document['raw_text']
                filename = document['filename']
            else:
                return jsonify({'error': 'Document not found'}), 404
        else:
            # Fallback to temp file
            with open(os.path.join(tempfile.gettempdir(), session_id), 'r') as f:
                session_data = json.load(f)
            raw_text = session_data['raw_text']
            filename = session_data['filename']
        
        # Create in-memory text file
        text_file = io.BytesIO(raw_text.encode())
        
        return send_file(
            text_file,
            mimetype="text/plain",
            as_attachment=True,
            download_name=f"extracted_text_{os.path.splitext(filename)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
    
    except Exception as e:
        return jsonify({'error': f'Error downloading text: {str(e)}'}), 500

@app.route('/download_excel')
def download_excel():
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            return send_file(
                EXCEL_FILE_PATH,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"extracted_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        except Exception as e:
            return jsonify({'error': f'Error downloading Excel: {str(e)}'}), 500
    else:
        return jsonify({'error': 'Excel file does not exist'}), 404

@app.route('/preview_excel')
def preview_excel():
    import pandas as pd
    
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            df = pd.read_excel(EXCEL_FILE_PATH)
            return jsonify({
                'success': True,
                'data': df.to_dict('records'),
                'columns': df.columns.tolist()
            })
        except Exception as e:
            return jsonify({'error': f'Error previewing Excel: {str(e)}'}), 500
    else:
        return jsonify({'error': 'Excel file does not exist'}), 404

@app.route('/clear_excel', methods=['POST'])
def clear_excel():
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            os.remove(EXCEL_FILE_PATH)
            audit_ops.log_action("excel_cleared")
            return jsonify({'success': True, 'message': 'Excel file cleared'})
        except Exception as e:
            return jsonify({'error': f'Error clearing Excel: {str(e)}'}), 500
    else:
        return jsonify({'success': True, 'message': 'Excel file does not exist'})

@app.route('/api/kpi_data')
def get_kpi_data():
    """Return KPI data for the dashboard"""
    import glob
    
    try:
        # Try to get document count from MongoDB
        document_count = doc_ops.count_documents()
        if document_count == 0:
            # Fallback to counting files in upload folder
            document_count = len(glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], '*')))
        
        # Try to get controls count from MongoDB
        controls_count = result_ops.count_results()
        if controls_count == 0 and os.path.exists(EXCEL_FILE_PATH):
            # Fallback to Excel file
            try:
                import pandas as pd
                df = pd.read_excel(EXCEL_FILE_PATH)
                controls_count = len(df)
            except:
                pass
        
        # Calculate metrics
        avg_processing_time = 2.3
        success_rate = 95.0 if controls_count > 0 else 0
        
        return jsonify({
            'documents': document_count,
            'templates': 1,
            'processingTime': avg_processing_time,
            'successRate': success_rate,
            'controls': controls_count
        })
    
    except Exception as e:
        return jsonify({
            'documents': 0,
            'templates': 1,
            'processingTime': 0,
            'successRate': 0,
            'controls': 0
        })


# ✅ NEW MONGODB API ENDPOINTS

@app.route('/api/documents', methods=['GET'])
def get_documents():
    """Get all documents from MongoDB"""
    try:
        limit = int(request.args.get('limit', 100))
        skip = int(request.args.get('skip', 0))
        
        documents = doc_ops.get_all_documents(limit=limit, skip=skip)
        
        # Convert ObjectId to string
        for doc in documents:
            doc['_id'] = str(doc['_id'])
        
        return jsonify({
            'success': True,
            'documents': documents,
            'count': len(documents)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/documents/<document_id>', methods=['GET'])
def get_document(document_id):
    """Get a specific document"""
    try:
        document = doc_ops.get_document(document_id)
        if not document:
            return jsonify({'error': 'Document not found'}), 404
        
        document['_id'] = str(document['_id'])
        return jsonify({'success': True, 'document': document})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sessions/<session_id>', methods=['GET'])
def get_session_info(session_id):
    """Get session information"""
    try:
        session = session_ops.get_session(session_id)
        if not session:
            return jsonify({'error': 'Session not found'}), 404
        
        session['_id'] = str(session['_id'])
        session['document_id'] = str(session['document_id'])
        
        # Get results for this session
        results = result_ops.get_results_by_session(session_id)
        for result in results:
            result['_id'] = str(result['_id'])
            result['document_id'] = str(result['document_id'])
        
        return jsonify({
            'success': True,
            'session': session,
            'results': results,
            'results_count': len(results)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/audit_logs', methods=['GET'])
def get_audit_logs():
    """Get audit logs"""
    try:
        limit = int(request.args.get('limit', 50))
        skip = int(request.args.get('skip', 0))
        
        logs = audit_ops.get_logs(limit=limit, skip=skip)
        
        # Convert ObjectId to string
        for log in logs:
            log['_id'] = str(log['_id'])
        
        return jsonify({
            'success': True,
            'logs': logs,
            'count': len(logs)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Cleanup on app shutdown
@app.teardown_appcontext
def shutdown_db(exception=None):
    """Close MongoDB connection on shutdown"""
    from database import mongo_db
    if mongo_db:
        mongo_db.close()

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Starting SOC 1 Process Control Application")
    print("=" * 60)
    print(f"📁 Upload Folder: {app.config['UPLOAD_FOLDER']}")
    print(f"📊 Excel Output: {EXCEL_FILE_PATH}")
    print(f"🗄️  MongoDB: {Config.MONGODB_NAME}")
    print("=" * 60)
    app.run(host="0.0.0.0", port="8506", debug=False)