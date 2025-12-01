import pytesseract
from PIL import Image
import pandas as pd
import io
import os
import fitz  # PyMuPDF
import tempfile
from datetime import datetime
import requests
import json
from docx import Document
from pdf2image import convert_from_bytes
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import numpy as np
import base64

pytesseract.pytesseract.tesseract_cmd = r'C:\Users\Tamanna.garg.HCLLP\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'  # Example for Windows

# Function to extract text from image using Tesseract
def extract_text_from_image(image):
    try:
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        return f"Error extracting text from image: {str(e)}"

def extract_text_from_pdf(pdf_file):
    try:
        pdf_file.seek(0)
        pdf_bytes = pdf_file.read()
        pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        extracted_text = ""
        all_extracted_data = []
        
        for page_num in range(len(pdf_doc)):
            page = pdf_doc[page_num]
            page_output = f"\n--- Page {page_num + 1} ---\n"
            
            # Extract text directly from PDF
            pdf_text = page.get_text()
            if pdf_text.strip():
                page_output += f"[PDF Text Extraction]\n{pdf_text}\n"
                all_extracted_data.append({
                    "page": page_num + 1,
                    "source": "pdf_text",
                    "content": pdf_text
                })
            
            # OCR from rendered page
            try:
                pix = page.get_pixmap(dpi=300)
                image = Image.open(io.BytesIO(pix.tobytes("png")))
                rendered_text = extract_text_from_image(image).strip()
                page_output += f"[OCR from rendered page]\n{rendered_text if rendered_text else 'No text found.'}\n"
                
                if rendered_text:
                    all_extracted_data.append({
                        "page": page_num + 1,
                        "source": "page_ocr",
                        "content": rendered_text
                    })
            except Exception as e:
                page_output += f"[Error rendering page for OCR]: {str(e)}\n"
            
            # Text from embedded images
            images = page.get_images(full=True)
            if images:
                for img_index, img in enumerate(images):
                    xref = img[0]
                    try:
                        base_image = pdf_doc.extract_image(xref)
                        image_bytes = base_image["image"]
                        image_ext = base_image["ext"].lower()
                        
                        if image_ext in ["jpeg", "jpg", "png"]:
                            img_obj = Image.open(io.BytesIO(image_bytes))
                            img_text = extract_text_from_image(img_obj).strip()
                            page_output += f"\n[OCR from image {xref}]\n{img_text if img_text else 'No text found in image.'}\n"
                            
                            if img_text:
                                all_extracted_data.append({
                                    "page": page_num + 1,
                                    "source": f"embedded_image_{xref}",
                                    "content": img_text
                                })
                        else:
                            page_output += f"\n[Unsupported image format: {image_ext} for image {xref}]\n"
                    except Exception as e:
                        page_output += f"\n[Error processing image {xref}]: {str(e)}\n"
            else:
                page_output += "\n[No embedded images found]\n"
            
            extracted_text += page_output
        
        pdf_doc.close()
        return extracted_text.strip(), all_extracted_data
    
    except Exception as e:
        return f"Error extracting content from PDF via OCR: {str(e)}", []

def extract_text_from_pdf_with_pdf2image(pdf_file):
    """Alternative PDF extraction using pdf2image"""
    try:
        pdf_file.seek(0)
        images = convert_from_bytes(pdf_file.read())
        extracted_text = ""
        all_extracted_data = []
        
        for i, img in enumerate(images):
            page_text = pytesseract.image_to_string(img)
            extracted_text += f"\n--- Page {i+1} ---\n{page_text}\n"
            
            if page_text.strip():
                all_extracted_data.append({
                    "page": i + 1,
                    "source": "pdf2image_ocr",
                    "content": page_text
                })
        
        return extracted_text.strip(), all_extracted_data
    except Exception as e:
        return f"Error extracting PDF with pdf2image: {str(e)}", []

def extract_text_from_docx(docx_file):
    """Extract text from Word document"""
    try:
        docx_file.seek(0)
        doc = Document(docx_file)
        extracted_text = "\n".join([para.text for para in doc.paragraphs])
        
        # Structure data similar to PDF extraction for consistency
        all_extracted_data = [{
            "page": 1,
            "source": "docx_text",
            "content": extracted_text
        }]
        
        return extracted_text, all_extracted_data
    except Exception as e:
        return f"Error extracting content from DOCX: {str(e)}", []

def clean_data_for_json(data):
    """Clean data to make it JSON serializable"""
    if isinstance(data, dict):
        return {key: clean_data_for_json(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [clean_data_for_json(item) for item in data]
    elif pd.isna(data):
        return ""
    elif isinstance(data, (pd.Timestamp, datetime)):
        return data.isoformat()
    elif isinstance(data, (np.integer, np.floating)):
        if pd.isna(data) or np.isinf(data):
            return ""
        return float(data) if isinstance(data, np.floating) else int(data)
    elif isinstance(data, np.ndarray):
        return clean_data_for_json(data.tolist())
    else:
        try:
            str_val = str(data).strip()
            return "" if str_val.lower() in ['nan', 'nat', 'none', 'null'] else str_val
        except:
            return ""

# def extract_text_from_excel(excel_file):
#     """Extract text from Excel file with consistent data structure"""
#     try:
#         excel_file.seek(0)
        
#         # Read all sheets from the Excel file
#         try:
#             df_dict = pd.read_excel(excel_file, sheet_name=None)
#         except Exception as e:
#             return f"Error reading Excel file: {str(e)}", []
        
#         extracted_text = ""
#         all_extracted_data = []
        
#         # Process each sheet
#         for sheet_name, df in df_dict.items():
#             # Clean the dataframe - replace NaN, inf, and other problematic values
#             df = df.replace([np.nan, np.inf, -np.inf], "")
            
#             # Convert empty strings and None values to empty string for consistency
#             df = df.fillna("")
            
#             # Convert datetime columns to strings to avoid serialization issues
#             for col in df.columns:
#                 if df[col].dtype == 'datetime64[ns]' or 'datetime' in str(df[col].dtype):
#                     df[col] = df[col].astype(str).replace('NaT', '')
            
#             # Create sheet header for text output
#             sheet_text = f"\n=== Sheet: {sheet_name} ===\n"
            
#             # Convert dataframe to readable text format
#             if not df.empty:
#                 # Add column headers
#                 headers = df.columns.tolist()
#                 sheet_text += "Columns: " + " | ".join(str(col) for col in headers) + "\n\n"
                
#                 # Add each row as structured text
#                 for idx, row in df.iterrows():
#                     # Clean row data for JSON serialization
#                     row_dict = {}
#                     row_text_parts = []
                    
#                     for col, val in row.items():
#                         cleaned_val = clean_data_for_json(val)
#                         row_dict[str(col)] = cleaned_val
                        
#                         if cleaned_val and str(cleaned_val).strip():
#                             row_text_parts.append(f"{col}: {cleaned_val}")
                    
#                     row_text = " | ".join(row_text_parts)
                    
#                     if row_text:  # Only add non-empty rows
#                         sheet_text += f"Row {idx + 1}: {row_text}\n"
                        
#                         # Add to structured data (consistent with PDF/DOCX format)
#                         all_extracted_data.append({
#                             "page": int(idx + 1),
#                             "sheet": str(sheet_name),
#                             "source": "excel_row",
#                             "content": row_text,
#                             "structured_data": row_dict  # Keep cleaned structured data
#                         })
                
#                 # Also add CSV representation for raw text
#                 try:
#                     csv_representation = df.to_csv(index=False)
#                     sheet_text += f"\nCSV Format:\n{csv_representation}\n"
#                 except:
#                     sheet_text += "\nCSV Format: Could not generate CSV representation\n"
#             else:
#                 sheet_text += "Sheet is empty\n"
            
#             extracted_text += sheet_text
        
#         # If no data was extracted, create a minimal structure
#         if not all_extracted_data:
#             all_extracted_data = [{
#                 "page": 1,
#                 "source": "excel_empty",
#                 "content": "Excel file appears to be empty or contains no readable data"
#             }]
        
#         # Final cleanup of all extracted data
#         all_extracted_data = clean_data_for_json(all_extracted_data)
        
#         return extracted_text.strip(), all_extracted_data
        
#     except Exception as e:
#         return f"Error extracting content from Excel: {str(e)}", []

def extract_text_from_excel(excel_file):
    """Extract text from Excel file, convert CSV to base64"""
    try:
        excel_file.seek(0)
        
        # Read all sheets from the Excel file
        df_dict = pd.read_excel(excel_file, sheet_name=None)
        
        extracted_text = ""
        all_extracted_data = []
        
        for sheet_name, df in df_dict.items():
            df = df.replace([np.nan, np.inf, -np.inf], "")
            df = df.fillna("")

            for col in df.columns:
                if df[col].dtype == 'datetime64[ns]' or 'datetime' in str(df[col].dtype):
                    df[col] = df[col].astype(str).replace('NaT', '')
            
            sheet_text = f"\n=== Sheet: {sheet_name} ===\n"
            
            if not df.empty:
                headers = df.columns.tolist()
                sheet_text += "Columns: " + " | ".join(str(col) for col in headers) + "\n\n"
                
                for idx, row in df.iterrows():
                    row_dict = {}
                    row_text_parts = []
                    
                    for col, val in row.items():
                        cleaned_val = clean_data_for_json(val)
                        row_dict[str(col)] = cleaned_val
                        
                        if cleaned_val and str(cleaned_val).strip():
                            row_text_parts.append(f"{col}: {cleaned_val}")
                    
                    row_text = " | ".join(row_text_parts)
                    
                    if row_text:
                        sheet_text += f"Row {idx + 1}: {row_text}\n"
                        all_extracted_data.append({
                            "page": int(idx + 1),
                            "sheet": str(sheet_name),
                            "source": "excel_row",
                            "content": row_text,
                            "structured_data": row_dict
                        })
                
                # Generate CSV representation
                csv_representation = df.to_csv(index=False)

                # Encode CSV representation to base64
                encoded_csv = base64.b64encode(csv_representation.encode('utf-8')).decode('utf-8')

                # Include encoded CSV in the extracted data
                all_extracted_data.append({
                    "page": sheet_name,
                    "source": "csv_base64",
                    "content": encoded_csv
                })

                sheet_text += f"\nCSV Format (Base64 encoded):\n{encoded_csv}\n"
            else:
                sheet_text += "Sheet is empty\n"
            
            extracted_text += sheet_text
        
        if not all_extracted_data:
            all_extracted_data = [{
                "page": 1,
                "source": "excel_empty",
                "content": "Excel file appears to be empty or contains no readable data"
            }]
        
        all_extracted_data = clean_data_for_json(all_extracted_data)
        
        return extracted_text.strip(), all_extracted_data
        
    except Exception as e:
        return f"Error extracting content from Excel: {str(e)}", []


def send_raw_to_n8n(raw_data, webhook_url, file_name):
    """
    Send raw extracted data to n8n workflow via webhook
    """
    try:
        headers = {
            "Content-Type": "application/json"
        }
        
        # Get file extension for specialized processing
        file_extension = os.path.splitext(file_name)[1].lower()
        
        payload = {
            "raw_data": raw_data,
            "timestamp": datetime.now().isoformat(),
            "file_name": file_name,
            "file_type": file_extension.replace(".", "")  # Add file type for workflow routing
        }
        
        response = requests.post(webhook_url, json=payload, headers=headers)
        
        if response.status_code == 200:
            return True, response.json(), f"Successfully sent raw data to n8n workflow. Status: {response.status_code}"
        else:
            return False, None, f"Error sending data to n8n workflow. Status: {response.status_code}, Response: {response.text}"
    
    except Exception as e:
        return False, None, f"Exception when sending data to n8n: {str(e)}"

def save_to_excel(new_data, excel_path, filename, current_row=1):
    """Save data to Excel with proper formatting and structure"""
    try:
        # Map JSON keys to desired column names
        renamed_data = []
        for item in new_data:
            renamed_data.append({
                "Process": item.get("Process", ""),
                "Description": item.get("Description", ""),
                "Objective": item.get("Objective", ""),
                "Control Design": item.get("Control Designed", ""),
                "Risks": item.get("Risks", ""),
                "Gaps": item.get("Gaps", ""),
                "Recommendation": item.get("Auditor's Recommendation", "")
            })
        
        df = pd.DataFrame(renamed_data)
        
        # Create workbook if it doesn't exist
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
        else:
            with open(excel_path, "rb") as f:
                wb = load_workbook(f)
            ws = wb.active
        
        # Get file name without extension
        file_name_without_ext = os.path.splitext(filename)[0]
        
        # Set file name as header row
        ws.cell(row=current_row, column=1, value=file_name_without_ext)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        current_row += 1
        
        # Write column headers
        headers = list(df.columns)
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col_index, value=header)
            ws.cell(row=current_row, column=col_index).font = Font(bold=True)
            ws.cell(row=current_row, column=col_index).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        current_row += 1
        
        # Write data
        for row_index, row_data in enumerate(df.values, start=0):
            for col_index, cell_value in enumerate(row_data, start=1):
                ws.cell(row=current_row + row_index, column=col_index, value=cell_value)
        
        # Calculate the next row for the next file
        next_row = current_row + len(df) + 1  # +1 for an empty row between files
        
        # Save the workbook
        with open(excel_path, "wb") as f:
            wb.save(f)
            
        return next_row
    
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        return current_row  # Return the original row on error