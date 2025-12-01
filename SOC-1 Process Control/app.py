# currently running on server
#app.py
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import os
import io
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import tempfile

# Import document processing logic
from document_processor import (
    extract_text_from_image,
    extract_text_from_pdf,
    extract_text_from_pdf_with_pdf2image,
    extract_text_from_docx,
    send_raw_to_n8n,
    save_to_excel
)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload
app.config['SECRET_KEY'] = 'your-secret-key-here'  # Change in production

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Define n8n webhook URLs
WEBHOOK_URLS = {
    '.pdf': "http://192.168.100.85:5678/webhook/83b73abd-95ae-468d-a930-883181f46b78",
    '.docx': "http://192.168.100.85:5678/webhook/a5350f42-6e00-4773-ba83-6c8f00ce31c0",
}

# Excel output path
EXCEL_FILE_PATH = "output.xlsx"

def allowed_file(filename):
    """Check if file has an allowed extension"""
    allowed_extensions = {'.pdf', '.docx'}
    _, extension = os.path.splitext(filename.lower())
    return extension in allowed_extensions

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Process the file
        file_extension = os.path.splitext(filename)[1].lower()
        extracted_data = []
        raw_text = ""
        
        try:
            if file_extension == '.pdf':
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_pdf(f)
                    # If PyMuPDF extraction returns little text, try pdf2image as fallback
                    if len(raw_text.strip()) < 100:  # Threshold for minimal text
                        alt_text, alt_data = extract_text_from_pdf_with_pdf2image(f)
                        if len(alt_text.strip()) > len(raw_text.strip()):
                            raw_text = alt_text
                            extracted_data = alt_data
            
            elif file_extension in ['.png', '.jpg', '.jpeg']:
                from PIL import Image
                image = Image.open(file_path)
                raw_text = extract_text_from_image(image)
                extracted_data = [{"page": 1, "source": "image_ocr", "content": raw_text}]
            
            elif file_extension == '.docx':
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_docx(f)
            
            else:
                return jsonify({'error': 'Unsupported file format'}), 400
            
            # Save processed data to session
            session_data = {
                'filename': filename,
                'raw_text': raw_text,
                'extracted_data': extracted_data,
                'file_path': file_path
            }
            
            # Use a temporary file to store session data
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
            with open(temp_file.name, 'w') as f:
                json.dump(session_data, f)
            
            return jsonify({
                'success': True,
                'filename': filename,
                'raw_text': raw_text,
                'session_id': os.path.basename(temp_file.name)
            })
            
        except Exception as e:
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500
    
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/send_to_n8n', methods=['POST'])
def send_to_n8n():
    data = request.json
    session_id = data.get('session_id')
    
    if not session_id:
        return jsonify({'error': 'No session ID provided'}), 400
    
    try:
        # Load session data
        with open(os.path.join(tempfile.gettempdir(), session_id), 'r') as f:
            session_data = json.load(f)
        
        filename = session_data['filename']
        extracted_data = session_data['extracted_data']
        
        # Get file extension
        file_extension = os.path.splitext(filename)[1].lower()
        
        # Get appropriate webhook URL
        webhook_url = WEBHOOK_URLS.get(file_extension)
        if not webhook_url:
            return jsonify({'error': 'No webhook URL for this file type'}), 400
        
        # Send data to n8n
        success, n8n_response, message = send_raw_to_n8n(extracted_data, webhook_url, filename)
        
        if success:
            # Save n8n response to Excel if it's valid
            excel_saved = False
            if n8n_response:
                try:
                    # Determine the current row in Excel
                    current_row = 1
                    if os.path.exists(EXCEL_FILE_PATH):
                        from openpyxl import load_workbook
                        if os.path.getsize(EXCEL_FILE_PATH) > 0:
                            wb = load_workbook(EXCEL_FILE_PATH)
                            ws = wb.active
                            if ws.max_row > 1:
                                current_row = ws.max_row + 2
                    
                    # Handle different response formats
                    if not isinstance(n8n_response, list):
                        n8n_response = [n8n_response] if isinstance(n8n_response, dict) else []
                    
                    # Save to Excel
                    if n8n_response:
                        next_row = save_to_excel(n8n_response, EXCEL_FILE_PATH, filename, current_row)
                        excel_saved = next_row > current_row
                except Exception as e:
                    return jsonify({
                        'success': True,
                        'n8n_success': True, 
                        'excel_saved': False,
                        'message': message,
                        'error': f'Error saving to Excel: {str(e)}',
                        'n8n_response': n8n_response
                    })
            
            return jsonify({
                'success': True,
                'n8n_success': True,
                'excel_saved': excel_saved,
                'message': message,
                'n8n_response': n8n_response
            })
        else:
            return jsonify({
                'success': True,
                'n8n_success': False,
                'message': message
            })
    
    except Exception as e:
        return jsonify({'error': f'Error sending to n8n: {str(e)}'}), 500

@app.route('/download_text/<session_id>')
def download_text(session_id):
    try:
        # Load session data
        with open(os.path.join(tempfile.gettempdir(), session_id), 'r') as f:
            session_data = json.load(f)
        
        raw_text = session_data['raw_text']
        filename = session_data['filename']
        
        # Create in-memory text file
        text_file = io.BytesIO(raw_text.encode())
        
        return send_file(
            text_file,
            mimetype="text/plain",
            as_attachment=True,
            download_name=f"extracted_text_{os.path.splitext(filename)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
    
    except Exception as e:
        return jsonify({'error': f'Error downloading text: {str(e)}'}), 500

@app.route('/download_excel')
def download_excel():
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            return send_file(
                EXCEL_FILE_PATH,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"extracted_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        except Exception as e:
            return jsonify({'error': f'Error downloading Excel: {str(e)}'}), 500
    else:
        return jsonify({'error': 'Excel file does not exist'}), 404

@app.route('/preview_excel')
def preview_excel():
    import pandas as pd
    
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            df = pd.read_excel(EXCEL_FILE_PATH)
            return jsonify({
                'success': True,
                'data': df.to_dict('records'),
                'columns': df.columns.tolist()
            })
        except Exception as e:
            return jsonify({'error': f'Error previewing Excel: {str(e)}'}), 500
    else:
        return jsonify({'error': 'Excel file does not exist'}), 404

@app.route('/clear_excel', methods=['POST'])
def clear_excel():
    if os.path.exists(EXCEL_FILE_PATH):
        try:
            os.remove(EXCEL_FILE_PATH)
            return jsonify({'success': True, 'message': 'Excel file cleared'})
        except Exception as e:
            return jsonify({'error': f'Error clearing Excel: {str(e)}'}), 500
    else:
        return jsonify({'success': True, 'message': 'Excel file does not exist'})

if __name__ == '__main__':
    print("Starting Flask app...")
    app.run(debug=True)