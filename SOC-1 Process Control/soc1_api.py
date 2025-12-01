"""
SOC-1 PROCESS CONTROL API
Exposes SOC-1 document processing as a REST API for orchestration
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import logging
from datetime import datetime
from werkzeug.utils import secure_filename
import tempfile

# Import SOC-1 modules
from config import Config
from document_processor import (
    extract_text_from_image,
    extract_text_from_pdf,
    extract_text_from_pdf_with_pdf2image,
    extract_text_from_docx,
    extract_text_from_excel,
    clean_data_for_json,
    send_raw_to_n8n,
    save_to_excel
)
from db_operations import doc_ops, session_ops, result_ops, audit_ops

# Initialize Flask
app = Flask(__name__)
app.config.from_object(Config)
CORS(app)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Constants
TOOL_ID = "soc1_process_control"
TOOL_VERSION = "1.0.0"
EXCEL_FILE_PATH = Config.EXCEL_FILE_PATH


def allowed_file(filename):
    """Check if file extension is allowed"""
    _, extension = os.path.splitext(filename.lower())
    return extension in Config.ALLOWED_EXTENSIONS


# ============================================
# ENDPOINT 1: HEALTH CHECK
# ============================================

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "healthy",
        "tool_id": TOOL_ID,
        "version": TOOL_VERSION,
        "timestamp": datetime.now().isoformat()
    }), 200


# ============================================
# ENDPOINT 2: TOOL INFO
# ============================================

@app.route('/api/v1/info', methods=['GET'])
def tool_info():
    """Get tool information and capabilities"""
    return jsonify({
        "tool_id": TOOL_ID,
        "name": "SOC-1 Process Control Analysis",
        "version": TOOL_VERSION,
        "description": "AI-powered document processing for SOC-1 audit controls. Extracts text from documents and generates control analysis reports.",
        "author": "Harshwal Consulting",
        "capabilities": {
            "document_types": list(Config.ALLOWED_EXTENSIONS),
            "max_file_size": "16MB",
            "features": [
                "OCR text extraction",
                "PDF processing",
                "Excel data extraction",
                "Word document processing",
                "AI control analysis",
                "Excel report generation"
            ]
        },
        "parameters": {
            "required": {
                "document_file": {
                    "type": "file",
                    "description": "Document file to process (PDF, DOCX, XLSX, XLS, PNG, JPG)",
                    "accepts": list(Config.ALLOWED_EXTENSIONS)
                }
            },
            "optional": {
                "repeat_count": {
                    "type": "integer",
                    "default": 1,
                    "description": "Number of times to process the document"
                },
                "finalize": {
                    "type": "boolean",
                    "default": False,
                    "description": "Whether to finalize and generate control sheet"
                }
            }
        },
        "endpoints": {
            "execute": "/api/v1/execute",
            "download": "/api/v1/download/<session_id>",
            "info": "/api/v1/info",
            "health": "/health"
        }
    }), 200


# ============================================
# ENDPOINT 3: EXECUTE (Main Processing)
# ============================================

@app.route('/api/v1/execute', methods=['POST'])
def execute():
    """
    Main execution endpoint - processes documents and generates analysis
    """
    logger.info("="*70)
    logger.info("SOC-1 PROCESS CONTROL: Starting execution")
    logger.info("="*70)
    
    try:
        # Get file from request
        if 'document_file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No document_file provided in request',
                'error_type': 'missing_parameter'
            }), 400
        
        file = request.files['document_file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'Empty filename',
                'error_type': 'invalid_file'
            }), 400
        
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': f'Unsupported file type. Allowed: {list(Config.ALLOWED_EXTENSIONS)}',
                'error_type': 'invalid_file_type'
            }), 400
        
        # Get optional parameters
        repeat_count = int(request.form.get('repeat_count', 1))
        finalize = request.form.get('finalize', 'false').lower() == 'true'
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)
        
        file_size = os.path.getsize(file_path)
        file_extension = os.path.splitext(filename)[1].lower()
        
        logger.info(f"File saved: {filename} ({file_size} bytes)")
        
        # ============================================
        # STEP 1: EXTRACT TEXT FROM DOCUMENT
        # ============================================
        
        extracted_data = []
        raw_text = ""
        
        try:
            if file_extension == '.pdf':
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_pdf(f)
                    if len(raw_text.strip()) < 100:
                        alt_text, alt_data = extract_text_from_pdf_with_pdf2image(f)
                        if len(alt_text.strip()) > len(raw_text.strip()):
                            raw_text = alt_text
                            extracted_data = alt_data
            
            elif file_extension == '.docx':
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_docx(f)
            
            elif file_extension in ['.xlsx', '.xls']:
                with open(file_path, 'rb') as f:
                    raw_text, extracted_data = extract_text_from_excel(f)
            
            elif file_extension in ['.png', '.jpg', '.jpeg']:
                from PIL import Image
                image = Image.open(file_path)
                raw_text = extract_text_from_image(image)
                extracted_data = [{"page": 1, "source": "image_ocr", "content": raw_text}]
            
            else:
                raise ValueError(f"Unsupported file extension: {file_extension}")
            
            logger.info(f"Text extracted: {len(raw_text)} characters")
            
        except Exception as e:
            logger.error(f"Error extracting text: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to extract text: {str(e)}',
                'error_type': 'extraction_error'
            }), 500
        
        # Clean data
        cleaned_extracted_data = clean_data_for_json(extracted_data)
        cleaned_raw_text = clean_data_for_json(raw_text)
        
        # ============================================
        # STEP 2: SAVE TO MONGODB
        # ============================================
        
        document_id = doc_ops.create_document(
            filename=filename,
            file_type=file_extension,
            file_size=file_size,
            raw_text=cleaned_raw_text,
            extracted_data=cleaned_extracted_data
        )
        
        session_id = f"session_{timestamp}_{os.urandom(4).hex()}"
        
        if document_id:
            session_ops.create_session(
                session_id=session_id,
                document_id=document_id,
                filename=filename,
                file_path=file_path
            )
            logger.info(f"✓ Document saved to MongoDB: {document_id}")
        
        # ============================================
        # STEP 3: SEND TO N8N FOR PROCESSING
        # ============================================
        
        webhook_url = Config.WEBHOOK_URLS.get(file_extension)
        if not webhook_url:
            return jsonify({
                'success': False,
                'error': f'No webhook configured for {file_extension}',
                'error_type': 'configuration_error'
            }), 500
        
        session_ops.update_repeat_count(session_id, repeat_count)
        session_ops.add_processing_step(
            session_id, "n8n_processing", "started",
            {"repeat_count": repeat_count}
        )
        
        # Initialize Excel file
        current_row = 1
        if not os.path.exists(EXCEL_FILE_PATH):
            from openpyxl import Workbook
            wb = Workbook()
            wb.save(EXCEL_FILE_PATH)
        else:
            from openpyxl import load_workbook
            try:
                wb = load_workbook(EXCEL_FILE_PATH)
                ws = wb.active
                if ws.max_row > 1:
                    current_row = ws.max_row + 2
                wb.close()
            except:
                pass
        
        all_results = []
        
        # Process with n8n (repeat if needed)
        for i in range(repeat_count):
            try:
                success, n8n_response, message = send_raw_to_n8n(
                    cleaned_extracted_data, webhook_url, f"{filename}_part{i+1}"
                )
                
                if not success:
                    session_ops.add_processing_step(
                        session_id, f"n8n_iteration_{i+1}", "failed",
                        {"error": message}
                    )
                    return jsonify({
                        'success': False,
                        'error': f'N8N processing failed: {message}',
                        'error_type': 'n8n_error'
                    }), 500
                
                if not isinstance(n8n_response, list):
                    n8n_response = [n8n_response] if isinstance(n8n_response, dict) else []
                
                if n8n_response and document_id:
                    # Save to MongoDB
                    result_ids = result_ops.bulk_create_results(
                        document_id=str(document_id),
                        session_id=session_id,
                        results_data=n8n_response
                    )
                    all_results.extend(result_ids)
                    
                    # Save to Excel
                    current_row = save_to_excel(
                        n8n_response, EXCEL_FILE_PATH,
                        f"{filename}_part{i+1}", current_row
                    )
                    
                    session_ops.add_processing_step(
                        session_id, f"n8n_iteration_{i+1}", "completed",
                        {"result_count": len(n8n_response)}
                    )
                
            except Exception as e:
                logger.error(f"Error in iteration {i+1}: {e}")
                return jsonify({
                    'success': False,
                    'error': f'Processing error: {str(e)}',
                    'error_type': 'processing_error'
                }), 500
        
        # Mark results as exported
        if all_results:
            result_ops.mark_as_excel_exported(all_results)
        
        # ============================================
        # STEP 4: FINALIZE (if requested)
        # ============================================
        
        control_data = None
        if finalize:
            try:
                import pandas as pd
                import requests
                from openpyxl import load_workbook
                from openpyxl.styles import Font
                
                df = pd.read_excel(EXCEL_FILE_PATH)
                structured_data = df.fillna("").to_dict(orient="records")
                
                control_response = requests.post(
                    Config.CONTROL_WEBHOOK_URL,
                    json={"rows": structured_data}
                )
                control_response.raise_for_status()
                control_data = control_response.json()
                
                # Write control sheet
                wb = load_workbook(EXCEL_FILE_PATH)
                if 'Control' in wb.sheetnames:
                    wb.remove(wb['Control'])
                ws = wb.create_sheet('Control')
                
                if isinstance(control_data, list) and control_data:
                    headers = list(control_data[0].keys())
                    for col_index, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_index, value=header)
                        ws.cell(row=1, column=col_index).font = Font(bold=True, size=14)
                    for row_index, row in enumerate(control_data, 2):
                        for col_index, header in enumerate(headers, 1):
                            ws.cell(row=row_index, column=col_index, value=row.get(header))
                
                wb.save(EXCEL_FILE_PATH)
                logger.info("✓ Control sheet created")
                
            except Exception as e:
                logger.error(f"Finalize error: {e}")
        
        # Update session status
        session_ops.update_session_status(session_id, "completed")
        
        # Log completion
        audit_ops.log_action(
            "soc1_processing_completed",
            details={
                "session_id": session_id,
                "filename": filename,
                "iterations": repeat_count,
                "results_count": len(all_results),
                "finalized": finalize
            }
        )
        
        # ============================================
        # STEP 5: PREPARE RESPONSE
        # ============================================
        
        response = {
            'success': True,
            'session_id': session_id,
            'document_id': str(document_id) if document_id else None,
            'filename': filename,
            'processing': {
                'iterations': repeat_count,
                'results_count': len(all_results),
                'text_extracted': len(raw_text),
                'finalized': finalize
            },
            'downloads': {
                'excel': f'/api/v1/download/{session_id}/excel',
                'text': f'/api/v1/download/{session_id}/text'
            }
        }
        
        if control_data:
            response['control_data'] = control_data
        
        logger.info("✓ Processing completed successfully")
        logger.info("="*70)
        
        # Cleanup uploaded file
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except:
            pass
        
        return jsonify(response), 200
        
    except Exception as e:
        logger.error(f"Unexpected error: {e}", exc_info=True)
        audit_ops.log_action(
            "soc1_processing_failed",
            details={"error": str(e)},
            status="error"
        )
        return jsonify({
            'success': False,
            'error': f'Unexpected error: {str(e)}',
            'error_type': 'internal_error'
        }), 500


# ============================================
# ENDPOINT 4: DOWNLOAD RESULTS
# ============================================

@app.route('/api/v1/download/<session_id>/<file_type>', methods=['GET'])
def download_result(session_id, file_type):
    """Download processed results"""
    try:
        if file_type == 'excel':
            if os.path.exists(EXCEL_FILE_PATH):
                return send_file(
                    EXCEL_FILE_PATH,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=f"soc1_results_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
            else:
                return jsonify({
                    'success': False,
                    'error': 'Excel file not found'
                }), 404
        
        elif file_type == 'text':
            session = session_ops.get_session(session_id)
            if not session:
                return jsonify({
                    'success': False,
                    'error': 'Session not found'
                }), 404
            
            document = doc_ops.get_document(str(session['document_id']))
            if not document:
                return jsonify({
                    'success': False,
                    'error': 'Document not found'
                }), 404
            
            import io
            text_file = io.BytesIO(document['raw_text'].encode())
            
            return send_file(
                text_file,
                mimetype="text/plain",
                as_attachment=True,
                download_name=f"extracted_text_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
        
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid file type. Use "excel" or "text"'
            }), 400
    
    except Exception as e:
        logger.error(f"Download error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================
# MAIN
# ============================================

if __name__ == '__main__':
    logger.info("="*70)
    logger.info("SOC-1 PROCESS CONTROL API - Starting")
    logger.info("="*70)
    logger.info("Port: 5007")
    logger.info("Endpoints:")
    logger.info("  GET  /health")
    logger.info("  GET  /api/v1/info")
    logger.info("  POST /api/v1/execute")
    logger.info("  GET  /api/v1/download/<session_id>/<file_type>")
    logger.info("="*70)
    
    app.run(host='0.0.0.0', port=5007, debug=True, use_reloader=False)